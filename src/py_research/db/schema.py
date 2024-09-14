"""Static schemas for universal relational databases."""

from __future__ import annotations

from collections.abc import Callable, Hashable, Iterable, Mapping, Sequence
from dataclasses import asdict, dataclass, field, fields
from datetime import date, datetime, time
from functools import cached_property, partial, reduce
from inspect import getmodule
from io import BytesIO
from pathlib import Path
from secrets import token_hex
from types import ModuleType
from typing import (
    TYPE_CHECKING,
    Any,
    ClassVar,
    Generic,
    Literal,
    Self,
    cast,
    dataclass_transform,
    get_args,
    get_origin,
    overload,
)
from uuid import UUID, uuid4

import openpyxl
import pandas as pd
import polars as pl
import sqlalchemy as sqla
import sqlalchemy.dialects.mysql as mysql
import sqlalchemy.dialects.postgresql as postgresql
import sqlalchemy.orm as orm
import sqlalchemy.sql.type_api as sqla_types
import sqlalchemy.sql.visitors as sqla_visitors
import yarl
from bidict import bidict
from cloudpathlib import CloudPath
from pandas.api.types import (
    is_bool_dtype,
    is_datetime64_dtype,
    is_integer_dtype,
    is_numeric_dtype,
    is_string_dtype,
)
from sqlalchemy_utils import UUIDType
from typing_extensions import TypeVar, TypeVarTuple
from xlsxwriter import Workbook as ExcelWorkbook

from py_research.files import HttpFile
from py_research.hashing import gen_int_hash, gen_str_hash
from py_research.reflect.ref import PyObjectRef
from py_research.reflect.runtime import get_subclasses
from py_research.reflect.types import has_type

DataFrame = pd.DataFrame | pl.DataFrame
Series = pd.Series | pl.Series

B_def = TypeVar("B_def", bound="Backend", default="Backend", covariant=True)
B_nul = TypeVar(
    "B_nul", bound="Backend | None", default="Backend | None", covariant=True
)
B2_opt = TypeVar("B2_opt", bound="Backend | None", default=None)

Key = TypeVar("Key", bound=Hashable)
Key2 = TypeVar("Key2", bound=Hashable)
Key3 = TypeVar("Key3", bound=Hashable)
Key4 = TypeVar("Key4", bound=Hashable)
Key_def = TypeVar("Key_def", contravariant=True, bound=Hashable, default=Any)

type Index = Hashable | BaseIdx | SingleIdx | FilteredIdx
Idx = TypeVar("Idx", bound=Index)
Idx_cov = TypeVar("Idx_cov", covariant=True, bound=Index)
Idx_def = TypeVar(
    "Idx_def",
    covariant=True,
    bound=Index,
    default=Index,
)

type KeyIdx = Hashable | SingleIdx
KeyIdx_def = TypeVar("KeyIdx_def", bound=KeyIdx, default=KeyIdx, covariant=True)

type InsertIdx = Hashable | BaseIdx
InsIdx = TypeVar("InsIdx", bound=InsertIdx, covariant=True)

IdxTup = TypeVarTuple("IdxTup")

type IdxStart[Key: Hashable] = Key | tuple[Key, *tuple[Any, ...]]
type IdxEnd[Key: Hashable] = tuple[*tuple[Any, ...], Key]
type IdxStartEnd[Key: Hashable, Key2: Hashable] = tuple[Key, *tuple[Any, ...], Key2]
type IdxTupStartEnd[*IdxTup, Key2: Hashable] = tuple[*IdxTup, *tuple[Any], Key2]

Val = TypeVar("Val")
Val2 = TypeVar("Val2")
Val3 = TypeVar("Val3")
Val_cov = TypeVar("Val_cov", covariant=True)
Val_def = TypeVar("Val_def", covariant=True, default=Any)
Val_defi = TypeVar("Val_defi", default=Any)

PVal = TypeVar("PVal", bound="Prop")

Rec = TypeVar("Rec", bound="Record")
Rec2 = TypeVar("Rec2", bound="Record")
Rec3 = TypeVar("Rec3", bound="Record")
Rec_cov = TypeVar("Rec_cov", covariant=True, bound="Record")
Rec_def = TypeVar("Rec_def", covariant=True, bound="Record", default="Record")
Rec2_def = TypeVar("Rec2_def", covariant=True, bound="Record", default="Record")
Rec3_def = TypeVar("Rec3_def", covariant=True, bound="Record", default="Record")
Rec2_nul = TypeVar(
    "Rec2_nul", bound="Record | None", default="Record | None", covariant=True
)

type RecordValue[Rec: Record] = Rec | Iterable[Rec] | Mapping[Any, Rec] | None
Recs_cov = TypeVar("Recs_cov", bound=RecordValue, covariant=True)

RM = TypeVar("RM", covariant=True, bound=tuple | None, default=tuple | None)

RelTup = TypeVarTuple("RelTup")

P = TypeVar("P", bound="Prop")
P_nul = TypeVar(
    "P_nul", bound="Prop[Any, R] | None", covariant=True, default="Prop[Any, R] | None"
)

R_def = TypeVar("R_def", bound="R", default="R", covariant=True)
RWT = TypeVar("RWT", bound="R", default="RW", covariant=True)

Df = TypeVar("Df", bound=DataFrame)
Dl = TypeVar("Dl", bound="DataFrame | Record")


class BaseIdx:
    """Singleton to mark dataset index as default index."""

    __hash__: ClassVar[None]  # type: ignore[assignment]


class SingleIdx:
    """Singleton to mark dataset index as a single value."""

    __hash__: ClassVar[None]  # type: ignore[assignment]


class FilteredIdx(Generic[Idx_cov]):
    """Singleton to mark dataset index as filtered."""

    __hash__: ClassVar[None]  # type: ignore[assignment]


@dataclass(frozen=True)
class Backend:
    """SQL backend for DB."""

    name: str | None = None
    """Unique name to identify this backend by."""

    url: sqla.URL | CloudPath | HttpFile | Path | None = None
    """Connection URL or path."""

    @cached_property
    def type(
        self,
    ) -> Literal["sql-connection", "sqlite-file", "excel-file", "in-memory"]:
        """Type of the backend."""
        match self.url:
            case Path() | CloudPath():
                return "excel-file" if "xls" in self.url.suffix else "sqlite-file"
            case sqla.URL():
                return (
                    "sqlite-file"
                    if self.url.drivername == "sqlite"
                    else "sql-connection"
                )
            case HttpFile():
                url = yarl.URL(self.url.url)
                typ = (
                    ("excel-file" if "xls" in Path(url.path).suffix else "sqlite-file")
                    if url.scheme in ("http", "https")
                    else None
                )
                if typ is None:
                    raise ValueError(f"Unsupported URL scheme: {url.scheme}")
                return typ
            case None:
                return "in-memory"


class R:
    """Read-only flag."""


class RO(R):
    """Read-only flag."""


class RW(RO):
    """Read-write flag."""


type RecInput[Rec: Record, Key: InsertIdx] = DataFrame | Iterable[Rec] | Mapping[
    Key, Rec
] | sqla.Select | Rec

type PartialRec[Rec: Record] = Mapping[Set[Rec, Any], Any]

type PartialRecInput[Rec: Record, Key: InsertIdx] = RecInput[Rec, Any] | Iterable[
    PartialRec[Rec]
] | Mapping[Key, PartialRec[Rec]] | PartialRec[Rec]

type ValInput[Val, Key: Hashable] = Series | Mapping[Key, Val] | sqla.Select[
    tuple[Key, Val]
] | Val


def map_df_dtype(c: pd.Series | pl.Series) -> type | None:
    """Map pandas dtype to Python type."""
    if isinstance(c, pd.Series):
        if is_datetime64_dtype(c):
            return datetime
        elif is_bool_dtype(c):
            return bool
        elif is_integer_dtype(c):
            return int
        elif is_numeric_dtype(c):
            return float
        elif is_string_dtype(c):
            return str
    else:
        if c.dtype.is_temporal():
            return datetime
        elif c.dtype.is_integer():
            return int
        elif c.dtype.is_float():
            return float
        elif c.dtype.is_(pl.String):
            return str

    return None


def map_col_dtype(c: sqla.ColumnElement) -> type | None:
    """Map sqla column type to Python type."""
    match c.type:
        case sqla.DateTime():
            return datetime
        case sqla.Date():
            return date
        case sqla.Time():
            return time
        case sqla.Boolean():
            return bool
        case sqla.Integer():
            return int
        case sqla.Float():
            return float
        case sqla.String() | sqla.Text() | sqla.Enum():
            return str
        case sqla.LargeBinary():
            return bytes
        case _:
            return None


def props_from_data(
    data: DataFrame | sqla.Select, foreign_keys: Mapping[str, ValueSet] | None = None
) -> list[Prop]:
    """Extract prop definitions from dataframe or query."""
    if isinstance(data, pd.DataFrame) and len(data.index.names) > 1:
        raise NotImplementedError("Multi-index not supported yet.")

    foreign_keys = foreign_keys or {}

    def _gen_prop(name: str, data: Series | sqla.ColumnElement) -> Prop:
        is_rel = name in foreign_keys
        value_type = (
            map_df_dtype(data) if isinstance(data, Series) else map_col_dtype(data)
        ) or Any
        attr = ValueSet(
            prop=Attr(primary_key=True, _name=name if not is_rel else f"fk_{name}"),
            typedef=PropType(Attr[value_type]),
            record_type=DynRecord,
        )
        return attr.prop if not is_rel else Rel(on={attr: foreign_keys[name]})

    columns = (
        [data[col] for col in data.columns]
        if isinstance(data, DataFrame)
        else list(data.columns)
    )

    return [
        *(
            (
                _gen_prop(level, data.index.get_level_values(level).to_series())
                for level in data.index.names
            )
            if isinstance(data, pd.DataFrame)
            else []
        ),
        *(_gen_prop(str(col.name), col) for col in columns),
    ]


def _remove_external_fk(table: sqla.Table):
    """Dirty vodoo to remove external FKs from existing table."""
    for c in table.columns.values():
        c.foreign_keys = set(
            fk
            for fk in c.foreign_keys
            if fk.constraint and fk.constraint.referred_table.schema == table.schema
        )

    table.foreign_keys = set(  # type: ignore
        fk
        for fk in table.foreign_keys
        if fk.constraint and fk.constraint.referred_table.schema == table.schema
    )

    table.constraints = set(
        c
        for c in table.constraints
        if not isinstance(c, sqla.ForeignKeyConstraint)
        or c.referred_table.schema == table.schema
    )


class Unloaded:
    """Singleton to mark unfetched values."""


@dataclass(frozen=True)
class PropType(Generic[P_nul]):
    """Reference to a type."""

    hint: str | type[P_nul] | None = None
    ctx: ModuleType | None = None

    def prop_type(self) -> type[Prop]:
        """Resolve the property type reference."""
        hint = self.hint or Prop

        if isinstance(hint, type):
            base = get_origin(hint)
            assert base is not None and issubclass(base, Prop)
            return base
        else:
            return Attr if "Attr" in hint else Rel if "Rel" in hint else Prop

    def value_type(self: PropType[Prop[Val, Any]]) -> type[Val]:
        """Resolve the value type reference."""
        hint = self.hint or Prop

        generic = (
            cast(type[P_nul], eval(hint, vars(self.ctx) if self.ctx else None))
            if isinstance(hint, str)
            else hint
        )

        assert issubclass(get_origin(generic) or generic, Prop)

        args = get_args(generic)
        return cast(type[Val], args[0] if len(args) > 0 else object)


@dataclass(eq=False)
class Prop(Generic[Val_cov, RWT]):
    """Reference a property of a record."""

    alias: str | None = None
    default: Val_cov | None = None
    default_factory: Callable[[], Val_cov] | None = None

    getter: Callable[[Record], Val_cov] | None = None
    setter: Callable[[Record, Val_cov], None] | None = None

    _name: str | None = None

    @property
    def name(self) -> str:
        """Property name."""
        if self.alias is not None:
            return self.alias
        elif self._name is not None:
            return self._name
        else:
            raise ValueError("Property name not set.")

    def __set_name__(self, _, name: str) -> None:  # noqa: D105
        self._name = name

    def __hash__(self) -> int:
        """Hash the Prop."""
        return gen_int_hash(self)

    @overload
    def __get__(self, instance: Record, owner: type[Record]) -> Val_cov: ...

    @overload
    def __get__(
        self, instance: None, owner: type[Rec]
    ) -> Set[Val_cov, Rec, None, Any, Prop[Val_cov, RWT]]: ...

    @overload
    def __get__(self, instance: object | None, owner: type | None) -> Self: ...

    def __get__(  # noqa: D105
        self, instance: object | None, owner: type | type[Rec] | None
    ) -> Val_cov | Set[Val_cov, Rec, None, Any, Prop[Val_cov, RWT]] | Self:
        if isinstance(instance, Record):
            if self.getter is not None:
                value = self.getter(instance)
            else:
                value = instance.__dict__[self.name]

            if isinstance(value, Unloaded) and instance._loader is not None:
                try:
                    value = instance._loader(getattr(owner, self.name), instance._index)
                except KeyError:
                    pass

            if isinstance(value, Unloaded):
                if self.default_factory is not None:
                    value = self.default_factory()
                elif self.default is not None:
                    value = self.default

            if isinstance(value, Unloaded):
                raise ValueError("Property value could not fetched.")

            instance.__dict__[self.name] = value
            return value
        elif owner is not None and issubclass(owner, Record):
            t = (
                ValueSet
                if isinstance(self, Attr)
                else RecordSet if isinstance(self, Rel) else Set
            )
            return t(
                prop=type(self)(
                    **{f.name: getattr(self, f.name) for f in fields(self)}
                ),  # type: ignore
                record_type=cast(type[Rec], owner),
                type=owner._prop_defs[self.name],  # type: ignore
            )
        return self

    def __set__(self: Prop[Val, RW], instance: Record, value: Val | Unloaded) -> None:
        """Set the value of the property."""
        if self.setter is not None and not isinstance(value, Unloaded):
            self.setter(instance, value)
        instance.__dict__[self.name] = value


@dataclass(eq=False)
class Attr(Prop[Val_cov, RWT]):
    """Define an attribute of a record."""

    index: bool = False
    primary_key: bool = False

    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement] | None = None

    @overload
    def __get__(self, instance: Record, owner: type[Record]) -> Val_cov: ...

    @overload
    def __get__(
        self, instance: None, owner: type[Rec]
    ) -> ValueSet[Val_cov, Rec, None, Any, RWT]: ...

    @overload
    def __get__(self, instance: object | None, owner: type | None) -> Self: ...

    def __get__(  # noqa: D105 # type: ignore
        self, instance: object | None, owner: type | type[Rec] | None
    ) -> Val_cov | Set[Val_cov, Rec, None, Any, Prop[Val_cov, RWT]] | Self:
        return super().__get__(instance, owner)


@dataclass(eq=False)
class Rel(Prop[Recs_cov, RWT], Generic[Recs_cov, Rec2_def, RWT, Rec_def]):
    """Define a relation to another record."""

    index: bool = False
    primary_key: bool = False

    on: (
        ValueSet
        | Iterable[ValueSet]
        | dict[ValueSet, ValueSet]
        | RelSet[Any, Any, None, Any, Rec_def]
        | RelSet[Rec_def, Any, None, Any, Rec2_def]
        | tuple[
            RelSet[Any, Any, None, Any, Rec2_def],
            RelSet[Rec_def, Any, None, Any, Rec2_def],
        ]
        | type[Rec_def]
        | type[Rec2_def]
        | None
    ) = None
    order_by: Mapping[ValueSet, int] | None = None
    map_by: ValueSet | None = None
    collection: Callable[[Any], Recs_cov] | None = None

    @property
    def dyn_target_type(self) -> type[Rec_def] | None:
        """Dynamic target type of the relation."""
        match self.on:
            case dict():
                return cast(type[Rec_def], next(iter(self.on.values())).record_type)
            case tuple():
                via_1 = self.on[1]
                assert isinstance(via_1, RecordSet)
                return via_1.target_type
            case type() | RecordSet() | ValueSet() | Iterable() | None:
                return None

    @overload
    def __get__(self, instance: Record, owner: type[Record]) -> Recs_cov: ...

    @overload
    def __get__(
        self: Rel[Mapping[Key, Rec2], Rec3, RWT, Rec2],
        instance: None,
        owner: type[Rec],
    ) -> RelSet[Rec2, Key, None, RWT, Rec, Rec3]: ...

    @overload
    def __get__(
        self: Rel[Sequence[Rec2], Rec3, RWT, Rec2],
        instance: None,
        owner: type[Rec],
    ) -> RelSet[Rec2, int, None, RWT, Rec, Rec3]: ...

    @overload
    def __get__(
        self: Rel[Iterable[Rec2], Rec3, RWT, Record[Key2]],
        instance: None,
        owner: type[Rec],
    ) -> RelSet[Rec2, Key2, None, RWT, Rec, Rec3]: ...

    @overload
    def __get__(
        self: Rel[Rec2, Rec3, RWT, Rec2],
        instance: None,
        owner: type[Rec],
    ) -> RelSet[Rec2, SingleIdx, None, RWT, Rec, Rec3]: ...

    @overload
    def __get__(
        self: Rel[Rec2 | None, Rec3, RWT, Rec2],
        instance: None,
        owner: type[Rec],
    ) -> RelSet[Rec2, SingleIdx | None, None, RWT, Rec, Rec3]: ...

    @overload
    def __get__(self, instance: object | None, owner: type | None) -> Self: ...

    def __get__(  # noqa: D105 # type: ignore
        self, instance: object | None, owner: type | type[Rec2] | None
    ) -> Recs_cov | Set[Recs_cov, Rec2, None, Any, Prop[Recs_cov, RWT]] | Self:
        return super().__get__(instance, owner)


type DirectLink[Rec: Record] = (
    ValueSet | Iterable[ValueSet] | dict[ValueSet, ValueSet[Rec, Any]]
)

type BackLink[Rec: Record] = (RecordSet[Any, Any, Any, Any, Rec] | type[Rec])

type BiLink[Rec: Record, Rec2: Record] = (
    RecordSet[Rec, Any, Any, Any, Rec2]
    | tuple[RecordSet[Any, Any, Any, Any, Rec2], RecordSet[Rec, Any, Any, Any, Rec2]]
    | type[Rec2]
)


@overload
def prop(
    *,
    default: Val | Rec | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: bool = ...,
    primary_key: bool = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: Literal[True],
) -> Any: ...


@overload
def prop(
    *,
    default: Val | Rec | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: bool = ...,
    primary_key: bool = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
) -> Attr[Val]: ...


@overload
def prop(
    *,
    default: Val | Rec | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: bool | Literal["fk"] = ...,
    primary_key: bool | Literal["fk"] = ...,
    collection: None = ...,
    link_on: DirectLink[Rec],
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
) -> Rel[Rec]: ...


@overload
def prop(
    *,
    default: Val | Rec | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal["fk"],
    primary_key: bool | Literal["fk"] = ...,
    collection: None = ...,
    link_on: DirectLink[Rec] | None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
) -> Rel[Rec]: ...


@overload
def prop(
    *,
    default: Val | Rec | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: bool | Literal["fk"] = ...,
    primary_key: Literal["fk"],
    collection: None = ...,
    link_on: DirectLink[Rec] | None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
) -> Rel[Rec]: ...


@overload
def prop(
    *,
    default: Val | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], Val],
    setter: None = ...,
    sql_getter: None = ...,
) -> Prop[Val, RO]: ...


@overload
def prop(
    *,
    default: Val | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], Val],
    setter: Callable[[Record, Val], None],
    sql_getter: None = ...,
) -> Prop[Val, RW]: ...


@overload
def prop(
    *,
    default: Val | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Val2] | None = ...,
    alias: str | None = ...,
    index: bool = ...,
    primary_key: Literal[False] = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], Val],
    setter: None = ...,
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement],
) -> Attr[Val, RO]: ...


@overload
def prop(
    *,
    default: Val | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Val2] | None = ...,
    alias: str | None = ...,
    index: bool = ...,
    primary_key: Literal[False] = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], Val],
    setter: Callable[[Record, Val], None],
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement],
) -> Attr[Val, RW]: ...


@overload
def prop(
    *,
    default: Val | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal[True] = ...,
    primary_key: Literal[False] = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], Val],
    setter: None = ...,
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement] | None = ...,
) -> Attr[Val, RO]: ...


@overload
def prop(
    *,
    default: Val | None = ...,  # type: ignore
    default_factory: Callable[[], Val | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal[True] = ...,
    primary_key: Literal[False] = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], Val],
    setter: Callable[[Record, Val], None],
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement] | None = ...,
) -> Attr[Val, RW]: ...


@overload
def prop(
    *,
    default: Rec | None = ...,  # type: ignore
    default_factory: Callable[[], Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: BackLink[Rec] | None = ...,
    link_via: BiLink[Rec, Rec2] | None = ...,  # type: ignore
    order_by: Mapping[ValueSet[Rec | Rec2, Any], int] | None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
) -> Rel[list[Rec], Rec2]: ...


@overload
def prop(
    *,
    default: Rec | None = ...,
    default_factory: Callable[[], Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    collection: Callable[[Iterable[Rec]], Recs_cov],
    link_on: None = ...,
    link_from: BackLink[Rec] | None = ...,
    link_via: BiLink[Rec, Rec2] | None = ...,  # type: ignore
    order_by: Mapping[ValueSet[Rec | Rec2, Any], int] | None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
) -> Rel[Recs_cov, Rec2]: ...


@overload
def prop(
    *,
    default: Rec | None = ...,
    default_factory: Callable[[], Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    collection: None = ...,
    link_on: None = ...,
    link_from: BackLink[Rec] | None = ...,
    link_via: BiLink[Rec, Rec2] | None = ...,
    order_by: None = ...,
    map_by: ValueSet[Rec | Rec2, Val3],
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
) -> Rel[dict[Val3, Rec], Rec2]: ...


@overload
def prop(
    *,
    default: Rec | None = ...,
    default_factory: Callable[[], Rec | Val2] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    collection: Callable[[Mapping[Val3, Rec]], Recs_cov],
    link_on: None = ...,
    link_from: BackLink[Rec] | None = ...,
    link_via: BiLink[Rec, Rec2] | None = ...,
    order_by: None = ...,
    map_by: ValueSet[Rec | Rec2, Val3],
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
) -> Rel[Recs_cov, Rec2]: ...


def prop(
    *,
    default: Any | None = None,
    default_factory: Callable[[], Any] | None = None,
    alias: str | None = None,
    index: bool | Literal["fk"] = False,
    primary_key: bool | Literal["fk"] = False,
    collection: Callable[[Any], Any] | None = None,
    link_on: DirectLink[Record] | None = None,
    link_from: BackLink[Record] | None = None,
    link_via: BiLink[Record, Record] | None = None,
    order_by: Mapping[ValueSet[Record, Any], int] | None = None,
    map_by: ValueSet[Record, Any] | None = None,
    getter: Callable[[Record], Any] | None = None,
    setter: Callable[[Record, Any], None] | None = None,
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement] | None = None,
    local: bool = False,
) -> Any:
    """Define a backlinking relation to another record."""
    if local:
        return Prop(
            default=default,
            default_factory=default_factory,
            alias=alias,
        )

    if any(a is not None for a in (link_on, link_from, link_via)) or any(
        a == "fk" for a in (index, primary_key)
    ):
        return Rel(
            default=default,
            default_factory=default_factory,
            alias=alias,
            index=index == "fk",
            primary_key=primary_key == "fk",
            on=(
                link_on
                if link_on is not None
                else link_from if link_from is not None else link_via
            ),  # type: ignore
            order_by=order_by,
            map_by=map_by,
            collection=collection,
        )

    if getter is not None and not (primary_key or index or sql_getter is not None):
        return Prop(
            default=default,
            default_factory=default_factory,
            alias=alias,
            getter=getter,
            setter=setter,
        )

    return Attr(
        default=default,
        default_factory=default_factory,
        alias=alias,
        index=index is not False,
        primary_key=primary_key is not False,
        getter=getter,
        setter=setter,
        sql_getter=sql_getter,
    )


@dataclass_transform(kw_only_default=True, field_specifiers=(prop,))
class RecordMeta(type):
    """Metaclass for record types."""

    def __new__(cls, name, bases, namespace, **_):
        """Create a new record type."""
        return super().__new__(cls, name, bases, namespace)

    @property
    def _prop_defs(cls) -> dict[str, PropType]:
        return {
            name: PropType(hint, ctx=getmodule(cls))
            for name, hint in cls.__annotations__.items()
            if issubclass(get_origin(hint) or type, Prop)
            or isinstance(hint, str)
            and ("Attr" in hint or "Rel" in hint)
        }

    def __init__(cls, name, bases, dct):
        """Initialize a new record type."""
        super().__init__(name, bases, dct)

        for name, type_ in cls._prop_defs.items():
            if name not in cls.__dict__:
                setattr(cls, name, type_.prop_type()(_name=name))

    @property
    def _record_bases(cls) -> list[type[Record]]:
        """Get all direct record superclasses of this class."""
        return [
            c
            for c in cls.__bases__
            if issubclass(c, Record) and c is not Record and c is not cls
        ]

    @property
    def _base_pks(cls: type[Record]) -> dict[type[Record], dict[ValueSet, ValueSet]]:
        return {
            base: {
                ValueSet(
                    prop=Attr(
                        _name=pk.name,
                        primary_key=True,
                    ),
                    record_type=cls,
                    typedef=pk.typedef,
                ): pk
                for pk in base._primary_keys.values()
            }
            for base in cls._record_bases
        }

    @property
    def _static_props(cls) -> dict[str, Set]:
        return {name: getattr(cls, name) for name in cls._prop_defs.keys()}

    @property
    def _dynamic_props(cls) -> dict[str, Set]:
        return {
            name: getattr(cls, name)
            for name, prop in cls.__dict__.items()
            if isinstance(prop, Prop)
        }

    @property
    def _defined_props(cls) -> dict[str, Set]:
        return {
            **{
                name: prop
                for name, prop in [
                    *cls._static_props.items(),
                    *cls._dynamic_props.items(),
                ]
            },
        }

    @property
    def _defined_attrs(cls) -> dict[str, ValueSet]:
        return {
            name: getattr(cls, name)
            for name, attr in cls._defined_props.items()
            if isinstance(attr, ValueSet)
        }

    @property
    def _defined_rels(cls) -> dict[str, RelSet]:
        return {
            name: getattr(cls, name)
            for name, rel in cls._defined_props.items()
            if isinstance(rel, RecordSet)
        }

    @property
    def _defined_rel_attrs(cls) -> dict[str, ValueSet]:
        return {
            a.name: a for rel in cls._defined_rels.values() for a in rel.fk_map.keys()
        }

    @property
    def _primary_keys(cls) -> dict[str, ValueSet]:
        """Return the defined primary key attributes of this record."""
        base_pks = {v.name: v for vs in cls._base_pks.values() for v in vs.keys()}

        if len(base_pks) > 0:
            assert all(not a.primary_key for a in cls._defined_attrs.values())
            return base_pks

        return {name: p for name, p in cls._defined_attrs.items() if p.primary_key}

    @property
    def _props(cls) -> dict[str, Set]:
        return reduce(
            lambda x, y: {**x, **y},
            (c._props for c in cls._record_bases),
            cls._defined_props,
        )

    @property
    def _attrs(cls) -> dict[str, ValueSet]:
        return reduce(
            lambda x, y: {**x, **y},
            (c._attrs for c in cls._record_bases),
            cls._defined_attrs,
        )

    @property
    def _rels(cls) -> dict[str, RelSet]:
        return reduce(
            lambda x, y: {**x, **y},
            (c._rels for c in cls._record_bases),
            cls._defined_rels,
        )

    @property
    def _rel_types(cls) -> set[type[Record]]:
        """Return all record types that are related to this record."""
        return {rel.target_type for rel in cls._rels.values()}


class Record(Generic[Key_def], metaclass=RecordMeta):
    """Schema for a record in a database."""

    _table_name: ClassVar[str]
    _type_map: ClassVar[dict[type, sqla.types.TypeEngine]] = {
        str: sqla.types.String().with_variant(
            sqla.types.String(50), "oracle"
        ),  # Avoid oracle error when VARCHAR has no size parameter
        UUID: UUIDType(binary=False),  # Binary type causes issues with DuckDB
    }

    _loader: Callable[[Set, Key_def], Any] | None = None
    _edge_dict: dict[str, RecordValue] = prop(default_factory=dict, local=True)

    @classmethod
    def _default_table_name(cls) -> str:
        """Return the name of the table for this schema."""
        return (
            cls._table_name
            if hasattr(cls, "_table_name")
            else PyObjectRef.reference(cls).fqn.replace(".", "_")
        )

    @classmethod
    def _sql_table_name(
        cls,
        subs: Mapping[type[Record], sqla.TableClause],
    ) -> str:
        """Return a SQLAlchemy table object for this schema."""
        sub = subs.get(cls)
        return sub.name if sub is not None else cls._default_table_name()

    @classmethod
    def _columns(cls, registry: orm.registry) -> list[sqla.Column]:
        """Columns of this record type's table."""
        table_attrs = (
            set(cls._defined_attrs.values())
            | set(cls._defined_rel_attrs.values())
            | set(cls._primary_keys.values())
        )

        return [
            sqla.Column(
                attr.name,
                (
                    registry._resolve_type(attr.prop_type.value_type())
                    if attr.prop_type
                    else None
                ),
                primary_key=attr.primary_key,
                index=attr.prop.index,
            )
            for attr in table_attrs
        ]

    @classmethod
    def _foreign_keys(
        cls, metadata: sqla.MetaData, subs: Mapping[type[Record], sqla.TableClause]
    ) -> list[sqla.ForeignKeyConstraint]:
        """Foreign key constraints for this record type's table."""
        return [
            *(
                sqla.ForeignKeyConstraint(
                    [attr.name for attr in rel.fk_map.keys()],
                    [attr.name for attr in rel.fk_map.values()],
                    table=rel.target_type._table(metadata, subs),
                    name=f"{cls._sql_table_name(subs)}_{rel.prop.name}_fk",
                )
                for rel in cls._defined_rels.values()
                if rel.fk_record_type is cls
            ),
            *(
                sqla.ForeignKeyConstraint(
                    [attr.name for attr in pks.keys()],
                    [attr.name for attr in pks.values()],
                    table=base._table(metadata, subs),
                    name=f"{cls._sql_table_name(subs)}_{base._sql_table_name(subs)}_inheritance_fk",
                )
                for base, pks in cls._base_pks.items()
            ),
        ]

    @classmethod
    def _table(
        cls,
        metadata: sqla.MetaData,
        subs: Mapping[type[Record], sqla.TableClause],
    ) -> sqla.Table:
        """Return a SQLAlchemy table object for this schema."""
        table_name = cls._sql_table_name(subs)

        if table_name in metadata.tables:
            # Return the table object from metadata if it already exists.
            # This is necessary to avoid circular dependencies.
            return metadata.tables[table_name]

        registry = orm.registry(metadata=metadata, type_annotation_map=cls._type_map)
        sub = subs.get(cls)

        # Create a partial SQLAlchemy table object from the class definition
        # without foreign keys to avoid circular dependencies.
        # This adds the table to the metadata.
        sqla.Table(
            table_name,
            registry.metadata,
            *cls._columns(registry),
            schema=(sub.schema if sub is not None else None),
        )

        # Re-create the table object with foreign keys and return it.
        return sqla.Table(
            table_name,
            registry.metadata,
            *cls._columns(registry),
            *cls._foreign_keys(metadata, subs),
            schema=(sub.schema if sub is not None else None),
            extend_existing=True,
        )

    @classmethod
    def _joined_table(
        cls,
        metadata: sqla.MetaData,
        subs: Mapping[type[Record], sqla.TableClause],
    ) -> sqla.Table | sqla.Join:
        """Recursively join all bases of this record to get the full data."""
        table = cls._table(metadata, subs)

        base_joins = [
            (base._joined_table(metadata, subs), pk_map)
            for base, pk_map in cls._base_pks.items()
        ]
        for target_table, pk_map in base_joins:
            table = table.join(
                target_table,
                reduce(
                    sqla.and_,
                    (
                        table.c[pk.name] == target_table.c[target_pk.name]
                        for pk, target_pk in pk_map.items()
                    ),
                ),
            )

        return table

    @classmethod
    def _backrels_to_rels(
        cls, target: type[Rec]
    ) -> set[RelSet[Self, SingleIdx, None, Any, Rec]]:
        """Get all direct relations from a target record type to this type."""
        return {
            RelSet[Self, SingleIdx, None, Any, Rec](
                prop=Rel(on=rel.prop.on),  # type: ignore
                record_type=cast(type[Rec], rel.fk_record_type),
                typedef=cls,
            )
            for rel in cls._rels.values()
            if issubclass(target, rel.fk_record_type)
        }

    @overload
    @classmethod
    def _rel(
        cls, other: type[Rec], index: None = ...
    ) -> RelSet[Rec, BaseIdx, None, R, Self]: ...

    @overload
    @classmethod
    def _rel(
        cls, other: type[Rec], index: type[Key]
    ) -> RelSet[Rec, SingleIdx, None, R, Self]: ...

    @classmethod
    def _rel(
        cls, other: type[Rec], index: type[Key] | None = None
    ) -> RelSet[Rec, Any, None, R, Self]:
        """Dynamically define a relation to another record type."""
        return RelSet[Rec, Any, None, R, Self](
            prop=Rel(on=other),
            record_type=cls,
            typedef=other,
        )

    @classmethod
    def __clause_element__(cls) -> sqla.TableClause:  # noqa: D105
        assert cls._default_table_name() is not None
        return sqla.table(cls._default_table_name())

    def __hash__(self) -> int:
        """Hash the record."""
        return gen_int_hash(self)

    def __eq__(self, value: Hashable) -> bool:
        """Check if the record is equal to another record."""
        return hash(self) == hash(value)

    @overload
    def _to_dict(
        self,
        name_keys: Literal[False] = ...,
        only_loaded: bool = ...,
        load: bool = ...,
        include: tuple[type[Prop], ...] = ...,
    ) -> dict[Set[Self, Any], Any]: ...

    @overload
    def _to_dict(
        self,
        name_keys: Literal[True],
        only_loaded: bool = ...,
        load: bool = ...,
        include: tuple[type[Prop], ...] = ...,
    ) -> dict[str, Any]: ...

    def _to_dict(
        self,
        name_keys: bool = False,
        only_loaded: bool = True,
        load: bool = False,
        include: tuple[type[Prop], ...] = (Attr, Rel),
    ) -> dict[Set[Self, Any], Any] | dict[str, Any]:
        """Convert the record to a dictionary."""

        def getter(r, n):
            return r.__dict__[n] if not load else getattr(r, n)

        vals = {
            p if not name_keys else p.name: getter(self, p.name)
            for p in type(self)._props.values()
            if isinstance(p, include)
        }

        return cast(
            dict,
            (
                vals
                if not only_loaded
                else {k: v for k, v in vals.items() if not isinstance(v, Unloaded)}
            ),
        )

    @property
    def _index(self: Record[Key]) -> Key:
        """Return the index of the record."""
        pks = type(self)._primary_keys
        if len(pks) == 1:
            return getattr(self, next(iter(pks)))
        return cast(Key, tuple(getattr(self, pk) for pk in pks))

    @classmethod
    def _index_from_dict(cls: type[Record[Key]], data: Mapping[Set, Any]) -> Key:
        """Return the index contained in a dict representation of this record."""
        pks = cls._primary_keys
        if len(pks) == 1:
            return data[next(iter(pks.values()))]
        return cast(Key, tuple(data[pk] for pk in pks.values()))

    @overload
    def _link(self, rel: RelSet[Rec, SingleIdx, None, RW, Self, Rec2]) -> Rec2: ...

    @overload
    def _link(
        self, rel: RelSet[Rec, SingleIdx | None, None, RW, Self, Rec2]
    ) -> Rec2 | None: ...

    @overload
    def _link(
        self, rel: RelSet[Rec, Key, None, RW, Self, Rec2]
    ) -> Mapping[Key, Rec2]: ...

    @overload
    def _link(self, rel: RelSet[Rec, BaseIdx, None, RW, Self, Rec2]) -> list[Rec2]: ...

    def _link(self, rel: RelSet[Rec, Any, None, RW, Self, Rec2]) -> RecordValue:
        """Return the model of the given relation, if any."""
        if rel.link_type is Record:
            return DynRecord()

        if self._loader is not None and rel.prop.name not in self._edge_dict:
            self._edge_dict[rel.prop.name] = self._loader(rel, self._index)

        return self._edge_dict[rel.prop.name]


class Schema:
    """Group multiple record types into a schema."""

    _record_types: set[type[Record]]
    _rel_record_types: set[type[Record]]

    def __init_subclass__(cls) -> None:  # noqa: D105
        subclasses = get_subclasses(cls, max_level=1)
        cls._record_types = {s for s in subclasses if isinstance(s, Record)}
        cls._rel_record_types = {rr for r in cls._record_types for rr in r._rel_types}
        super().__init_subclass__()


@dataclass
class Require:
    """Mark schema or record type as required."""

    present: bool = True


class RecUUID(Record[UUID]):
    """Dynamically defined record type."""

    _id: Attr[UUID] = prop(primary_key=True, default_factory=uuid4)


class Scalar(Record[Key_def], Generic[Val, Key_def]):
    """Dynamically defined record type."""

    _id: Attr[Key_def] = prop(primary_key=True, default_factory=uuid4)
    _value: Attr[Val]


class DynRecordMeta(RecordMeta):
    """Metaclass for dynamically defined record types."""

    def __getitem__(cls: type[Record], name: str) -> ValueSet:
        """Get dynamic attribute by dynamic name."""
        return ValueSet(prop=Attr(_name=name), record_type=cls, typedef=object)

    def __getattr__(cls: type[Record], name: str) -> ValueSet:
        """Get dynamic attribute by name."""
        if not TYPE_CHECKING and name.startswith("__"):
            return super().__getattribute__(name)

        return ValueSet(prop=Attr(_name=name), record_type=cls, typedef=object)


class DynRecord(Record, metaclass=DynRecordMeta):
    """Dynamically defined record type."""


a = DynRecord


def dynamic_record_type(name: str, props: Iterable[Prop] = []) -> type[DynRecord]:
    """Create a dynamically defined record type."""
    return type(name, (DynRecord,), {p.name: p for p in props})


type DataDict = dict[RelSet, DataDict]


@dataclass
class RelTree(Generic[*RelTup]):
    """Tree of relations starting from the same root."""

    rels: Iterable[RelSet] = field(default_factory=set)

    def __post_init__(self) -> None:  # noqa: D105
        assert all(
            rel.rel_path[0].record_type == self.root for rel in self.rels
        ), "Relations in set must all start from same root."
        self.targets = [rel.target_type for rel in self.rels]

    @cached_property
    def root(self) -> type[Record]:
        """Root record type of the set."""
        root = list(self.rels)[-1].rel_path[0].record_type
        assert issubclass(root, Record)
        return root

    @cached_property
    def dict(self) -> DataDict:
        """Tree representation of the relation set."""
        tree: DataDict = {}

        for rel in self.rels:
            subtree = tree
            for ref in rel.rel_path:
                if ref not in subtree:
                    subtree[ref] = {}
                subtree = subtree[ref]

        return tree

    def prefix(
        self, prefix: type[Record] | RecordSet[Any, Any, Any, Any, Any, Any, Any]
    ) -> Self:
        """Prefix all relations in the set with given relation."""
        rels = {rel.prefix(prefix) for rel in self.rels}
        return cast(Self, RelTree(rels))

    def __rmul__(self, other: RelSet[Rec] | RelTree) -> RelTree[*RelTup, Rec]:
        """Append more rels to the set."""
        other = other if isinstance(other, RelTree) else RelTree([other])
        return RelTree([*self.rels, *other.rels])

    def __or__(
        self: RelTree[Rec], other: RelSet[Rec2] | RelTree[Rec2]
    ) -> RelTree[Rec | Rec2]:
        """Add more rels to the set."""
        other = other if isinstance(other, RelTree) else RelTree([other])
        return RelTree([*self.rels, *other.rels])


@dataclass(kw_only=True, eq=False)
class Set(Generic[Val_def, KeyIdx_def, B_nul, Rec2_nul, P_nul]):
    """Reference a property of a record."""

    typedef: PropType[Prop[Val_def, Any]] | type[Val_def]
    record_type: type[Rec2_nul] = type(None)
    prop: P_nul = None

    schema_types: set[type[Record]] = field(default_factory=set)

    overlay: str | None = None
    overlay_with_schemas: bool = True
    subs: dict[type[Record], sqla.TableClause] = field(default_factory=dict)
    merges: RelTree = field(default_factory=RelTree)
    filters: list[sqla.ColumnElement[bool]] = field(default_factory=list)
    keys: Sequence[slice | list[Hashable] | Hashable | sqla.ColumnElement] = field(
        default_factory=list
    )

    backend: B_nul = None
    create_cross_fk: bool = True

    @staticmethod
    def get_tag(rec_type: type[Record]) -> RelSet | None:
        """Retrieve relation-tag of a record type, if any."""
        try:
            rel = getattr(rec_type, "_rel")
            return rel if isinstance(rel, RelSet) else None
        except AttributeError:
            return None

    @cached_property
    def meta(self) -> sqla.MetaData:
        """Metadata object for this DB instance."""
        return sqla.MetaData()

    @property
    def target_type(self) -> type[Val_def]:
        """Value type of the property."""
        return (
            self.typedef.value_type()
            if isinstance(self.typedef, PropType)
            else self.typedef
        )

    @cached_property
    def path_idx(self) -> list[ValueSet] | None:
        """Get the path index of the relation."""
        if not isinstance(self, RelSet):
            return None

        p = [
            a
            for rel in self.rel_path
            if rel.prop is not None
            for a in (
                [rel.prop.map_by]
                if rel.prop.map_by is not None
                else rel.prop.order_by.keys() if rel.prop.order_by is not None else []
            )
        ]

        if len(p) == 0:
            return None

        return [
            *self.rel_path[0].record_type._primary_keys.values(),
            *p,
        ]

    @cached_property
    def path_str(self) -> str | None:
        """String representation of the relation path."""
        if self.prop is None or not isinstance(self, RelSet):
            return None

        prefix = (
            self.record_type.__name__
            if len(self.rel_path) == 0
            else self.rel_path[-1].path_str
        )
        return f"{prefix}.{self.prop.name}"

    @cached_property
    def idx(self) -> set[ValueSet]:
        """Return the index attrs."""
        path_idx = self.path_idx
        if path_idx is not None:
            return {p for p in path_idx}

        if issubclass(self.target_type, Record):
            return set(self.target_type._primary_keys.values())

        assert issubclass(self.record_type, Record)
        return set(self.record_type._primary_keys.values())

    @cached_property
    def base_table(self) -> sqla.FromClause:
        """Get the main table for the current selection."""
        if isinstance(self, RelSet):
            rec = self.rel_path[0].record_type
            if issubclass(rec, Record):
                return self._get_random_alias(rec)
        elif issubclass(self.target_type, Record):
            return self._get_random_alias(self.target_type)

        assert issubclass(self.record_type, Record)
        return self._get_random_alias(self.record_type)

    def prefix(
        self, left: type[Record] | RecordSet[Any, Any, Any, Any, Any, Any, Any]
    ) -> Self:
        """Prefix this prop with a relation or record type."""
        assert self.prop is not None and issubclass(self.record_type, Record)

        rel_path = (
            self.rel_path
            if isinstance(self, RelSet)
            else (
                self.record_set.rel_path
                if isinstance(self, ValueSet) and isinstance(self.record_set, RelSet)
                else []
            )
        )
        current_root = rel_path[0] if len(rel_path) > 0 else self.record_type
        new_root = (
            left
            if isinstance(left, RecordSet)
            else left._rel(
                current_root.target_type
                if isinstance(current_root, RecordSet)
                else current_root
            )
        )

        prefixed_rel = reduce(
            lambda r1, r2: RecordSet(**asdict(r2), record_type=r1.t),  # type: ignore
            rel_path,
            new_root,
        )

        return cast(
            Self,
            (
                prefixed_rel
                if isinstance(self, RecordSet)
                else getattr(prefixed_rel.rec, self.prop.name)
            ),
        )

    @cached_property
    def engine(self) -> sqla.engine.Engine:
        """SQLA Engine for this DB."""
        # Create engine based on backend type
        # For Excel-backends, use duckdb in-memory engine
        return (
            sqla.create_engine(
                self.backend.url
                if isinstance(self.backend.url, sqla.URL)
                else str(self.backend.url)
            )
            if self.backend is not None
            and (
                self.backend.type == "sql-connection"
                or self.backend.type == "sqlite-file"
            )
            else (sqla.create_engine("duckdb:///:memory:"))
        )

    @cached_property
    def db(self: Set[Any, Any, B_def]) -> DB[B_def]:
        """Get the DB object."""
        return DB(**{**asdict(self), **dict(typedef=Record)})

    def __hash__(self) -> int:  # noqa: D105
        return gen_int_hash(self)

    def __setitem__(self, key: Any, other: Set) -> None:
        """Catchall setitem."""
        raise NotImplementedError()

    def _get_table(self, rec: type[Record], writable: bool = False) -> sqla.Table:
        if writable and self.overlay is not None and rec not in self.subs:
            # Create an empty overlay table for the record type
            self.subs[rec] = sqla.table(
                (
                    (self.overlay + "_" + rec._default_table_name())
                    if self.overlay_with_schemas
                    else rec._default_table_name()
                ),
                schema=self.overlay if self.overlay_with_schemas else None,
            )

        table = rec._table(self.meta, self.subs)

        # Create any missing tables in the database.
        self.meta.create_all(self.engine)

        return table

    def _get_joined_table(self, rec: type[Record]) -> sqla.Table | sqla.Join:
        table = rec._joined_table(self.meta, self.subs)

        # Create any missing tables in the database.
        self.meta.create_all(self.engine)

        return table

    def _get_alias(self, rel: RelSet) -> sqla.FromClause:
        """Get alias for a relation reference."""
        return self._get_joined_table(rel.target_type).alias(gen_str_hash(rel, 8))

    def _get_random_alias(self, rec: type[Record]) -> sqla.FromClause:
        """Get random alias for a type."""
        return self._get_joined_table(rec).alias(token_hex(4))

    def _parse_merge_tree(self, merge: RelSet | RelTree | None) -> RelTree:
        """Parse merge argument and prefix with current selection."""
        assert issubclass(self.record_type, Record) and issubclass(
            self.target_type, Record
        )

        merge = (
            merge
            if isinstance(merge, RelTree)
            else RelTree({merge}) if merge is not None else RelTree()
        )

        return merge.prefix(cast(RelSet, self))

    def _gen_idx_match_expr(
        self,
        values: Sequence[slice | list[Hashable] | Hashable | sqla.ColumnElement],
    ) -> sqla.ColumnElement[bool] | None:
        """Generate SQL expression for matching index values."""
        if values == slice(None):
            return None

        exprs = [
            (
                idx.label(None).in_(val)
                if isinstance(val, list)
                else (
                    idx.label(None).between(val.start, val.stop)
                    if isinstance(val, slice)
                    else idx.label(None) == val
                )
            )
            for idx, val in zip(self.idx, values)
        ]

        if len(exprs) == 0:
            return None

        return reduce(sqla.and_, exprs)

    def _replace_attr(
        self,
        element: sqla_visitors.ExternallyTraversible,
        reflist: set[RelSet] = set(),
        **kw: Any,
    ) -> sqla.ColumnElement | None:
        if isinstance(element, ValueSet):
            if isinstance(element.record_set, RelSet):
                reflist.add(element.record_set)

            if issubclass(self.record_type, Record):
                element = element.prefix(cast(RelSet, self))

            table = (
                self._get_alias(element.record_set)
                if isinstance(element.record_set, RelSet)
                else self._get_table(element.record_type)
            )
            return table.c[element.prop.name]

        return None

    def _parse_filter(
        self,
        key: sqla.ColumnElement[bool],
    ) -> tuple[sqla.ColumnElement[bool], RelTree]:
        """Parse filter argument and return SQL expression and join operations."""
        reflist: set[RelSet] = set()
        replace_func = partial(self._replace_attr, reflist=reflist)
        filt = sqla_visitors.replacement_traverse(key, {}, replace=replace_func)
        merge = RelTree(reflist)

        return filt, merge

    def _parse_schema_items(
        self,
        element: sqla_visitors.ExternallyTraversible,
        **kw: Any,
    ) -> sqla.ColumnElement | sqla.FromClause | None:
        if isinstance(element, RelSet):
            return self._get_alias(element)
        elif isinstance(element, ValueSet):
            table = (
                self._get_alias(element.record_set)
                if isinstance(element.record_set, RelSet)
                else self._get_table(element.record_type)
            )
            return table.c[element.name]
        elif has_type(element, type[Record]):
            return self._get_table(element)

        return None

    def _parse_expr[CE: sqla.ClauseElement](self, expr: CE) -> CE:
        """Parse an expression in this database's context."""
        return cast(
            CE,
            sqla_visitors.replacement_traverse(
                expr, {}, replace=self._parse_schema_items
            ),
        )

    def _ensure_schema_exists(self, schema_name: str) -> str:
        """Ensure that the table exists in the database, then return it."""
        if not sqla.inspect(self.engine).has_schema(schema_name):
            with self.engine.begin() as conn:
                conn.execute(sqla.schema.CreateSchema(schema_name))

        return schema_name

    def _table_exists(self, sqla_table: sqla.Table) -> bool:
        """Check if a table exists in the database."""
        return sqla.inspect(self.engine).has_table(
            sqla_table.name, schema=sqla_table.schema
        )

    def _create_sqla_table(self, sqla_table: sqla.Table) -> None:
        """Create SQL-side table from Table class."""
        if not self.create_cross_fk:
            # Create a temporary copy of the table object and remove external FKs.
            # That way, local metadata will retain info on the FKs
            # (for automatic joins) but the FKs won't be created in the DB.
            sqla_table = sqla_table.to_metadata(sqla.MetaData())  # temporary metadata
            _remove_external_fk(sqla_table)

        sqla_table.create(self.engine)

    def _load_from_excel(self, record_types: list[type[Record]] | None = None) -> None:
        """Load all tables from Excel."""
        assert self.backend is not None
        assert self.backend.type == "excel-file", "Backend must be an Excel file."
        assert isinstance(self.backend.url, Path | CloudPath | HttpFile)

        path = (
            self.backend.url.get()
            if isinstance(self.backend.url, HttpFile)
            else self.backend.url
        )

        with open(path, "rb") as file:
            for rec in record_types or self.schema_types:
                pl.read_excel(
                    file, sheet_name=rec._default_table_name()
                ).write_database(str(self._get_table(rec)), str(self.engine.url))

    def _save_to_excel(
        self, record_types: Iterable[type[Record]] | None = None
    ) -> None:
        """Save all (or selected) tables to Excel."""
        assert self.backend is not None
        assert self.backend.type == "excel-file", "Backend must be an Excel file."
        assert isinstance(self.backend.url, Path | CloudPath | HttpFile)

        file = (
            BytesIO()
            if isinstance(self.backend.url, HttpFile)
            else self.backend.url.open("wb")
        )

        with ExcelWorkbook(file) as wb:
            for rec in record_types or self.schema_types:
                pl.read_database(
                    f"SELECT * FROM {self._get_table(rec)}",
                    self.engine,
                ).write_excel(wb, worksheet=rec._default_table_name())

        if isinstance(self.backend.url, HttpFile):
            assert isinstance(file, BytesIO)
            self.backend.url.set(file)

    def _delete_from_excel(self, record_types: Iterable[type[Record]]) -> None:
        """Delete selected table from Excel."""
        assert self.backend is not None
        assert self.backend.type == "excel-file", "Backend must be an Excel file."
        assert isinstance(self.backend.url, Path | CloudPath | HttpFile)

        file = (
            BytesIO()
            if isinstance(self.backend.url, HttpFile)
            else self.backend.url.open("wb")
        )

        wb = openpyxl.load_workbook(file)
        for rec in record_types or self.schema_types:
            del wb[rec._default_table_name()]

        if isinstance(self.backend.url, HttpFile):
            assert isinstance(file, BytesIO)
            self.backend.url.set(file)

        raise TypeError("Invalid property reference.")

    def _load_prop(
        self: Set[Any, Any, Backend],
        p: Set[Val, Record[Key]],
        parent_idx: Key,
    ) -> Val:
        base = self.db[p.record_type]
        base_record = base[parent_idx]

        if isinstance(p, ValueSet):
            return getattr(base_record.load(), p.name)
        elif isinstance(p, RecordSet):
            recs = base_record[p].load()
            recs_type = p.target_type

            if (
                isinstance(recs, dict)
                and not issubclass(recs_type, Mapping)
                and issubclass(recs_type, Iterable)
            ):
                recs = list(recs.values())

            if p.prop is not None and p.prop.collection is not None:
                recs = p.prop.collection(recs)

            return cast(Val, recs)

        raise TypeError("Invalid property reference.")

    @staticmethod
    def _normalize_rel_data(
        rec_data: RecordValue[Record],
        parent_idx: Hashable,
        list_idx: bool,
        covered: set[int],
    ) -> dict[Any, Record]:
        """Convert record data to dictionary."""
        res = {}

        if isinstance(rec_data, Record) and id(rec_data) not in covered:
            res = {parent_idx: rec_data}
        elif isinstance(rec_data, Mapping):
            res = {
                (parent_idx, idx): rec
                for idx, rec in rec_data.items()
                if id(rec) not in covered
            }
        elif isinstance(rec_data, Iterable):
            res = {
                (parent_idx, idx if list_idx else rec._index): rec
                for idx, rec in enumerate(rec_data)
                if id(rec) not in covered
            }

        covered |= set(id(r) for r in res.values())
        return res

    @staticmethod
    def _get_record_rels(
        rec_data: dict[Any, dict[Set, Any]], list_idx: bool, covered: set[int]
    ) -> dict[RelSet, dict[Any, Record]]:
        """Get relation data from record data."""
        rel_data: dict[RelSet, dict[Any, Record]] = {}

        for idx, rec in rec_data.items():
            for prop, prop_val in rec.items():
                if isinstance(prop, RecordSet) and not isinstance(prop_val, Unloaded):
                    prop = cast(RelSet, prop)
                    rel_data[prop] = {
                        **rel_data.get(prop, {}),
                        **Set._normalize_rel_data(prop_val, idx, list_idx, covered),
                    }

        return rel_data


@dataclass(kw_only=True, eq=False)
class ValueSet(
    Set[Val_defi, KeyIdx_def, B_nul, Rec_def, Attr[Val_defi, R_def]],
    sqla.ColumnClause[Val_defi],
):
    """Reference an attribute of a record."""

    def __post_init__(self) -> None:  # noqa: D105
        # Initialize fields required by SQLAlchemy superclass.
        self.table = None
        self.is_literal = False

    # Replace fields of SQLAlchemy superclass with properties:

    @property
    def name(self) -> str:
        """Column key."""
        assert Prop.name.fget is not None
        return Prop.name.fget(self)

    @property
    def key(self) -> str:
        """Column key."""
        return self.name

    @cached_property
    def sql_type(self) -> sqla_types.TypeEngine:
        """Column key."""
        return sqla_types.to_instance(self.target_type)  # type: ignore

    def all(self) -> sqla.CollectionAggregate[bool]:
        """Return a SQL ALL expression for this attribute."""
        return sqla.all_(self)

    def any(self) -> sqla.CollectionAggregate[bool]:
        """Return a SQL ANY expression for this attribute."""
        return sqla.any_(self)

    # Value set interface:

    record_set: RecordSet[Any, KeyIdx_def, B_nul, R_def, Rec_def] | None = None

    # Plural selection
    @overload
    def __getitem__(
        self: ValueSet[Any, Key],
        key: (
            Iterable[Key_def]
            | (ValueSet[bool, Key_def] | ValueSet[Key_def, Any])
            | slice
            | tuple[slice, ...]
        ),
    ) -> ValueSet[Val_def, Key, B_nul, Rec_def]: ...

    # Single value selection
    @overload
    def __getitem__(
        self: ValueSet[Any, Key], key: Key
    ) -> ValueSet[Val_def, SingleIdx, B_nul, Rec_def]: ...

    # Implementation:

    def __getitem__(  # noqa: D105
        self: ValueSet,
        key: ValueSet | list[Hashable] | slice | tuple[slice, ...] | Hashable,
    ) -> ValueSet:
        vs = ValueSet(
            **{
                **asdict(self),
                **dict(
                    keys=[key],
                ),
            }
        )

        if self.record_set is not None:
            return vs.prefix(self.record_set)
        elif issubclass(self.record_type, Record):
            return vs.prefix(cast(RelSet, self))

        return vs

    @cached_property
    def _idx_cols(self) -> list[sqla.ColumnElement]:
        """Return the index columns."""
        return [
            *(
                col.label(f"{self.record_type._default_table_name()}.{col_name}")
                for col_name, col in self.base_table.columns.items()
                if col_name == self.prop.name
                or self.record_type._attrs[col_name] in self.idx
            ),
            *(
                col.label(f"{rel.path_str}.{col_name}")
                for rel in self.merges.rels
                for col_name, col in self._get_alias(rel).columns.items()
                if rel.target_type._attrs[col_name] in self.idx
            ),
        ]

    def select(self) -> sqla.Select:
        """Return select statement for this dataset."""
        selection_table = self.base_table
        assert selection_table is not None

        select = sqla.select(
            selection_table.c[self.prop.name], *self._idx_cols
        ).select_from(selection_table)

        for join in self.joins():
            select = select.join(*join)

        for filt in self.filters:
            select = select.where(filt)

        return select

    @overload
    def load(  # type: ignore
        self: ValueSet[Val_cov, SingleIdx, Backend],
        kind: type[Val_cov] = ...,
    ) -> Val_cov: ...

    @overload
    def load(self: ValueSet[Any, Any, Backend], kind: type[Series]) -> Series: ...

    @overload
    def load(
        self: ValueSet[Val_cov, Key, Backend], kind: type[Val_cov] = ...
    ) -> dict[Key, Val_cov]: ...

    def load(
        self: ValueSet[Any, Any, Backend],
        kind: type[Val_cov | Series] = Record,
    ) -> Val_cov | Series | dict[Any, Val_cov]:
        """Download selection."""
        select = self.select()

        if kind is pd.Series:
            with self.engine.connect() as con:
                return pd.read_sql(select, con).set_index(
                    [c.key for c in self._idx_cols]
                )[self.prop.name]
        elif kind is pl.Series:
            return pl.read_database(select, self.engine)[self.prop.name]

        with self.engine.connect() as con:
            return (
                pd.read_sql(select, con)
                .set_index([c.key for c in self._idx_cols])[self.prop.name]
                .to_dict()
            )

    def __imatmul__(
        self: ValueSet[Val_cov, KeyIdx_def, B_def, Rec_def, R_def],
        value: ValInput | ValueSet[Val_cov, Key_def, B_def],
    ) -> ValueSet[Val_cov, KeyIdx_def, B_def, Rec_def, R_def]:
        """Do item-wise, broadcasting assignment on this dataset."""
        # Load data into a temporary table.
        value_df = None
        if isinstance(value, ValueSet):
            value_set = value
        elif isinstance(value, sqla.Select):
            value_set = self.db.dataset(value)
        else:
            value_df = (
                value.to_frame()
                if isinstance(value, Series)
                else (
                    pd.DataFrame.from_records(value)
                    if isinstance(value, Mapping)
                    else pd.DataFrame({self.prop.name: [value]})
                )
            )
            value_set = self.db.dataset(value_df)

        # Derive current select statement and join with value table, if exists.
        select = self.select()
        if value_set is not None:
            select = select.join(
                value_set.base_table,
                reduce(
                    sqla.and_,
                    (
                        self.base_table.c[idx_col.name] == idx_col
                        for idx_col in value_set.base_table.primary_key
                    ),
                ),
            )

        assert isinstance(self.base_table, sqla.Table), "Base must be a table."

        if self.engine.dialect.name in (
            "postgres",
            "postgresql",
            "duckdb",
            "mysql",
            "mariadb",
        ):
            # Update-from.
            statement = (
                self.base_table.update()
                .values(
                    {c_name: c for c_name, c in value_set.base_table.columns.items()}
                )
                .where(
                    reduce(
                        sqla.and_,
                        (
                            self.base_table.c[col.name] == select.c[col.name]
                            for col in self.base_table.primary_key.columns
                        ),
                    )
                )
            )

            with self.engine.begin() as con:
                con.execute(statement)
        else:
            # Correlated update.
            raise NotImplementedError("Correlated update not supported yet.")

        # Drop the temporary table, if any.
        if value_set is not None and value_set is not value:
            cast(sqla.Table, value_set.base_table).drop(self.engine)

        return self


type AggMap[Rec: Record] = dict[ValueSet[Rec, Any], ValueSet | sqla.Function]


@dataclass(frozen=True)
class Agg(Generic[Rec]):
    """Define an aggregation map."""

    target: type[Rec]
    map: AggMap[Rec]


type Join = tuple[sqla.FromClause, sqla.ColumnElement[bool]]


@dataclass(kw_only=True, eq=False)
class RecordSet(
    Set[Rec_cov, Rec2_nul, B_nul, Rec2_nul, Rel[Any, Record, R_def, Record] | None],
    Generic[Rec_cov, Idx_def, B_nul, R_def, Rec2_nul, Rec3_def, RM],
):
    """Dataset."""

    @cached_property
    def rec(self: RecordSet[Rec, Any, Any, Any, Any, Any]) -> type[Rec]:
        """Reference props of the target record type."""
        assert issubclass(self.target_type, Record)
        return cast(type[Rec], type(token_hex(5), (self.target_type,), {"_rel": self}))

    # Overloads: attribute selection:

    # 1. DB-level type selection
    @overload
    def __getitem__(  # type: ignore
        self,
        key: type[Rec],
    ) -> RecordSet[Rec, Idx_def, B_nul, R_def, Rec2_nul, Rec2_def]: ...

    # 2. Top-level attribute selection, custom index
    @overload
    def __getitem__(
        self: RecordSet[Any, Key | FilteredIdx[Key]],
        key: ValueSet[Val, Rec_cov],
    ) -> ValueSet[Val, Key, B_nul, Rec_cov, R_def]: ...

    # 3. Top-level attribute selection, base index
    @overload
    def __getitem__(
        self: RecordSet[Record[Key2], Any],
        key: ValueSet[Val, Rec_cov],
    ) -> ValueSet[Val, Key2, B_nul, Rec_cov, R_def]: ...

    # 4. Nested attribute selection, custom index
    @overload
    def __getitem__(
        self: RecordSet[Any, Key | FilteredIdx[Key]],
        key: ValueSet[Val, Any],
    ) -> ValueSet[Val, Key | IdxStart[Key], B_nul, Record, R_def]: ...

    # 5. Nested attribute selection, base index
    @overload
    def __getitem__(
        self: RecordSet[Record[Key2], Any],
        key: ValueSet[Val, Any],
    ) -> ValueSet[Val, Key2 | IdxStart[Key2], B_nul, Record, R_def]: ...

    # Overloads: relation selection:

    # 6. Top-level relation selection, singular, base index
    @overload
    def __getitem__(  # type: ignore
        self: RecordSet[Record[Key2], BaseIdx | FilteredIdx[BaseIdx]],
        key: RelSet[Rec2, SingleIdx | None, None, Any, Rec_cov, Rec3],
    ) -> RecordSet[Rec2, Key2, B_nul, R_def, Rec_cov, Rec3]: ...

    # 7. Top-level relation selection, singular, single index
    @overload
    def __getitem__(  # type: ignore
        self: RecordSet[Any, SingleIdx],
        key: RelSet[Rec2, SingleIdx, None, Any, Rec_cov, Rec3],
    ) -> RecordSet[Rec2, SingleIdx, B_nul, R_def, Rec_cov, Rec3]: ...

    # 8. Top-level relation selection, singular nullable, single index
    @overload
    def __getitem__(
        self: RecordSet[Any, SingleIdx],
        key: RelSet[Rec2, SingleIdx | None, None, Any, Rec_cov, Rec3],
    ) -> RecordSet[Rec2 | Scalar[None], SingleIdx, B_nul, R_def, Rec_cov, Rec3]: ...

    # 9. Top-level relation selection, singular, custom index
    @overload
    def __getitem__(  # type: ignore
        self: RecordSet[Any, Key | FilteredIdx[Key]],
        key: RelSet[Rec2, SingleIdx | None, None, Any, Rec_cov, Rec3],
    ) -> RecordSet[Rec2, Key, B_nul, R_def, Rec_cov, Rec3]: ...

    # 10. Top-level relation selection, plural, base index
    @overload
    def __getitem__(
        self: RecordSet[Record[Key2], BaseIdx | FilteredIdx[BaseIdx]],
        key: RelSet[Rec2, Key3, None, Any, Rec_cov, Rec3],
    ) -> RecordSet[Rec2, tuple[Key2, Key3], B_nul, R_def, Rec_cov, Rec3]: ...

    # 11. Top-level relation selection, plural, single index
    @overload
    def __getitem__(
        self: RecordSet[Any, SingleIdx],
        key: RelSet[Rec2, Key3, None, Any, Rec_cov, Rec3],
    ) -> RecordSet[Rec2, Key3, B_nul, R_def, Rec_cov, Rec3]: ...

    # 12. Top-level relation selection, plural, tuple index
    @overload
    def __getitem__(
        self: RecordSet[Record[Key2], tuple[*IdxTup] | FilteredIdx[tuple[*IdxTup]]],
        key: RelSet[Rec2, Key3, None, Any, Rec_cov, Rec3],
    ) -> RecordSet[
        Rec2, tuple[*IdxTup, Key3] | tuple[Key2, Key3], B_nul, R_def, Rec_cov, Rec3
    ]: ...

    # 13. Top-level relation selection, plural, custom index
    @overload
    def __getitem__(
        self: RecordSet[Any, Key | FilteredIdx[Key]],
        key: RelSet[Rec2, Key3, None, Any, Rec_cov, Rec3],
    ) -> RecordSet[Rec2, tuple[Key, Key3], B_nul, R_def, Rec_cov, Rec3]: ...

    # 14. Nested relation selection, base index
    @overload
    def __getitem__(
        self: RecordSet[Record[Key2], BaseIdx | FilteredIdx[BaseIdx]],
        key: RelSet[Rec2, Key3 | SingleIdx, None, Any, Rec, Rec3],
    ) -> RecordSet[Rec2, IdxStartEnd[Key2, Key3], B_nul, R_def, Rec, Rec3]: ...

    # 15. Nested relation selection, single index
    @overload
    def __getitem__(
        self: RecordSet[Any, SingleIdx],
        key: RelSet[Rec2, Key3 | SingleIdx, None, Any, Rec, Rec3],
    ) -> RecordSet[Rec2, IdxEnd[Key3], B_nul, R_def, Rec, Rec3]: ...

    # 16. Nested relation selection, tuple index
    @overload
    def __getitem__(
        self: RecordSet[Any, tuple[*IdxTup] | FilteredIdx[tuple[*IdxTup]]],
        key: RelSet[Rec2, Key3 | SingleIdx, None, Any, Rec, Rec3],
    ) -> RecordSet[Rec2, IdxTupStartEnd[*IdxTup, Key3], B_nul, R_def, Rec, Rec3]: ...

    # 17. Nested relation selection, custom index
    @overload
    def __getitem__(
        self: RecordSet[Any, Key | FilteredIdx[Key]],
        key: RelSet[Rec2, Key3 | SingleIdx, None, Any, Rec, Rec3],
    ) -> RecordSet[Rec2, IdxStartEnd[Key, Key3], B_nul, R_def, Rec, Rec3]: ...

    # 18. Default relation selection
    @overload
    def __getitem__(
        self: RecordSet,
        key: RelSet[Rec2, Any, None, Any, Rec, Rec3],
    ) -> RecordSet[Rec2, Any, B_nul, R_def, Rec, Rec3]: ...

    # 19. Merge selection, single index
    @overload
    def __getitem__(
        self: RecordSet[Any, SingleIdx, Any, Any, Any, Any, None],
        key: RelTree[Rec_cov, *RelTup],
    ) -> RecordSet[
        Rec_cov, BaseIdx, B_nul, R_def, Rec2_nul, Rec3_def, tuple[Rec_cov, *RelTup]
    ]: ...

    # 20. Merge selection, default
    @overload
    def __getitem__(
        self: RecordSet[Any, Any, Any, Any, Any, Any, None],
        key: RelTree[Rec_cov, *RelTup],
    ) -> RecordSet[
        Rec_cov, Idx_def, B_nul, R_def, Rec2_nul, Rec3_def, tuple[Rec_cov, *RelTup]
    ]: ...

    # 21. Expression filtering, keep index
    @overload
    def __getitem__(
        self: RecordSet, key: sqla.ColumnElement[bool]
    ) -> RecordSet[Rec_cov, FilteredIdx[Idx_def], B_nul, R_def, Rec2_nul, Rec3_def]: ...

    # 22. List selection
    @overload
    def __getitem__(
        self: RecordSet[Record[Key2], Key | Index], key: Iterable[Key | Key2]
    ) -> RecordSet[Rec_cov, FilteredIdx[Idx_def], B_nul, R_def, Rec2_nul, Rec3_def]: ...

    # 23. Filtering based on dataset, base index
    @overload
    def __getitem__(
        self: RecordSet[Record[Key2], BaseIdx | FilteredIdx[BaseIdx]],
        key: (
            RecordSet[Scalar[bool, Key2], BaseIdx]
            | RecordSet[Scalar[bool], Key2]
            | RecordSet[Scalar[Key2], Any]
        ),
    ) -> RecordSet[Rec_cov, FilteredIdx[Idx_def], B_nul, R_def, Rec2_nul, Rec3_def]: ...

    # 24. Filtering based on dataset, custom index
    @overload
    def __getitem__(
        self: RecordSet[Any, Key | FilteredIdx[Key]],
        key: (
            RecordSet[Scalar[bool, Key], BaseIdx]
            | RecordSet[Scalar[bool], Key]
            | RecordSet[Scalar[Key], Any]
        ),
    ) -> RecordSet[Rec_cov, FilteredIdx[Idx_def], B_nul, R_def, Rec2_nul, Rec3_def]: ...

    # 25. Slice selection
    @overload
    def __getitem__(
        self: RecordSet[Any], key: slice | tuple[slice, ...]
    ) -> RecordSet[Rec_cov, FilteredIdx[Idx_def], B_nul, R_def, Rec2_nul, Rec3_def]: ...

    # 26. Index value selection
    @overload
    def __getitem__(
        self: RecordSet[Record[Key2], Key | Index], key: Key | Key2
    ) -> RecordSet[Rec_cov, SingleIdx, B_nul, R_def, Rec2_nul, Rec3_def]: ...

    # Implementation:

    def __getitem__(  # noqa: D105
        self: RecordSet[Any, Any, Any, Any, Any, Any, Any],
        key: (
            type[Record]
            | ValueSet
            | RecordSet
            | RelTree
            | sqla.ColumnElement[bool]
            | list[Hashable]
            | slice
            | tuple[slice, ...]
            | Hashable
        ),
    ) -> (
        ValueSet[Any, Any, Any, Any, Any] | RecordSet[Any, Any, Any, Any, Any, Any, Any]
    ):
        merge = self._parse_merge_tree(key) if isinstance(key, RelTree) else None
        filt = key if isinstance(key, sqla.ColumnElement) else None

        keys = None
        if isinstance(key, tuple):
            # Selection by tuple of index values.
            keys = list(key)

        if isinstance(key, list | slice) and not isinstance(key, sqla.ColumnElement):
            # Selection by index value list, slice or single value.
            keys = [key]

        t = (
            RelSet
            if isinstance(key, RelSet)
            else ValueSet if isinstance(key, ValueSet) else RecordSet
        )

        return t(
            **{
                **asdict(self),
                **dict(
                    typedef=(
                        key
                        if isinstance(key, type) and self.prop is None
                        else self.typedef
                    ),
                    merges=(
                        self.merges * merge
                        if isinstance(merge, RelTree)
                        else self.merges
                    ),
                    filters=self.filters + [filt] if filt is not None else self.filters,
                    keys=keys if keys is not None else self.keys,
                ),
            }
        ).prefix(self)

    def select(
        self,
        *,
        index_only: bool = False,
    ) -> sqla.Select:
        """Return select statement for this dataset."""
        selection_table = self.base_table
        assert selection_table is not None

        select = sqla.select(
            *(
                col.label(f"{self.target_type._default_table_name()}.{col_name}")
                for col_name, col in selection_table.columns.items()
                if not index_only or self.target_type._attrs[col_name] in self.idx
            ),
            *(
                col.label(f"{rel.path_str}.{col_name}")
                for rel in self.merges.rels
                for col_name, col in self._get_alias(rel).columns.items()
                if not index_only or rel.target_type._attrs[col_name] in self.idx
            ),
        ).select_from(selection_table)

        for join in self._joins():
            select = select.join(*join)

        for filt in self.filters:
            select = select.where(filt)

        return select

    @overload
    def load(  # type: ignore
        self: RecordSet[Rec_cov, SingleIdx, Backend, Any, Any, Any, tuple[*RelTup]],
        kind: type[Record] = ...,
    ) -> tuple[Rec_cov, *RelTup]: ...

    @overload
    def load(  # type: ignore
        self: RecordSet[Rec_cov, SingleIdx, Backend],
        kind: type[Record] = ...,
    ) -> Rec_cov: ...

    @overload
    def load(
        self: RecordSet[Record, Any, Backend, Any, Any, Any, tuple[*RelTup]],
        kind: type[Df],
    ) -> tuple[Df, ...]: ...

    @overload
    def load(self: RecordSet[Record, Any, Backend], kind: type[Df]) -> Df: ...

    @overload
    def load(
        self: RecordSet[Record, Key, Backend, Any, Any, Any, tuple[*RelTup]],
        kind: type[Record] = ...,
    ) -> dict[Key, tuple[Rec_cov, *RelTup]]: ...

    @overload
    def load(
        self: RecordSet[Record[Key2], Any, Backend, Any, Any, Any, tuple[*RelTup]],
        kind: type[Record] = ...,
    ) -> dict[Key2, tuple[Rec_cov, *RelTup]]: ...

    @overload
    def load(
        self: RecordSet[Rec_cov, Key, Backend], kind: type[Record] = ...
    ) -> dict[Key, Rec_cov]: ...

    @overload
    def load(
        self: RecordSet[Record[Key2], Any, Backend], kind: type[Record] = ...
    ) -> dict[Key2, Rec_cov]: ...

    def load(
        self: RecordSet[Record, Any, Backend, Any, Any, Any, Any],
        kind: type[Record | Df] = Record,
    ) -> (
        Rec_cov
        | tuple[Rec_cov, *tuple[Any, ...]]
        | Df
        | tuple[Df, ...]
        | dict[Any, Rec_cov]
        | dict[Any, tuple[Rec_cov, *tuple[Any, ...]]]
    ):
        """Download selection."""
        select = self.select()

        idx_cols = [
            f"{rel.path_str}.{pk}"
            for rel in self.merges.rels
            for pk in rel.target_type._primary_keys
        ]

        main_cols = {
            col: col.lstrip(self.record_type._default_table_name() + ".")
            for col in select.columns.keys()
            if col.startswith(self.record_type._default_table_name())
        }

        extra_cols = {
            rel: {
                col: col.lstrip(rel.path_str + ".")
                for col in select.columns.keys()
                if col.startswith(rel.path_str)
            }
            for rel in self.merges.rels
            if rel.path_str is not None
        }

        merged_df = None
        if kind is pd.DataFrame:
            with self.engine.connect() as con:
                merged_df = pd.read_sql(select, con)
                merged_df = merged_df.set_index(idx_cols)
        else:
            merged_df = pl.read_database(select, self.engine)

        if issubclass(kind, Record):
            assert isinstance(merged_df, pl.DataFrame)

            rec_types = {
                self.record_type: main_cols,
                **{rel.target_type: cols for rel, cols in extra_cols.items()},
            }

            loaded: dict[type[Record], dict[Hashable, Record]] = {
                r: {} for r in rec_types
            }
            records: dict[Any, Record | tuple[Record, ...]] = {}

            for row in merged_df.iter_rows(named=True):
                idx = tuple(row[i] for i in idx_cols)
                idx = idx[0] if len(idx) == 1 else idx

                rec_list = []
                for rec_type, cols in rec_types.items():
                    rec_data: dict[Set, Any] = {
                        getattr(rec_type, attr): row[col] for col, attr in cols.items()
                    }
                    rec_idx = self.record_type._index_from_dict(rec_data)
                    rec = loaded[rec_type].get(rec_idx) or self.record_type(
                        _loader=self._load_prop,
                        **{p.name: v for p, v in rec_data.items()},  # type: ignore
                        **{r: Unloaded() for r in self.record_type._rels},
                    )

                    rec_list.append(rec)

                records[idx] = tuple(rec_list) if len(rec_list) > 1 else rec_list[0]

            return cast(
                dict[Any, Rec_cov] | dict[Any, tuple[Rec_cov, *tuple[Any, ...]]],
                records,
            )

        main_df, *extra_dfs = cast(
            tuple[Df, ...],
            (
                merged_df[list(main_cols.keys())].rename(main_cols),
                *(
                    merged_df[list(cols.keys())].rename(cols)
                    for cols in extra_cols.values()
                ),
            ),
        )

        return main_df, *extra_dfs

    def __iand__(
        self: RecordSet[Any, InsIdx, B_def, RWT],
        other: RecordSet[Rec_cov, InsIdx, B_def] | RecInput[Rec_cov, InsIdx],
    ) -> RecordSet[Any, InsIdx, B_def, RWT, Rec2_nul, Rec3_def, RM]:
        """Replacing assignment."""
        raise NotImplementedError("Replace not supported yet.")

    def __ior__(
        self: RecordSet[Any, InsIdx, B_def, RWT],
        other: RecordSet[Rec_cov, InsIdx, B_def] | RecInput[Rec_cov, InsIdx],
    ) -> RecordSet[Any, InsIdx, B_def, RWT, Rec2_nul, Rec3_def, RM]:
        """Upserting assignment."""
        raise NotImplementedError("Upsert not supported yet.")

    def __iadd__(
        self: RecordSet[Any, InsIdx, B_def, RWT],
        other: RecordSet[Rec_cov, InsIdx, B_def] | RecInput[Rec_cov, InsIdx],
    ) -> RecordSet[Any, InsIdx, B_def, RWT, Rec2_nul, Rec3_def, RM]:
        """Inserting assignment."""
        raise NotImplementedError("Insert not supported yet.")

    def __isub__(
        self: RecordSet[Any, InsIdx, B_def, RWT],
        other: RecordSet[Rec_cov, InsIdx, B_def] | Iterable[InsIdx] | InsIdx,
    ) -> RecordSet[Any, InsIdx, B_def, RWT, Rec2_nul, Rec3_def, RM]:
        """Deletion."""
        raise NotImplementedError("Delete not supported yet.")

    @overload
    def __lshift__(
        self: RecordSet[Any, Key, B_def, RWT],
        other: RecordSet[Rec_cov, Key, B_def] | RecInput[Rec_cov, Key],
    ) -> list[Key]: ...

    @overload
    def __lshift__(
        self: RecordSet[Record[Key2], Any, B_def, RWT],
        other: RecordSet[Rec_cov, Any, B_def] | RecInput[Rec_cov, Key2],
    ) -> list[Key2]: ...

    def __lshift__(
        self: RecordSet[Any, Any, B_def, RWT],
        other: RecordSet[Rec_cov, Any, B_def] | RecInput[Rec_cov, Any],
    ) -> list:
        """Injection."""
        raise NotImplementedError("Inject not supported yet.")

    @overload
    def __rshift__(
        self: RecordSet[Any, Key, B_def, RW], other: Key | Iterable[Key]
    ) -> dict[Key, Rec_cov]: ...

    @overload
    def __rshift__(
        self: RecordSet[Record[Key2], Any, B_def, RW], other: Key2 | Iterable[Key2]
    ) -> dict[Key2, Rec_cov]: ...

    def __rshift__(
        self: RecordSet[Any, Any, B_def, RW], other: Hashable | Iterable[Hashable]
    ) -> dict[Any, Rec_cov]:
        """Extraction."""
        raise NotImplementedError("Extract not supported yet.")

    # 1. Type deletion
    @overload
    def __delitem__(self: RecordSet[Rec, Any, Backend, RW], key: type[Rec]) -> None: ...

    # 2. List deletion
    @overload
    def __delitem__(
        self: RecordSet[
            Record[Key2],
            BaseIdx | FilteredIdx[BaseIdx] | Key | FilteredIdx[Key],
            Backend,
            RW,
        ],
        key: Iterable[Key | Key2],
    ) -> None: ...

    # 3. Index value deletion
    @overload
    def __delitem__(
        self: RecordSet[
            Record[Key2],
            BaseIdx | FilteredIdx[BaseIdx] | Key | FilteredIdx[Key],
            Backend,
            RW,
        ],
        key: Key | Key2,
    ) -> None: ...

    # 4. Slice deletion
    @overload
    def __delitem__(
        self: RecordSet[Any, Any, Backend, RW], key: slice | tuple[slice, ...]
    ) -> None: ...

    # 5. Expression filter deletion
    @overload
    def __delitem__(
        self: RecordSet[Any, Any, Backend, RW], key: sqla.ColumnElement[bool]
    ) -> None: ...

    # Implementation:

    def __delitem__(  # noqa: D105
        self: RecordSet[Any, Any, Backend, RW, Any, Any],
        key: (
            type[Record]
            | list[Hashable]
            | Hashable
            | slice
            | tuple[slice, ...]
            | sqla.ColumnElement[bool]
        ),
    ) -> None:
        if not isinstance(key, slice) or key != slice(None):
            del self[key][:]  # type: ignore

        select = self.select()

        tables = {self._get_table(rec) for rec in self.record_type._record_bases}

        statements = []

        for table in tables:
            # Prepare delete statement.
            if self.engine.dialect.name in (
                "postgres",
                "postgresql",
                "duckdb",
                "mysql",
                "mariadb",
            ):
                # Delete-from.
                statements.append(
                    table.delete().where(
                        reduce(
                            sqla.and_,
                            (
                                table.c[col.name] == select.c[col.name]
                                for col in table.primary_key.columns
                            ),
                        )
                    )
                )
            else:
                # Correlated update.
                raise NotImplementedError("Correlated update not supported yet.")

        # Execute delete statements.
        with self.engine.begin() as con:
            for statement in statements:
                con.execute(statement)

    def copy(
        self: RecordSet[Rec_cov, Idx_def, Backend],
        backend: B2_opt | None = None,
        overlay: str | bool = True,
    ) -> RecordSet[Rec_cov, Idx_def, B2_opt, R_def, Rec2_nul, Rec3_def, RM]:
        """Transfer the DB to a different backend (defaults to in-memory)."""
        other = RecordSet[Rec_cov, Idx_def, B2_opt, R_def, Rec2_nul, Rec3_def, RM](
            **asdict(self)
        )
        other.backend = backend if backend is not None else cast(B2_opt, self.backend)

        if other.backend is not None and other.backend != self.backend:
            for rec in self.schema_types:
                self[rec].load(kind=pl.DataFrame).write_database(
                    str(self._get_table(rec)), str(other.backend.url)
                )
        self.db

        return other

    def extract(
        self: RecordSet[Any, Any, B_def],
        use_schema: bool | type[Schema] = False,
        aggs: Mapping[RelSet, Agg] | None = None,
    ) -> DB[B_def]:
        """Extract a new database instance from the current selection."""
        assert issubclass(self.record_type, Record), "Record type must be defined."

        # Get all rec types in the schema.
        rec_types = (
            use_schema._record_types
            if isinstance(use_schema, type)
            else (
                self.schema_types
                if use_schema
                else ({self.record_type, *self.record_type._rel_types})
            )
        )

        # Get the entire subdag from this selection.
        all_paths_rels = {
            r
            for rel in self.record_type._rels.values()
            for r in rel._get_subdag(rec_types)
        }

        # Extract rel paths, which contain an aggregated rel.
        aggs_per_type: dict[type[Record], list[tuple[RelSet, Agg]]] = {}
        if aggs is not None:
            for rel, agg in aggs.items():
                for path_rel in all_paths_rels:
                    if rel in path_rel.rel_path:
                        aggs_per_type[rel.record_type] = [
                            *aggs_per_type.get(rel.record_type, []),
                            (rel, agg),
                        ]
                        all_paths_rels.remove(path_rel)

        replacements: dict[type[Record], sqla.Select] = {}
        for rec in rec_types:
            # For each table, create a union of all results from the direct routes.
            selects = [
                self[rel].select()
                for rel in all_paths_rels
                if issubclass(rec, rel.target_type)
            ]
            replacements[rec] = sqla.union(*selects).select()

        aggregations: dict[type[Record], sqla.Select] = {}
        for rec, rec_aggs in aggs_per_type.items():
            selects = []
            for rel, agg in rec_aggs:
                src_select = self[rel].select()
                selects.append(
                    sqla.select(
                        *[
                            (
                                src_select.c[sa.name]
                                if isinstance(sa, ValueSet)
                                else sqla_visitors.replacement_traverse(
                                    sa,
                                    {},
                                    replace=lambda element, **kw: (
                                        src_select.c[element.name]
                                        if isinstance(element, ValueSet)
                                        else None
                                    ),
                                )
                            ).label(ta.name)
                            for ta, sa in agg.map.items()
                        ]
                    )
                )

            aggregations[rec] = sqla.union(*selects).select()

        # Create a new database overlay for the results.
        new_db = DB(
            typedef=Record, backend=self.backend, overlay=f"temp_{token_hex(10)}"
        )

        # Overlay the new tables onto the new database.
        for rec in rec_types:
            if rec in replacements:
                new_db[rec] &= replacements[rec]

        for rec, agg_select in aggregations.items():
            new_db[rec] &= agg_select

        return new_db

    def __or__(
        self, other: RecordSet[Rec_cov, Idx, B_nul]
    ) -> RecordSet[Rec_cov, Idx_def | Idx]:
        """Union two datasets."""
        raise NotImplementedError("Union not supported yet.")

    def __len__(self: RecordSet[Any, Any, Backend]) -> int:
        """Return the number of records in the dataset."""
        with self.engine.connect() as conn:
            res = conn.execute(
                sqla.select(sqla.func.count()).select_from(self.select().subquery())
            ).scalar()
            assert isinstance(res, int)
            return res

    def __clause_element__(self) -> sqla.Subquery:
        """Return subquery for the current selection to be used inside SQL clauses."""
        return self.select().subquery()

    def _joins(self, _subtree: DataDict | None = None) -> list[Join]:
        """Extract join operations from the relation tree."""
        if self.base_table is None:
            return []

        joins = []
        _subtree = _subtree or self.merges.dict

        for rel, next_subtree in _subtree.items():
            parent = (
                self._get_alias(rel.parent)
                if isinstance(rel.parent, RelSet)
                else self.base_table
            )

            temp_alias_map = {
                rec: self._get_random_alias(rec) for rec in rel.inter_joins.keys()
            }

            joins.extend(
                (
                    temp_alias_map[rec],
                    reduce(
                        sqla.or_,
                        (
                            reduce(
                                sqla.and_,
                                (
                                    parent.c[lk.name] == temp_alias_map[rec].c[rk.name]
                                    for lk, rk in join_on.items()
                                ),
                            )
                            for join_on in joins
                        ),
                    ),
                )
                for rec, joins in rel.inter_joins.items()
            )

            target_table = self._get_alias(rel)

            joins.append(
                (
                    target_table,
                    reduce(
                        sqla.or_,
                        (
                            reduce(
                                sqla.and_,
                                (
                                    temp_alias_map[lk.record_type].c[lk.name]
                                    == target_table.c[rk.name]
                                    for lk, rk in join_on.items()
                                ),
                            )
                            for join_on in rel.target_joins
                        ),
                    ),
                )
            )

            joins.extend(self._joins(next_subtree))

        return joins

    def _get_subdag(
        self,
        backlink_records: set[type[Record]] | None = None,
        _traversed: set[RelSet] | None = None,
    ) -> set[RelSet]:
        """Find all paths to the target record type."""
        assert issubclass(self.record_type, Record)

        backlink_records = backlink_records or set()
        _traversed = _traversed or set()

        # Get relations of the target type as next relations
        next_rels = set(self.record_type._rels.values())

        for backlink_record in backlink_records:
            next_rels |= backlink_record._backrels_to_rels(self.record_type)

        # Filter out already traversed relations
        next_rels = {rel for rel in next_rels if rel not in _traversed}

        # Add next relations to traversed set
        _traversed |= next_rels

        # Prefix next relations with current relation
        prefixed_rels = {rel.prefix(self) for rel in next_rels}

        # Return next relations + recurse
        return prefixed_rels | {
            rel
            for next_rel in next_rels
            for rel in next_rel._get_subdag(backlink_records, _traversed)
        }

    def _set(  # noqa: C901, D105
        self: RecordSet[Rec_cov, Any, Backend, RW],
        value: RecordSet | PartialRecInput[Rec_cov, Any] | ValInput,
        mode: Literal["update", "upsert", "replace"] = "update",
        covered: set[int] | None = None,
    ) -> None:
        assert issubclass(self.record_type, Record), "Record type must be defined."

        covered = covered or set()

        record_data: dict[Any, dict[Set, Any]] | None = None
        rel_data: dict[RelSet, dict[Any, Record]] | None = None
        df_data: DataFrame | None = None
        partial: bool = False

        list_idx = (
            self.path_idx is not None
            and len(self.path_idx) == 1
            and issubclass(self.path_idx[0].prop_type.value_type(), int)
        )

        if isinstance(value, Record):
            record_data = {value._index: value._to_dict()}
        elif isinstance(value, Mapping):
            if has_type(value, Mapping[Set, Any]):
                record_data = {
                    self.record_type._index_from_dict(value): {
                        p: v for p, v in value.items()
                    }
                }
                partial = True
            else:
                assert has_type(value, Mapping[Any, PartialRec[Record]])

                record_data = {}
                for idx, rec in value.items():
                    if isinstance(rec, Record):
                        rec_dict = rec._to_dict()
                    else:
                        rec_dict = {p: v for p, v in rec.items()}
                        partial = True
                    record_data[idx] = rec_dict

        elif isinstance(value, DataFrame):
            df_data = value
        elif isinstance(value, Series):
            df_data = value.rename("_value").to_frame()
        elif isinstance(value, Iterable):
            record_data = {}
            for idx, rec in enumerate(value):
                if isinstance(rec, Record):
                    rec_dict = rec._to_dict()
                    rec_idx = rec._index
                else:
                    rec_dict = {p: v for p, v in rec.items()}
                    rec_idx = self.record_type._index_from_dict(rec)
                    partial = True

                record_data[idx if list_idx else rec_idx] = rec_dict

        assert not (
            mode in ("upsert", "replace") and partial
        ), "Partial record input requires update mode."

        if record_data is not None:
            # Split record data into attribute and relation data.
            attr_data = {
                idx: {p.name: v for p, v in rec.items() if isinstance(p, ValueSet)}
                for idx, rec in record_data.items()
            }
            rel_data = self._get_record_rels(record_data, list_idx, covered)

            # Transform attribute data into DataFrame.
            df_data = pd.DataFrame.from_records(attr_data)

        if rel_data is not None:
            # Recurse into relation data.
            for r, r_data in rel_data.items():
                self[r]._set(r_data, mode="replace", covered=covered)

        # Load data into a temporary table.
        if df_data is not None:
            value_set = self.db.dataset(df_data)
        elif isinstance(value, sqla.Select):
            value_set = self.db.dataset(value)
        else:
            assert isinstance(value, RecordSet)
            value_set = value

        attrs_by_table = {
            self._get_table(rec): {
                a for a in self.record_type._attrs.values() if a.record_type is rec
            }
            for rec in self.record_type._record_bases
        }

        statements = []

        if mode == "replace":
            # Delete all records in the current selection.
            select = self.select()

            for table in attrs_by_table:
                # Prepare delete statement.
                if self.engine.dialect.name in (
                    "postgres",
                    "postgresql",
                    "duckdb",
                    "mysql",
                    "mariadb",
                ):
                    # Delete-from.
                    statements.append(
                        table.delete().where(
                            reduce(
                                sqla.and_,
                                (
                                    table.c[col.name] == select.c[col.name]
                                    for col in table.primary_key.columns
                                ),
                            )
                        )
                    )
                else:
                    # Correlated update.
                    raise NotImplementedError("Correlated update not supported yet.")

        if mode in ("replace", "upsert"):
            # Construct the insert statements.

            assert len(self.filters) == 0, "Can only upsert into unfiltered datasets."

            for table, attrs in attrs_by_table.items():
                # Do an insert-from-select operation, which updates on conflict:
                statement = table.insert().from_select(
                    [a.name for a in attrs],
                    value_set.select().subquery(),
                )

                if isinstance(statement, postgresql.Insert):
                    # For Postgres / DuckDB, use: https://docs.sqlalchemy.org/en/20/dialects/postgresql.html#updating-using-the-excluded-insert-values
                    statement = statement.on_conflict_do_update(
                        index_elements=[col.name for col in table.primary_key.columns],
                        set_=dict(statement.excluded),
                    )
                elif isinstance(statement, mysql.Insert):
                    # For MySQL / MariaDB, use: https://docs.sqlalchemy.org/en/20/dialects/mysql.html#insert-on-duplicate-key-update-upsert
                    statement = statement.prefix_with("INSERT INTO")
                    statement = statement.on_duplicate_key_update(**statement.inserted)
                else:
                    # For others, use CTE: https://docs.sqlalchemy.org/en/20/core/selectable.html#sqlalchemy.sql.expression.Select.cte
                    raise NotImplementedError(
                        "Upsert not supported for this database dialect."
                    )

                statements.append(statement)
        else:
            # Construct the update statements.

            assert isinstance(self.base_table, sqla.Table), "Base must be a table."

            # Derive current select statement and join with value table, if exists.
            select = self.select().join(
                value_set.base_table,
                reduce(
                    sqla.and_,
                    (
                        self.base_table.c[idx_col.name] == idx_col
                        for idx_col in value_set.base_table.primary_key
                    ),
                ),
            )

            for table, attrs in attrs_by_table.items():
                attr_names = {a.name for a in attrs}
                values = {
                    c_name: c
                    for c_name, c in value_set.base_table.columns.items()
                    if c_name in attr_names
                }

                # Prepare update statement.
                if self.engine.dialect.name in (
                    "postgres",
                    "postgresql",
                    "duckdb",
                    "mysql",
                    "mariadb",
                ):
                    # Update-from.
                    statements.append(
                        table.update()
                        .values(values)
                        .where(
                            reduce(
                                sqla.and_,
                                (
                                    table.c[col.name] == select.c[col.name]
                                    for col in table.primary_key.columns
                                ),
                            )
                        )
                    )
                else:
                    # Correlated update.
                    raise NotImplementedError("Correlated update not supported yet.")

        # Execute delete / insert / update statements.
        with self.engine.begin() as con:
            for statement in statements:
                con.execute(statement)

        # Drop the temporary table, if any.
        if value_set is not None:
            cast(sqla.Table, value_set.base_table).drop(self.engine)

        if mode in ("replace", "upsert") and isinstance(self, RelSet):
            # Update incoming relations from parent records.
            if self.direct_rel is not None:
                if issubclass(self.fk_record_type, self.record_type):
                    # Case: parent links directly to child (n -> 1)
                    for fk, pk in self.direct_rel.fk_map.items():
                        self.parent_set[fk] @= value_set[pk]
                else:
                    # Case: parent and child are linked via assoc table (n <--> m)
                    # Update link table with new child indexes.
                    assert self.link_set is not None
                    self.link_set &= value_set.select()

            # Note that the (1 <- n) case is already covered by updating
            # the child record directly, which includes all its foreign keys.


@dataclass(unsafe_hash=True)
class RelSet(RecordSet[Rec_def, Idx_def, B_nul, R_def, Rec2_def, Rec3_def, RM]):
    """Relation set class."""

    prop: Rel[Rec2_def, Rec3_def, R_def, Rec_def]  # type: ignore
    link_type: type[Rec3_def] = Record

    @cached_property
    def parent(self) -> RelSet | type[Record]:
        """Parent relation of this Rel."""
        return self.get_tag(self.record_type) or self.record_type

    @cached_property
    def rel_path(self) -> list[RelSet]:
        """Path from base record type to this Rel."""
        return [
            *(self.parent.rel_path if isinstance(self.parent, RelSet) else []),
            self,
        ]

    @cached_property
    def parent_set(self) -> RecordSet[Rec2_def, Any, B_nul, R_def]:
        """Parent set of this Rel."""
        t = RelSet if len(self.rel_path) > 1 else RecordSet

        arg_dict = asdict(self)
        del arg_dict["link_type"]

        if len(self.rel_path) > 1:
            parent_rel = self.rel_path[-2]
            arg_dict["prop"] = parent_rel.prop
            arg_dict["record_type"] = parent_rel.record_type

        return cast(RecordSet[Rec2_def, Any, B_nul, R_def], t(**arg_dict))

    @cached_property
    def link_set(
        self,
    ) -> RecordSet[Rec3_def, Any, B_nul, R_def, Rec2_def] | None:
        """Get the link set."""
        assert isinstance(self, RelSet)

        if (
            self.link_type is None
            or self.direct_rel is None
            or not issubclass(self.direct_rel.record_type, self.link_type)
        ):
            return None

        return self.parent_set[self.record_type._rel(self.link_type)]

    @cached_property
    def link(self) -> type[Rec3_def]:
        """Reference props of the link record type."""
        return (
            self.link_set.rec
            if self.link_set is not None
            else cast(type[Rec3_def], Record)
        )

    @cached_property
    def fk_record_type(self) -> type[Record]:
        """Record type of the foreign key."""
        match self.prop.on:
            case type():
                return self.prop.on
            case RelSet():
                return self.prop.on.record_type
            case tuple():
                link = self.prop.on[0]
                assert isinstance(link, RecordSet)
                return link.record_type
            case dict() | ValueSet() | Iterable() | None:
                return self.record_type

    @cached_property
    def direct_rel(
        self,
    ) -> RelSet[Rec2_def, SingleIdx, B_nul, R_def, Rec_def | Rec3_def] | None:
        """Direct rel."""
        match self.prop.on:
            case type():
                rels = [
                    r
                    for r in self.prop.on._rels.values()
                    if isinstance(r.target_type, self.record_type)
                    and r.direct_rel is None
                ]
                assert len(rels) == 1, "Direct relation must be unique."
                return cast(
                    RelSet[Rec2_def, SingleIdx, B_nul, R_def, Rec_def | Rec3_def],
                    rels[0],
                )
            case RelSet():
                return cast(
                    RelSet[Rec2_def, SingleIdx, B_nul, R_def, Rec_def | Rec3_def],
                    self.prop.on,
                )
            case tuple():
                link = self.prop.on[0]
                assert isinstance(link, RelSet)
                return cast(
                    RelSet[Rec2_def, SingleIdx, B_nul, R_def, Rec_def | Rec3_def], link
                )
            case dict() | ValueSet() | Iterable() | None:
                return None

    @cached_property
    def counter_rel(self) -> RelSet[Rec2_def, Any, B_nul, R_def, Rec_def]:
        """Counter rel."""
        if self.direct_rel is not None and issubclass(
            self.direct_rel.record_type, self.target_type
        ):
            return cast(RelSet[Rec2_def, Any, B_nul, R_def, Rec_def], self.direct_rel)

        return cast(
            RelSet[Rec2_def, Any, B_nul, R_def, Rec_def],
            self.target_type._rel(self.record_type),
        )

    @cached_property
    def fk_map(self) -> bidict[ValueSet, ValueSet]:
        """Map source foreign keys to target attrs."""
        target = self.target_type

        match self.prop.on:
            case type() | RecordSet() | tuple():
                return bidict()
            case dict():
                return bidict(
                    {
                        ValueSet(
                            prop=Attr(_name=fk.name),
                            record_type=self.record_type,
                            typedef=fk.prop_type,
                        ): pk
                        for fk, pk in self.prop.on.items()
                    }
                )
            case ValueSet() | Iterable():
                attrs = (
                    self.prop.on
                    if isinstance(self.prop.on, Iterable)
                    else [self.prop.on]
                )
                source_attrs: list[ValueSet] = [
                    ValueSet(
                        prop=Attr(_name=attr.name),
                        record_type=self.record_type,
                        typedef=attr.prop_type,
                    )
                    for attr in attrs
                ]
                target_attrs = target._primary_keys.values()
                fk_map = dict(zip(source_attrs, target_attrs))

                assert all(
                    issubclass(
                        self.record_type._static_props[
                            fk_attr.name
                        ].typedef.value_type(),
                        pk_attr.typedef.value_type(),
                    )
                    for fk_attr, pk_attr in fk_map.items()
                    if pk_attr.typedef is not None
                ), "Foreign key value types must match primary key value types."

                return bidict(fk_map)
            case None:
                return bidict(
                    {
                        ValueSet(
                            prop=Attr(_name=f"{self.prop.name}_{target_attr.name}"),
                            record_type=self.record_type,
                            typedef=target_attr.prop_type,
                        ): target_attr
                        for target_attr in target._primary_keys.values()
                    }
                )

    @cached_property
    def inter_joins(
        self,
    ) -> dict[type[Record], list[Mapping[ValueSet, ValueSet]]]:
        """Intermediate joins required by this rel."""
        match self.prop.on:
            case RecordSet():
                # Relation is defined via other, backlinking relation
                other_rel = self.prop.on
                assert isinstance(
                    other_rel, RecordSet
                ), "Back-reference must be an explicit relation"

                if issubclass(other_rel.target_type, self.record_type):
                    # Supplied record type object is a backlinking relation
                    return {}
                else:
                    # Supplied record type object is a forward relation
                    # on a relation table
                    back_rels = [
                        rel
                        for rel in other_rel.record_type._rels.values()
                        if issubclass(rel.target_type, self.record_type)
                        and len(rel.fk_map) > 0
                    ]

                    return {
                        other_rel.record_type: [
                            back_rel.fk_map.inverse for back_rel in back_rels
                        ]
                    }
            case type():
                if issubclass(self.prop.on, self.target_type):
                    # Relation is defined via all direct backlinks of given record type.
                    return {}

                # Relation is defined via relation table
                back_rels = [
                    rel
                    for rel in self.prop.on._rels.values()
                    if issubclass(rel.target_type, self.record_type)
                    and len(rel.fk_map) > 0
                ]

                return {
                    self.prop.on: [back_rel.fk_map.inverse for back_rel in back_rels]
                }
            case tuple() if has_type(self.prop.on, tuple[RelSet, RelSet]):
                # Relation is defined via back-rel + forward-rel
                # on a relation table.
                back, _ = self.prop.on
                assert len(back.fk_map) > 0, "Back relation must be direct."
                assert issubclass(back.record_type, Record)
                return {back.record_type: [back.fk_map.inverse]}
            case _:
                # Relation is defined via foreign key attributes
                return {}

    @cached_property
    def target_joins(self) -> list[Mapping[ValueSet, ValueSet]]:
        """Mappings of column keys to join the target on."""
        match self.prop.on:
            case RecordSet():
                # Relation is defined via other relation or relation table
                other_rel = self.prop.on
                assert (
                    len(other_rel.fk_map) > 0
                ), "Backref or forward-ref on relation table must be a direct relation"
                return [
                    (
                        other_rel.fk_map.inverse
                        if issubclass(other_rel.target_type, self.record_type)
                        else other_rel.fk_map
                    )
                ]

            case type():
                if issubclass(self.prop.on, self.target_type):
                    # Relation is defined via all direct backlinks of given record type.
                    back_rels = [
                        rel
                        for rel in self.prop.on._rels.values()
                        if issubclass(rel.target_type, self.record_type)
                        and len(rel.fk_map) > 0
                    ]
                    assert len(back_rels) > 0, "No direct backlinks found."
                    return [back_rel.fk_map.inverse for back_rel in back_rels]

                # Relation is defined via relation table
                fwd_rels = [
                    rel
                    for rel in self.prop.on._rels.values()
                    if issubclass(rel.target_type, self.record_type)
                    and len(rel.fk_map) > 0
                ]
                assert (
                    len(fwd_rels) > 0
                ), "No direct forward rels found on relation table."
                return [fwd_rel.fk_map for fwd_rel in fwd_rels]

            case tuple():
                assert has_type(self.prop.on, tuple[RelSet, RelSet])
                # Relation is defined via back-rel + forward-rel
                # on a relation table.
                _, fwd = self.prop.on
                assert len(fwd.fk_map) > 0, "Forward relation must be direct."
                return [fwd.fk_map]

            case _:
                # Relation is defined via foreign key attributes
                return [self.fk_map]


@dataclass(unsafe_hash=True)
class DB(RecordSet[Record, BaseIdx, B_def, RWT, Record, Record, None]):
    """Database class."""

    schema: (
        type[Schema]
        | Mapping[
            type[Schema],
            Require | str,
        ]
        | None
    ) = None
    records: (
        Mapping[
            type[Record],
            Require | str | sqla.TableClause,
        ]
        | None
    ) = None

    validate_on_init: bool = True

    def __post_init__(self):  # noqa: D105
        if self.records is not None:
            self.subs = {
                **self.subs,
                **{
                    rec: (sub if isinstance(sub, sqla.TableClause) else sqla.table(sub))
                    for rec, sub in self.records.items()
                    if not isinstance(sub, Require)
                },
            }
            self.schema_types |= set(self.records.keys())

        if self.schema is not None:
            if isinstance(self.schema, Mapping):
                self.subs = {
                    **self.subs,
                    **{
                        rec: sqla.table(rec._table_name, schema=schema_name)
                        for schema, schema_name in self.schema.items()
                        for rec in schema._record_types
                    },
                }
                schemas = (
                    set(self.schema.keys())
                    if isinstance(self.schema, Mapping)
                    else set([self.schema])
                )
            else:
                schemas = {self.schema}

            self.schema_types |= {
                rec for schema in schemas for rec in schema._record_types
            }

        if self.validate_on_init:
            self.validate()

        if self.overlay is not None and self.overlay_with_schemas:
            self._ensure_schema_exists(self.overlay)

    @cached_property
    def assoc_types(self) -> set[type[Record]]:
        """Set of all association tables in this DB."""
        assoc_types = set()
        for rec in self.schema_types:
            pks = set([attr.name for attr in rec._primary_keys.values()])
            fks = set(
                [attr.name for rel in rec._rels.values() for attr in rel.fk_map.keys()]
            )
            if pks == fks:
                assoc_types.add(rec)

        return assoc_types

    @cached_property
    def relation_map(self) -> dict[type[Record], set[RelSet]]:
        """Maps all tables in this DB to their outgoing or incoming relations."""
        rels: dict[type[Record], set[RelSet]] = {
            table: set() for table in self.schema_types
        }

        for rec in self.schema_types:
            for rel in rec._rels.values():
                rels[rec].add(rel)
                rels[rel.target_type].add(rel)

        return rels

    def describe(self) -> dict[str, str | dict[str, str] | None]:
        """Return a description of this database.

        Returns:
            Mapping of table names to table descriptions.
        """
        schema_desc = {}
        if self.schema is not None:
            schema_ref = PyObjectRef.reference(self.schema)

            schema_desc = {
                "schema": {
                    "repo": schema_ref.repo,
                    "package": schema_ref.package,
                    "class": f"{schema_ref.module}.{schema_ref.object}",
                }
            }

            if schema_ref.object_version is not None:
                schema_desc["schema"]["version"] = schema_ref.object_version
            elif schema_ref.package_version is not None:
                schema_desc["schema"]["version"] = schema_ref.package_version

            if schema_ref.repo_revision is not None:
                schema_desc["schema"]["revision"] = schema_ref.repo_revision

            if schema_ref.docs_url is not None:
                schema_desc["schema"]["docs"] = schema_ref.docs_url

        return {
            **schema_desc,
            "backend": (
                asdict(self.backend)
                if self.backend is not None and self.backend.type != "in-memory"
                else None
            ),
        }

    def execute[
        *T
    ](
        self, stmt: sqla.Select[tuple[*T]] | sqla.Insert | sqla.Update | sqla.Delete
    ) -> sqla.Result[tuple[*T]]:
        """Execute a SQL statement in this database's context."""
        stmt = self._parse_expr(stmt)
        with self.engine.begin() as conn:
            return conn.execute(self._parse_expr(stmt))

    def dataset(
        self: RecordSet[Any, Any, B_def, RW],
        data: DataFrame | sqla.Select,
        foreign_keys: Mapping[str, ValueSet] | None = None,
    ) -> RecordSet[DynRecord, Any, B_def, RO]:
        """Create a temporary dataset instance from a DataFrame or SQL query."""
        name = (
            f"temp_df_{gen_str_hash(data, 10)}"
            if isinstance(data, DataFrame)
            else f"temp_{token_hex(5)}"
        )

        rec = dynamic_record_type(name, props=props_from_data(data, foreign_keys))
        self._get_table(rec, writable=True)
        ds: RecordSet[DynRecord, BaseIdx, B_def, RW] = RecordSet(
            typedef=rec,
            backend=self.backend,
            overlay=self.overlay,
            subs=self.subs,
        )
        ds &= data

        return ds  # type: ignore

    def to_graph(
        self, nodes: Sequence[type[Rec_cov]]
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """Export links between select database objects in a graph format.

        E.g. for usage with `Gephi`_

        .. _Gephi: https://gephi.org/
        """
        node_tables = [self[n] for n in nodes]

        # Concat all node tables into one.
        node_dfs = [
            n.load(kind=pd.DataFrame)
            .reset_index()
            .assign(table=n.target_type._default_table_name())
            for n in node_tables
        ]
        node_df = (
            pd.concat(node_dfs, ignore_index=True)
            .reset_index()
            .rename(columns={"index": "node_id"})
        )

        directed_edges = reduce(set.union, (self.relation_map[n] for n in nodes))

        undirected_edges: dict[type[Record], set[tuple[RelSet, ...]]] = {
            t: set() for t in nodes
        }
        for n in nodes:
            for at in self.assoc_types:
                if len(at._rels) == 2:
                    left, right = (r for r in at._rels.values())
                    assert left is not None and right is not None
                    if left.target_type == n:
                        undirected_edges[n].add((left, right))
                    elif right.target_type == n:
                        undirected_edges[n].add((right, left))

        # Concat all edges into one table.
        edge_df = pd.concat(
            [
                *[
                    node_df.loc[
                        node_df["table"]
                        == str((rel.record_type or Record)._default_table_name())
                    ]
                    .rename(columns={"node_id": "source"})
                    .merge(
                        node_df.loc[
                            node_df["table"]
                            == str(rel.target_type._default_table_name())
                        ],
                        left_on=[c.name for c in rel.fk_map.keys()],
                        right_on=[c.name for c in rel.fk_map.values()],
                    )
                    .rename(columns={"node_id": "target"})[["source", "target"]]
                    .assign(ltr=",".join(c.name for c in rel.fk_map.keys()), rtl=None)
                    for rel in directed_edges
                ],
                *[
                    self[assoc_table]
                    .load(kind=pd.DataFrame)
                    .merge(
                        node_df.loc[
                            node_df["table"]
                            == str(left_rel.target_type._default_table_name())
                        ].dropna(axis="columns", how="all"),
                        left_on=[c.name for c in left_rel.fk_map.keys()],
                        right_on=[c.name for c in left_rel.fk_map.values()],
                        how="inner",
                    )
                    .rename(columns={"node_id": "source"})
                    .merge(
                        node_df.loc[
                            node_df["table"]
                            == str(left_rel.target_type._default_table_name())
                        ].dropna(axis="columns", how="all"),
                        left_on=[c.name for c in right_rel.fk_map.keys()],
                        right_on=[c.name for c in right_rel.fk_map.values()],
                        how="inner",
                    )
                    .rename(columns={"node_id": "target"})[
                        list(
                            {
                                "source",
                                "target",
                                *(a for a in (left_rel.record_type or Record)._attrs),
                            }
                        )
                    ]
                    .assign(
                        ltr=",".join(c.name for c in right_rel.fk_map.keys()),
                        rtl=",".join(c.name for c in left_rel.fk_map.keys()),
                    )
                    for assoc_table, rels in undirected_edges.items()
                    for left_rel, right_rel in rels
                ],
            ],
            ignore_index=True,
        )

        return node_df, edge_df

    def validate(self) -> None:
        """Perform pre-defined schema validations."""
        types = {}

        if self.records is not None:
            types |= {
                rec: isinstance(req, Require) and req.present
                for rec, req in self.records.items()
            }

        if isinstance(self.schema, Mapping):
            types |= {
                rec: isinstance(req, Require) and req.present
                for schema, req in self.schema.items()
                for rec in schema._record_types
            }

        tables = {self._get_table(rec): required for rec, required in types.items()}

        inspector = sqla.inspect(self.engine)

        # Iterate over all tables and perform validations for each
        for table, required in tables.items():
            has_table = inspector.has_table(table.name, table.schema)

            if not has_table and not required:
                continue

            # Check if table exists
            assert has_table

            db_columns = {
                c["name"]: c for c in inspector.get_columns(table.name, table.schema)
            }
            for column in table.columns:
                # Check if column exists
                assert column.name in db_columns

                db_col = db_columns[column.name]

                # Check if column type and nullability match
                assert isinstance(db_col["type"], type(column.type))
                assert db_col["nullable"] == column.nullable or column.nullable is None

            # Check if primary key is compatible
            db_pk = inspector.get_pk_constraint(table.name, table.schema)
            if len(db_pk["constrained_columns"]) > 0:  # Allow source tbales without pk
                assert set(db_pk["constrained_columns"]) == set(
                    table.primary_key.columns.keys()
                )

            # Check if foreign keys are compatible
            db_fks = inspector.get_foreign_keys(table.name, table.schema)
            for fk in table.foreign_key_constraints:
                matches = [
                    (
                        set(db_fk["constrained_columns"]) == set(fk.column_keys),
                        (
                            db_fk["referred_table"].lower()
                            == fk.referred_table.name.lower()
                        ),
                        set(db_fk["referred_columns"])
                        == set(f.column.name for f in fk.elements),
                    )
                    for db_fk in db_fks
                ]

                assert any(all(m) for m in matches)

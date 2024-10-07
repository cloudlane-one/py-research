"""Static schemas for universal relational databases."""

from __future__ import annotations

from abc import ABC
from collections.abc import (
    Callable,
    Generator,
    Hashable,
    Iterable,
    Iterator,
    Mapping,
    Sequence,
)
from dataclasses import Field, dataclass, field
from datetime import date, datetime, time
from enum import Enum, auto
from functools import cached_property, partial, reduce
from inspect import get_annotations, getmodule
from io import BytesIO
from itertools import groupby
from pathlib import Path
from secrets import token_hex
from types import GenericAlias, ModuleType, NoneType, UnionType, new_class
from typing import (
    TYPE_CHECKING,
    Any,
    ClassVar,
    ForwardRef,
    Generic,
    Literal,
    LiteralString,
    ParamSpec,
    Self,
    TypeVar,
    TypeVarTuple,
    cast,
    dataclass_transform,
    get_args,
    get_origin,
    overload,
)
from uuid import UUID, uuid4

import openpyxl
import pandas as pd
import polars as pl
import sqlalchemy as sqla
import sqlalchemy.dialects.mysql as mysql
import sqlalchemy.dialects.postgresql as postgresql
import sqlalchemy.orm as orm
import sqlalchemy.sql.type_api as sqla_types
import sqlalchemy.sql.visitors as sqla_visitors
import yarl
from bidict import bidict
from cloudpathlib import CloudPath
from pandas.api.types import (
    is_bool_dtype,
    is_datetime64_dtype,
    is_integer_dtype,
    is_numeric_dtype,
    is_string_dtype,
)
from sqlalchemy_utils import UUIDType
from typing_extensions import TypeVar as TypeVarX
from xlsxwriter import Workbook as ExcelWorkbook

from py_research.data import copy_and_override
from py_research.enums import StrEnum
from py_research.files import HttpFile
from py_research.hashing import gen_int_hash, gen_str_hash
from py_research.reflect.ref import PyObjectRef
from py_research.reflect.runtime import get_subclasses
from py_research.reflect.types import (
    SingleTypeDef,
    extract_nullable_type,
    get_lowest_common_base,
    has_type,
    is_subtype,
)

VarT = TypeVar("VarT", covariant=True)
VarTi = TypeVarX("VarTi", default=Any)
VarT2 = TypeVar("VarT2")
VarT3 = TypeVar("VarT3")
VarT4 = TypeVar("VarT4")

RwT = TypeVarX("RwT", bound="R", default="RW", covariant=True)

KeyT = TypeVarX("KeyT", bound=Hashable, default=Any)
KeyT2 = TypeVar("KeyT2", bound=Hashable)
KeyT3 = TypeVar("KeyT3", bound=Hashable)

IdxT = TypeVarX(
    "IdxT",
    covariant=True,
    bound="Index",
    default="BaseIdx",
)
IdxTt = TypeVarTuple("IdxTt")

SelT = TypeVarX("SelT", covariant=True, bound="Record | tuple | None", default=None)

MdxT = TypeVar("MdxT", bound="MutateIdx", covariant=True)
MdxT2 = TypeVar("MdxT2", bound="MutateIdx")

DfT = TypeVar("DfT", bound=pd.DataFrame | pl.DataFrame)

CdxT = TypeVarX("CdxT", bound="ColIdx", covariant=True, default=Hashable)

Params = ParamSpec("Params")

type LocalDyn = Literal[LocalBackend.dynamic]
type LocalStat = Literal[LocalBackend.static]
type DynBackendID = LiteralString | LocalDyn
type StatBackendID = DynBackendID | LocalStat


type Index = Hashable | BaseIdx | SingleIdx | Filt
type ColIdx = Hashable | SingleIdx
type MutateIdx = Hashable | BaseIdx

type IdxStart[Key: Hashable] = Key | tuple[Key, *tuple[Any, ...]]
type IdxEnd[Key: Hashable] = tuple[*tuple[Any, ...], Key]
type IdxStartEnd[Key: Hashable, Key2: Hashable] = tuple[Key, *tuple[Any, ...], Key2]
type IdxTupStart[*IdxTup] = tuple[*IdxTup, *tuple[Any]]
type IdxTupStartEnd[*IdxTup, Key2: Hashable] = tuple[*IdxTup, *tuple[Any], Key2]

type RecordValue[Rec: Record] = Rec | Iterable[Rec] | Mapping[Any, Rec] | None


class State(Enum):
    """Demark load status."""

    undef = auto()


class BaseIdx:
    """Singleton to mark dataset index as default index."""

    __hash__: ClassVar[None]  # type: ignore[assignment]


class SingleIdx:
    """Singleton to mark dataset index as a single value."""

    __hash__: ClassVar[None]  # type: ignore[assignment]


class Filt(Generic[MdxT]):
    """Singleton to mark dataset index as filtered."""

    __hash__: ClassVar[None]  # type: ignore[assignment]


class R:
    """Read-only flag."""


class RO(R):
    """Read-only flag."""


class RW(RO):
    """Read-write flag."""


type RecInput[Rec: Record, Key: MutateIdx] = pd.DataFrame | pl.DataFrame | Iterable[
    Rec
] | Mapping[Key, Rec] | sqla.Select | Rec


type ValInput[Val, Key: Hashable] = pd.Series | pl.Series | Mapping[
    Key, Val
] | sqla.Select[tuple[Val] | tuple[Key, Val]] | Val


def map_df_dtype(c: pd.Series | pl.Series) -> type | None:
    """Map pandas dtype to Python type."""
    if isinstance(c, pd.Series):
        if is_datetime64_dtype(c):
            return datetime
        elif is_bool_dtype(c):
            return bool
        elif is_integer_dtype(c):
            return int
        elif is_numeric_dtype(c):
            return float
        elif is_string_dtype(c):
            return str
    else:
        if c.dtype.is_temporal():
            return datetime
        elif c.dtype.is_integer():
            return int
        elif c.dtype.is_float():
            return float
        elif c.dtype.is_(pl.String):
            return str

    return None


def map_col_dtype(c: sqla.ColumnElement) -> type | None:
    """Map sqla column type to Python type."""
    match c.type:
        case sqla.DateTime():
            return datetime
        case sqla.Date():
            return date
        case sqla.Time():
            return time
        case sqla.Boolean():
            return bool
        case sqla.Integer():
            return int
        case sqla.Float():
            return float
        case sqla.String() | sqla.Text() | sqla.Enum():
            return str
        case sqla.LargeBinary():
            return bytes
        case _:
            return None


def props_from_data(
    data: pd.DataFrame | pl.DataFrame | sqla.Select,
    foreign_keys: Mapping[str, Col] | None = None,
) -> list[Prop]:
    """Extract prop definitions from dataframe or query."""
    foreign_keys = foreign_keys or {}

    def _gen_prop(
        name: str, data: pd.Series | pl.Series | sqla.ColumnElement
    ) -> Prop[Any, Any]:
        is_rel = name in foreign_keys
        value_type = (
            map_df_dtype(data)
            if isinstance(data, pd.Series | pl.Series)
            else map_col_dtype(data)
        ) or Any
        col = Col[value_type](
            primary_key=True,
            _name=name if not is_rel else f"fk_{name}",
            _type=PropType(Col[value_type]),
            _parent_type=DynRecord,
        )
        return (
            col
            if not is_rel
            else RelSet(
                on={col: foreign_keys[name]},
                _type=PropType(RelSet[cast(type, foreign_keys[name].value_type)]),
                _name=f"rel_{name}",
                _parent_type=DynRecord,
            )
        )

    columns = (
        [data[col] for col in data.columns]
        if isinstance(data, pd.DataFrame | pl.DataFrame)
        else list(data.columns)
    )

    index_props = []
    if isinstance(data, pd.DataFrame):
        levels = {name: name for name in data.index.names}
        if any(lvl is None for lvl in levels.values()):
            levels = {i: f"index_{i}" for i in range(len(levels))}

        for level_name, prop_name in levels.items():
            index_props.append(
                _gen_prop(
                    prop_name, data.index.get_level_values(level_name).to_series()
                )
            )

    return [
        *index_props,
        *(_gen_prop(str(col.name), col) for col in columns),
    ]


def _remove_external_fk(table: sqla.Table):
    """Dirty vodoo to remove external FKs from existing table."""
    for c in table.columns.values():
        c.foreign_keys = set(
            fk
            for fk in c.foreign_keys
            if fk.constraint and fk.constraint.referred_table.schema == table.schema
        )

    table.foreign_keys = set(  # type: ignore[reportAttributeAccessIssue]
        fk
        for fk in table.foreign_keys
        if fk.constraint and fk.constraint.referred_table.schema == table.schema
    )

    table.constraints = set(
        c
        for c in table.constraints
        if not isinstance(c, sqla.ForeignKeyConstraint)
        or c.referred_table.schema == table.schema
    )


@dataclass(frozen=True)
class PropType(Generic[VarT]):
    """Reference to a type."""

    hint: str | type[Prop[VarT]] | None = None
    ctx: ModuleType | None = None
    typevar_map: dict[TypeVar, SingleTypeDef] = field(default_factory=dict)

    def prop_type(self) -> type[Prop] | type[None]:
        """Resolve the property type reference."""
        hint = self.hint

        if hint is None:
            return NoneType

        if isinstance(hint, type | GenericAlias):
            base = get_origin(hint)
            if base is None or not issubclass(base, Prop):
                return NoneType

            return base
        else:
            return (
                Col
                if "Col" in hint
                else (
                    RelSet
                    if "RelSet" in hint
                    else Rel if "Rel" in hint else Prop if "Prop" in hint else NoneType
                )
            )

    def _to_typedef(
        self, hint: SingleTypeDef | UnionType | TypeVar | str | ForwardRef
    ) -> SingleTypeDef:
        typedef = hint

        if isinstance(typedef, str):
            typedef = eval(
                typedef, {**globals(), **(vars(self.ctx) if self.ctx else {})}
            )

        if isinstance(typedef, TypeVar):
            typedef = self.typevar_map.get(typedef) or typedef.__bound__ or object

        if isinstance(typedef, ForwardRef):
            typedef = typedef._evaluate(
                {**globals(), **(vars(self.ctx) if self.ctx else {})},
                {},
                recursive_guard=frozenset(),
            )

        if isinstance(typedef, UnionType):
            union_types: set[type] = {
                get_origin(union_arg) or union_arg for union_arg in get_args(typedef)
            }
            typedef = get_lowest_common_base(union_types)

        return cast(SingleTypeDef, typedef)

    def _to_type(self, hint: SingleTypeDef | UnionType | TypeVar) -> type:
        if isinstance(hint, type):
            return hint

        typedef = (
            self._to_typedef(hint) if isinstance(hint, UnionType | TypeVar) else hint
        )
        orig = get_origin(typedef)

        if orig is None or orig is Literal:
            return object

        assert isinstance(orig, type)
        return new_class(
            orig.__name__ + "_" + token_hex(5),
            (hint,),
            None,
            lambda ns: ns.update({"_src_mod": self.ctx}),
        )

    @cached_property
    def _generic_type(self) -> SingleTypeDef | UnionType:
        """Resolve the generic property type reference."""
        hint = self.hint or Prop
        generic = self._to_typedef(hint)
        assert is_subtype(generic, Prop)

        return generic

    @cached_property
    def _generic_args(self) -> tuple[SingleTypeDef | UnionType | TypeVar, ...]:
        """Resolve the generic property type reference."""
        args = get_args(self._generic_type)
        return tuple(self._to_typedef(hint) for hint in args)

    def value_type(self: PropType[VarT2]) -> SingleTypeDef[VarT2]:
        """Resolve the value type reference."""
        args = self._generic_args
        if len(args) == 0:
            return cast(type[VarT2], object)

        arg = args[0]
        return cast(SingleTypeDef[VarT2], self._to_typedef(arg))

    def record_type(
        self: PropType[Record | Iterable[Record | None] | None],
    ) -> type[Record]:
        """Resolve the record type reference."""
        assert is_subtype(self._generic_type, RelSet)

        args = self._generic_args
        assert len(args) > 0

        recs = args[0]
        if isinstance(recs, TypeVar):
            recs = self._to_type(recs)

        rec_args = get_args(recs)
        rec_arg = extract_nullable_type(recs)
        assert rec_arg is not None

        if is_subtype(recs, Iterable):
            assert len(rec_args) >= 1
            rec_arg = rec_args[0]

        rec_type = self._to_type(rec_arg)
        assert issubclass(rec_type, Record)

        return rec_type

    def link_type(
        self: PropType[Record | Iterable[Record | None] | None],
    ) -> type[Record | None]:
        """Resolve the record type reference."""
        assert is_subtype(self._generic_type, RelSet)
        args = self._generic_args
        rec = args[1]
        rec_type = self._to_type(rec)

        if not issubclass(rec_type, Record):
            return NoneType

        return rec_type


ParT = TypeVarX("ParT", covariant=True, bound="Record", default="Record")
ParT2 = TypeVar("ParT2", bound="Record")


@dataclass(eq=False)
class Prop(ABC, Generic[VarT, RwT, ParT]):
    """Reference a property of a record."""

    _type: PropType[VarT] = field(default_factory=PropType[VarT])
    _name: str | None = None
    _parent_type: type[ParT] | None = None

    alias: str | None = None
    default: VarT | State = State.undef
    default_factory: Callable[[], VarT] | None = None
    init: bool = True

    getter: Callable[[Record], VarT] | None = None
    setter: Callable[[Record, VarT], None] | None = None

    local: bool = False

    @property
    def name(self) -> str:
        """Property name."""
        if self.alias is not None:
            return self.alias

        assert self._name is not None
        return self._name

    @cached_property
    def value_type(self) -> SingleTypeDef[VarT]:
        """Value type of the property."""
        return self._type.value_type()

    @cached_property
    def parent_type(self) -> type[ParT]:
        """Parent record type."""
        return cast(type[ParT], self._parent_type or Record)

    def __set_name__(self, _, name: str) -> None:  # noqa: D105
        if self._name is None:
            self._name = name
        else:
            assert name == self._name

    def __hash__(self: Prop[Any, Any, Any]) -> int:
        """Hash the Prop."""
        return gen_int_hash((self.parent_type, self.name))

    def __eq__(self, other: object) -> bool:
        """Hash the Prop."""
        return hash(self) == hash(other)


class Attr(Prop[VarT, RwT, ParT]):
    """Record attribute."""

    @overload
    def __get__(self, instance: None, owner: type[RecT2]) -> Attr[VarT, RwT, RecT2]: ...

    @overload
    def __get__(self, instance: RecT2, owner: type[RecT2]) -> VarT: ...

    @overload
    def __get__(self, instance: object | None, owner: type | None) -> Self: ...

    def __get__(  # noqa: D105
        self, instance: object | None, owner: type | type[RecT2] | None
    ) -> Attr[Any, Any, Any] | VarT | Self:
        if isinstance(instance, Record):
            if self.getter is not None:
                value = self.getter(instance)
            else:
                value = instance.__dict__.get(self.name, State.undef)

            if value is State.undef:
                if self.default_factory is not None:
                    value = self.default_factory()
                else:
                    value = self.default

                assert (
                    value is not State.undef
                ), f"Property value for `{self.name}` could not be fetched."
                instance.__dict__[self.name] = value

            return value
        elif owner is not None and issubclass(owner, Record):
            rec_type = cast(type[RecT2], owner)
            return cast(
                Attr[VarT, RwT, RecT2],
                copy_and_override(
                    self,
                    type(self),
                    _parent_type=rec_type,  # type: ignore
                ),
            )

        return self

    def __set__(self: Prop[VarT2, RW], instance: Record, value: VarT2 | State) -> None:
        """Set the value of the column."""
        if value is State.undef:
            if self.name in instance.__dict__:
                value = instance.__dict__[self.name]
            elif self.default_factory is not None:
                value = self.default_factory()
            else:
                value = self.default

            assert (
                value is not State.undef
            ), f"Property value for `{self.name}` could not be fetched."

        if self.setter is not None:
            self.setter(instance, cast(VarT2, value))
        else:
            instance.__dict__[self.name] = value


type DirectLink[Rec: Record] = (
    Col[Any, Any, Any, Any, Any] | dict[Col, Col[Any, Any, Any, Any, Rec]]
)

type BackLink[Rec: Record] = (RelSet[Any, Any, Any, Any, LocalStat, Rec] | type[Rec])

type BiLink[Rec: Record, Rec2: Record] = (
    RelSet[Rec | None, Any, Any, Any, LocalStat, Rec2]
    | tuple[
        RelSet[Any, Any, Any, Any, LocalStat, Rec2],
        RelSet[Rec | None, Any, Any, Any, LocalStat, Rec2],
    ]
    | type[Rec2]
)


@overload
def prop(
    *,
    default: VarT2 | RecT2 | None = ...,
    default_factory: Callable[[], VarT2 | RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: bool = ...,
    primary_key: bool = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: Literal[True],
    init: bool = ...,
) -> Any: ...


@overload
def prop(
    *,
    default: VarT2 | RecT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: bool = ...,
    primary_key: bool = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> Col[VarT2]: ...


@overload
def prop(
    *,
    default: VarT2 | RecT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: bool | Literal["fk"] = ...,
    primary_key: bool | Literal["fk"] = ...,
    link_on: DirectLink[RecT2],
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> Rel[RecT2]: ...


@overload
def prop(
    *,
    default: VarT2 | RecT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal["fk"],
    primary_key: bool | Literal["fk"] = ...,
    link_on: DirectLink[RecT2] | None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> Rel[RecT2]: ...


@overload
def prop(
    *,
    default: VarT2 | RecT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: bool | Literal["fk"] = ...,
    primary_key: Literal["fk"],
    link_on: DirectLink[RecT2] | None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> Rel[RecT2]: ...


@overload
def prop(
    *,
    default: VarT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], VarT2],
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> Prop[VarT2, RO]: ...


@overload
def prop(
    *,
    default: VarT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], VarT2],
    setter: Callable[[Record, VarT2], None],
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> Prop[VarT2, RW]: ...


@overload
def prop(
    *,
    default: VarT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: bool = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], VarT2],
    setter: None = ...,
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement],
    local: bool = ...,
    init: bool = ...,
) -> Col[VarT2, RO]: ...


@overload
def prop(
    *,
    default: VarT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: bool = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], VarT2],
    setter: Callable[[Record, VarT2], None],
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement],
    local: bool = ...,
    init: bool = ...,
) -> Col[VarT2, RW]: ...


@overload
def prop(
    *,
    default: VarT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal[True] = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], VarT2],
    setter: None = ...,
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement] | None = ...,
    local: bool = ...,
    init: bool = ...,
) -> Col[VarT2, RO]: ...


@overload
def prop(
    *,
    default: VarT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], VarT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal[True] = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: None = ...,
    link_via: None = ...,
    order_by: None = ...,
    map_by: None = ...,
    getter: Callable[[Record], VarT2],
    setter: Callable[[Record, VarT2], None],
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement] | None = ...,
    local: bool = ...,
    init: bool = ...,
) -> Col[VarT2, RW]: ...


@overload
def prop(  # type: ignore[reportOverlappingOverload]
    *,
    default: RecT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: BackLink[RecT2] | None = ...,
    link_via: BiLink[RecT2, Link],
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> RelSet[RecT2, None]: ...


@overload
def prop(
    *,
    default: RecT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: BackLink[RecT2] | None = ...,
    link_via: BiLink[RecT2, RecT3] | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    order_by: None = ...,
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> RelSet[RecT2, RecT3]: ...


@overload
def prop(
    *,
    default: RecT2 | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    default_factory: Callable[[], RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: BackLink[RecT2] | None = ...,
    link_via: BiLink[RecT2, RecT3] | None = ...,  # type: ignore[reportInvalidTypeVarUse]
    order_by: Mapping[Col[Any, Any, Any, LocalStat, RecT2 | RecT3], int],
    map_by: None = ...,
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> RelSet[RecT2, RecT3, int]: ...


@overload
def prop(
    *,
    default: RecT2 | None = ...,
    default_factory: Callable[[], RecT2 | VarT3] | None = ...,
    alias: str | None = ...,
    index: Literal[False] = ...,
    primary_key: Literal[False] = ...,
    link_on: None = ...,
    link_from: BackLink[RecT2] | None = ...,
    link_via: BiLink[RecT2, RecT3] | None = ...,
    order_by: None = ...,
    map_by: Col[VarT4, Any, Any, LocalStat, RecT2 | RecT3],
    getter: None = ...,
    setter: None = ...,
    sql_getter: None = ...,
    local: bool = ...,
    init: bool = ...,
) -> RelSet[RecT2, RecT3, VarT4]: ...


def prop(
    *,
    default: Any | None = None,
    default_factory: Callable[[], Any] | None = None,
    alias: str | None = None,
    index: bool | Literal["fk"] = False,
    primary_key: bool | Literal["fk"] = False,
    link_on: DirectLink[Record] | None = None,
    link_from: BackLink[Record] | None = None,
    link_via: BiLink[Record, Any] | None = None,
    order_by: Mapping[Col, int] | None = None,
    map_by: Col | None = None,
    getter: Callable[[Record], Any] | None = None,
    setter: Callable[[Record, Any], None] | None = None,
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement] | None = None,
    local: bool = False,
    init: bool = True,
) -> Any:
    """Define a property."""
    if local:
        return Attr(
            default=default,
            default_factory=default_factory,
            alias=alias,
            init=init,
            getter=getter,
            setter=setter,
            _type=PropType(Attr[object]),
            local=local,
        )

    if any(
        a is not None for a in (link_on, link_from, link_via, order_by, map_by)
    ) or any(a == "fk" for a in (index, primary_key)):
        return RelSet(
            default=default,
            default_factory=default_factory,
            alias=alias,
            init=init,
            index=index == "fk",
            primary_key=primary_key == "fk",
            on=(
                link_on
                if link_on is not None
                else link_from if link_from is not None else link_via
            ),  # type: ignore[reportArgumentType]
            order_by=order_by,
            map_by=map_by,
            _type=PropType(RelSet[Record]),
        )

    return Col(
        default=default,
        default_factory=default_factory,
        alias=alias,
        init=init,
        index=index is not False,
        primary_key=primary_key is not False,
        getter=getter,
        setter=setter,
        sql_getter=sql_getter,
        _type=PropType(Col[object]),
        local=local,
    )


class RecordMeta(type):
    """Metaclass for record types."""

    _record_bases: list[type[Record]]
    _base_pks: dict[
        type[Record],
        dict[Col[Hashable], Col[Hashable]],
    ]

    _defined_props: dict[str, Col | RelSet]
    _defined_cols: dict[str, Col]
    _defined_rels: dict[str, RelSet]

    _is_record: bool = False
    _template: bool = False
    _src_mod: ModuleType | None = None

    def __init__(cls, name, bases, dct):
        """Initialize a new record type."""
        super().__init__(name, bases, dct)

        if "_src_mod" not in cls.__dict__:
            cls._src_mod = getmodule(cls)

        prop_defs = {
            name: pt
            for name, hint in get_annotations(cls).items()
            if issubclass((pt := PropType(hint, ctx=cls._src_mod)).prop_type(), Prop)
            or (isinstance(hint, str) and ("Attr" in hint or "Rel" in hint))
        }

        for prop_name, prop_type in prop_defs.items():
            if prop_name not in cls.__dict__:
                pt = prop_type.prop_type()
                assert issubclass(pt, Prop)
                setattr(
                    cls,
                    prop_name,
                    pt(_name=prop_name, _type=prop_type),
                )
            else:
                prop = cls.__dict__[prop_name]
                assert isinstance(prop, Prop)
                prop._type = prop_type

        cls._record_bases = []
        base_types: Iterable[type | GenericAlias] = (
            cls.__dict__["__orig_bases__"]
            if "__orig_bases__" in cls.__dict__
            else cls.__bases__
        )
        for c in base_types:
            orig = get_origin(c) if not isinstance(c, type) else c
            if orig is not c and hasattr(orig, "__parameters__"):
                typevar_map = dict(zip(getattr(orig, "__parameters__"), get_args(c)))
            else:
                typevar_map = {}

            if not isinstance(orig, RecordMeta) or orig._is_record:
                continue

            if orig._template:
                for prop_name, prop_set in orig._defined_props.items():
                    prop_defs[prop_name] = prop_set._type
                    if prop_name not in cls.__dict__:
                        if isinstance(prop_set, Col):
                            prop_set = copy_and_override(
                                prop_set,
                                type(prop_set),
                                _type=copy_and_override(
                                    prop_set._type,
                                    type(prop_set._type),
                                    typevar_map=typevar_map,
                                    ctx=cls._src_mod,
                                ),
                            )
                        else:
                            prop_set = copy_and_override(
                                prop_set,
                                type(prop_set),
                                _parent_type=prop_set.parent_type,
                                _type=copy_and_override(
                                    prop_set._type,
                                    type(prop_set._type),
                                    typevar_map=typevar_map,
                                    ctx=cls._src_mod,
                                ),
                            )
                        setattr(cls, prop_name, prop_set)
            else:
                assert orig is c
                cls._record_bases.append(orig)

        cls._base_pks = (
            {
                base: {
                    Col[Hashable, Any, Record](
                        _name=pk.name,
                        primary_key=True,
                        _type=PropType(Col[pk.value_type, Any, Record]),
                        _parent_type=cast(type[Record], cls),
                    ): pk
                    for pk in base._primary_keys.values()
                }
                for base in cls._record_bases
            }
            if not cls._is_record
            else {}
        )

        cls._defined_props = {name: getattr(cls, name) for name in prop_defs.keys()}

        cls._defined_cols = {
            name: col
            for name, col in cls._defined_props.items()
            if isinstance(col, Col)
        }

        cls._defined_rels = {
            name: rel
            for name, rel in cls._defined_props.items()
            if isinstance(rel, RelSet)
        }

    @property
    def _defined_rel_cols(cls) -> dict[str, Col[Hashable]]:
        return {
            a.name: a
            for rel in cls._defined_rels.values()
            if rel.direct_rel is True
            for a in rel.fk_map.keys()
        }

    @property
    def _primary_keys(cls: type[Record]) -> dict[str, Col[Hashable]]:
        base_pk_cols = {v.name: v for vs in cls._base_pks.values() for v in vs.keys()}
        defined_pks = {
            name: a for name, a in cls._defined_cols.items() if a.primary_key
        }

        if len(base_pk_cols) > 0 and len(defined_pks) > 0:
            assert set(defined_pks.keys()).issubset(
                set(base_pk_cols.keys())
            ), f"Primary keys of {cls} are ambiguous. "
            return base_pk_cols

        return {**base_pk_cols, **defined_pks}

    @property
    def _props(cls) -> dict[str, Col | RelSet]:
        return reduce(
            lambda x, y: {**x, **y},
            (c._props for c in cls._record_bases),
            cls._defined_props,
        )

    @property
    def _cols(cls) -> dict[str, Col]:
        return reduce(
            lambda x, y: {**x, **y},
            (c._cols for c in cls._record_bases),
            cls._defined_cols,
        )

    @property
    def _rels(cls) -> dict[str, RelSet]:
        return reduce(
            lambda x, y: {**x, **y},
            (c._rels for c in cls._record_bases),
            cls._defined_rels,
        )

    @property
    def _rel_types(cls) -> set[type[Record]]:
        return {rel.record_type for rel in cls._rels.values()}


class RecDB:
    """Descriptor to access the database of a record."""

    @overload
    def __get__(  # noqa: D105
        self, instance: None, owner: type[Record]
    ) -> DB[RW, LocalStat]: ...

    @overload
    def __get__(  # noqa: D105
        self, instance: Record, owner: type[Record]
    ) -> DB[RW, DynBackendID]: ...

    def __get__(  # noqa: D105
        self, instance: Record | None, owner: type[Record]
    ) -> DB[RW, LocalStat] | DB[RW, DynBackendID]:
        if instance is None:
            return DB(b_id=LocalBackend.static, types={owner})

        if "__db" not in instance.__dict__:
            db = DB(b_id=LocalBackend.dynamic, types={owner})
            instance.__dict__["__db"] = db
            instance._connected = True
            instance._root = True
            db[owner][instance._index] @= instance

        return instance.__dict__["__db"]

    def __set__(self, instance: Record, value: DB[RW, Any]) -> None:  # noqa: D105
        instance.__dict__["__db"] = value
        instance._connected = True
        instance._root = False


@dataclass_transform(kw_only_default=True, field_specifiers=(prop,), eq_default=False)
class Record(Generic[KeyT], metaclass=RecordMeta):
    """Schema for a record in a database."""

    _template: ClassVar[bool]
    _table_name: ClassVar[str] | None = None
    _type_map: ClassVar[dict[type, sqla.types.TypeEngine]] = {
        str: sqla.types.String().with_variant(
            sqla.types.String(50), "oracle"
        ),  # Avoid oracle error when VARCHAR has no size parameter
        UUID: UUIDType(binary=False),  # Binary type causes issues with DuckDB
    }
    _is_record: ClassVar[bool] = True
    __dataclass_fields__: ClassVar[dict[str, Field[Any]]]

    _db: RecDB = RecDB()
    _connected: bool = prop(default=False, local=True)
    _root: bool = prop(default=True, local=True)

    def __init_subclass__(cls, **kwargs: Any) -> None:
        """Initialize a new record subclass."""
        super().__init_subclass__(**kwargs)
        cls._is_record = False

        cls.__dataclass_fields__ = {
            **{
                name: Field(
                    p.default,
                    p.default_factory,  # type: ignore[reportArgumentType]
                    p.init,
                    repr=True,
                    hash=p.primary_key,
                    metadata={},
                    compare=True,
                    kw_only=True,
                )
                for name, p in cls._cols.items()
            },
            **{
                name: Field(
                    p.default,
                    p.default_factory,  # type: ignore[reportArgumentType]
                    p.init,
                    repr=True,
                    hash=False,
                    metadata={},
                    compare=True,
                    kw_only=True,
                )
                for name, p in cls._rels.items()
            },
        }

    def __init__(self, **kwargs: Any) -> None:
        """Initialize a new record instance."""
        super().__init__()
        for name, value in kwargs.items():
            setattr(self, name, value)

    @classmethod
    def _default_table_name(cls) -> str:
        """Return the name of the table for this schema."""
        if cls._table_name is not None:
            return cls._table_name

        fqn_parts = PyObjectRef.reference(cls).fqn.split(".")

        name = fqn_parts[-1]
        for part in reversed(fqn_parts[:-1]):
            name = part + "_" + name
            if len(name) > 40:
                break

        return name

    @classmethod
    def _sql_table_name(
        cls,
        subs: Mapping[type[Record], sqla.TableClause],
    ) -> str:
        """Return a SQLAlchemy table object for this schema."""
        sub = subs.get(cls)
        return sub.name if sub is not None else cls._default_table_name()

    @classmethod
    def _sql_cols(cls, registry: orm.registry) -> dict[str, sqla.Column]:
        """Columns of this record type's table."""
        table_cols: dict[str, Col] = {
            **cls._defined_cols,
            **cls._defined_rel_cols,
            **cls._primary_keys,
        }

        return {
            name: sqla.Column(
                col.name,
                registry._resolve_type(col.value_type),  # type: ignore
                primary_key=col.primary_key,
                autoincrement=False,
                index=col.index,
                nullable=has_type(None, col.value_type),
            )
            for name, col in table_cols.items()
        }

    @classmethod
    def _foreign_keys(
        cls, metadata: sqla.MetaData, subs: Mapping[type[Record], sqla.TableClause]
    ) -> list[sqla.ForeignKeyConstraint]:
        """Foreign key constraints for this record type's table."""
        fks: list[sqla.ForeignKeyConstraint] = []

        for rel in cls._defined_rels.values():
            if len(rel.fk_map) > 0:
                rel_table = rel.record_type._table(metadata, subs)
                fks.append(
                    sqla.ForeignKeyConstraint(
                        [col.name for col in rel.fk_map.keys()],
                        [rel_table.c[col.name] for col in rel.fk_map.values()],
                        name=f"{cls._sql_table_name(subs)}_{rel.name}_fk",
                    )
                )

        for base, pks in cls._base_pks.items():
            base_table = base._table(metadata, subs)

            fks.append(
                sqla.ForeignKeyConstraint(
                    [col.name for col in pks.keys()],
                    [base_table.c[col.name] for col in pks.values()],
                    name=(
                        cls._sql_table_name(subs)
                        + "_base_fk_"
                        + gen_str_hash(base._sql_table_name(subs), 5)
                    ),
                )
            )

        return fks

    @classmethod
    def _table(
        cls,
        metadata: sqla.MetaData,
        subs: Mapping[type[Record], sqla.TableClause],
    ) -> sqla.Table:
        """Return a SQLAlchemy table object for this schema."""
        table_name = cls._sql_table_name(subs)

        if table_name in metadata.tables:
            # Return the table object from metadata if it already exists.
            # This is necessary to avoid circular dependencies.
            return metadata.tables[table_name]

        registry = orm.registry(metadata=metadata, type_annotation_map=cls._type_map)
        sub = subs.get(cls)

        cols = cls._sql_cols(registry)

        # Create a partial SQLAlchemy table object from the class definition
        # without foreign keys to avoid circular dependencies.
        # This adds the table to the metadata.
        sqla.Table(
            table_name,
            registry.metadata,
            *cols.values(),
            schema=(sub.schema if sub is not None else None),
        )

        fks = cls._foreign_keys(metadata, subs)

        # Re-create the table object with foreign keys and return it.
        return sqla.Table(
            table_name,
            registry.metadata,
            *cols.values(),
            *fks,
            schema=(sub.schema if sub is not None else None),
            extend_existing=True,
        )

    @classmethod
    def _joined_table(
        cls,
        metadata: sqla.MetaData,
        subs: Mapping[type[Record], sqla.TableClause],
    ) -> sqla.Table | sqla.Join:
        """Recursively join all bases of this record to get the full data."""
        table = cls._table(metadata, subs)

        base_joins = [
            (base._joined_table(metadata, subs), pk_map)
            for base, pk_map in cls._base_pks.items()
        ]
        for target_table, pk_map in base_joins:
            table = table.join(
                target_table,
                reduce(
                    sqla.and_,
                    (
                        table.c[pk.name] == target_table.c[target_pk.name]
                        for pk, target_pk in pk_map.items()
                    ),
                ),
            )

        return table

    @classmethod
    def _backrels_to_rels(
        cls, target: type[RecT2]
    ) -> set[RelSet[Self | None, Any, SingleIdx, Any, LocalStat, RecT2]]:
        """Get all direct relations from a target record type to this type."""
        rels: set[RelSet[Self | None, Any, SingleIdx, Any, LocalStat, RecT2]] = set()
        for rel in cls._rels.values():
            if issubclass(target, rel.fk_record_type):
                rel = cast(
                    RelSet[Self | None, Any, Any, Any, LocalStat, RecT2, Self], rel
                )
                rels.add(
                    RelSet[Self | None, Any, SingleIdx, Any, LocalStat, RecT2, Self](
                        on=rel.on,
                        _type=PropType(
                            RelSet[cls | None, Any, Any, Any, Any, target, cls]
                        ),
                        _parent_type=cast(type[RecT2], rel.fk_record_type),
                    )
                )

        return rels

    @classmethod
    def _rel(
        cls, other: type[RecT2]
    ) -> RelSet[RecT2, Any, BaseIdx, Any, LocalStat, Self]:
        """Dynamically define a relation to another record type."""
        return RelSet[RecT2, Any, Any, Any, LocalStat, Self](
            on=other,
            _type=PropType(RelSet[other, Any, BaseIdx, Any, LocalStat, cls]),
            _parent_type=cls,
        )

    @classmethod
    def __clause_element__(cls) -> sqla.TableClause:  # noqa: D105
        assert cls._default_table_name() is not None
        return sqla.table(cls._default_table_name())

    @classmethod  # type: ignore[reportArgumentType]
    def _copy(
        cls: Callable[Params, RecT2],
        rec: Record[KeyT],
        idx: KeyT,
        *args: Params.args,
        **kwargs: Params.kwargs,
    ) -> RecT2:
        rec_type = type(rec)
        target_props = rec_type._props

        rec_props = {p: getattr(rec, p.name) for p in target_props.values() if p.init}
        rec_kwargs = {p.name: v for p, v in rec_props.items()}

        if len(rec_type._primary_keys) > 1:
            assert isinstance(idx, Iterable)
            new_idx = dict(zip(rec_type._primary_keys, idx))
        else:
            new_idx = {next(iter(rec_type._primary_keys)): idx}

        new_kwargs = {**rec_kwargs, **kwargs, **new_idx}

        return cls(*args, **new_kwargs)

    def __hash__(self) -> int:
        """Identify the record by database and id."""
        return gen_int_hash((self._db if self._connected else None, self._index))

    def __eq__(self, value: Hashable) -> bool:
        """Check if the record is equal to another record."""
        return hash(self) == hash(value)

    @overload
    def _to_dict(
        self,
        name_keys: Literal[False] = ...,
        include: set[type[Prop]] | None = ...,
    ) -> dict[Col | RelSet, Any]: ...

    @overload
    def _to_dict(
        self,
        name_keys: Literal[True],
        include: set[type[Prop]] | None = ...,
    ) -> dict[str, Any]: ...

    def _to_dict(
        self,
        name_keys: bool = True,
        include: set[type[Prop]] | None = None,
    ) -> dict[Col | RelSet, Any] | dict[str, Any]:
        """Convert the record to a dictionary."""
        include_types: tuple[type[Prop], ...] = (
            tuple(include) if include is not None else (Col,)
        )

        vals = {
            p if not name_keys else p.name: getattr(self, p.name)
            for p in type(self)._props.values()
            if isinstance(p, include_types)
        }

        return cast(dict, vals)

    @property
    def _index(self) -> KeyT:
        """Return the index of the record."""
        pks = type(self)._primary_keys
        if len(pks) == 1:
            return getattr(self, next(iter(pks)))
        return cast(KeyT, tuple(getattr(self, pk) for pk in pks))

    @classmethod
    def _from_partial_dict(
        cls,
        data: Mapping[Col | RelSet, Any] | Mapping[str, Any],
        name_keys: bool = True,
    ) -> Self:
        """Return the index contained in a dict representation of this record."""
        args = {
            p.name: data.get(p if not name_keys else p.name, State.undef)  # type: ignore[reportArgumentType]
            for p in cls._props.values()
        }
        return cls(**args)

    def __repr__(self) -> str:  # noqa: D105
        return self._to_dict(name_keys=True).__repr__()


ResT = TypeVarX("ResT", covariant=True, bound="Record | None", default="Record | None")
ResT2 = TypeVarX("ResT2", bound="Record | None")

ResTd = TypeVarX(
    "ResTd", covariant=True, bound="Record | None", default="Record | None"
)

RecT = TypeVarX("RecT", bound="Record", default="Record")
RecTt = TypeVarTuple("RecTt")
RecT2 = TypeVar("RecT2", bound="Record")
RecT3 = TypeVar("RecT3", bound="Record")

BsT = TypeVarX(
    "BsT",
    bound="StatBackendID",
    default="LocalStat",
    covariant=True,
)

BkT = TypeVarX(
    "BkT",
    bound="StatBackendID",
    default="LocalDyn",
    covariant=True,
)
BtT = TypeVar(
    "BtT",
    bound="DynBackendID",
)

LnT = TypeVarX("LnT", covariant=True, bound="Record | None", default="Record | None")
LnT2 = TypeVar("LnT2", bound="Record | None")


@dataclass
class RelTree(Generic[*RecTt]):
    """Tree of relations starting from the same root."""

    rels: Iterable[RelSet[Record | None, Any, Any, Any]] = field(default_factory=set)

    def __post_init__(self) -> None:  # noqa: D105
        assert all(
            rel.root_set == self.root_set for rel in self.rels
        ), "Relations in set must all start from same root."
        self.targets = [rel.record_type for rel in self.rels]

    @cached_property
    def root_set(self) -> RecSet[Record | None, Any, Any]:
        """Root record type of the set."""
        return list(self.rels)[-1].root_set

    @cached_property
    def dict(self) -> DataDict:
        """Tree representation of the relation set."""
        tree: DataDict = {}

        for rel in self.rels:
            subtree = tree
            if len(rel._rel_path) > 1:
                for ref in rel._rel_path[1:]:
                    if ref not in subtree:
                        subtree[ref] = {}
                    subtree = subtree[ref]

        return tree

    def prefix(
        self, prefix: type[Record] | RecSet[Any, Any, Any, Any, Any, Any]
    ) -> Self:
        """Prefix all relations in the set with given relation."""
        rels = {rel.prefix(prefix) for rel in self.rels}
        return cast(Self, RelTree(rels))

    def __mul__(self, other: RelSet[RecT2] | RelTree) -> RelTree[*RecTt, RecT2]:
        """Append more rels to the set."""
        other = other if isinstance(other, RelTree) else RelTree([other])
        return RelTree([*self.rels, *other.rels])

    def __rmul__(self, other: RelSet[RecT2] | RelTree) -> RelTree[RecT2, *RecTt]:
        """Append more rels to the set."""
        other = other if isinstance(other, RelTree) else RelTree([other])
        return RelTree([*other.rels, *self.rels])

    def __or__(
        self: RelTree[RecT2], other: RelSet[RecT3] | RelTree[RecT3]
    ) -> RelTree[RecT2 | RecT3]:
        """Add more rels to the set."""
        other = other if isinstance(other, RelTree) else RelTree([other])
        return RelTree([*self.rels, *other.rels])

    @property
    def types(self) -> tuple[*RecTt]:
        """Return the record types in the relation tree."""
        return cast(tuple[*RecTt], tuple(r.record_type for r in self.rels))


type AggMap[Rec: Record] = dict[Col[Rec, Any], Col | sqla.Function]


@dataclass(kw_only=True, frozen=True)
class Agg(Generic[RecT]):
    """Define an aggregation map."""

    target: type[RecT]
    map: AggMap[RecT]


type Join = tuple[sqla.FromClause, sqla.ColumnElement[bool]]


@dataclass(kw_only=True, eq=False)
class RecSet(
    Generic[ResT, IdxT, RwT, BsT, ResTd, SelT],
):
    """Record dataset."""

    db: DB[RwT, BsT] = field(default_factory=lambda: DB[RwT, BsT]())
    res_type: type[ResT] = NoneType

    keys: Sequence[slice | list[Hashable] | Hashable] = field(default_factory=list)
    filters: list[sqla.ColumnElement[bool]] = field(default_factory=list)
    merges: RelTree = field(default_factory=RelTree)
    sel_type: type[SelT] = NoneType

    def __hash__(self: RecSet[Any, Any, Any, Any, Any, Any]) -> int:
        """Hash the RecSet."""
        return gen_int_hash(
            (self.db, self.record_type, self.keys, self.filters, self.merges)
        )

    def __eq__(self, other: object) -> bool:
        """Hash the Prop."""
        return hash(self) == hash(other)

    @property
    def record_type(self: RecSet[RecT2 | None, Any, Any, Any, Any, Any]) -> type[RecT2]:
        """Set of all record types in this DB."""
        t = extract_nullable_type(self.res_type)
        assert t is not None
        return t

    @cached_property
    def _sql_alias(self) -> sqla.FromClause:
        """Get random alias for a type."""
        if isinstance(self, RelSet):
            return self.db._get_joined_table(self.record_type).alias(
                gen_str_hash(self.to_static(), 8)
            )

        return self.db._get_joined_table(self.record_type).alias(token_hex(4))

    @cached_property
    def _idx_cols(self: RecSet[Record, Any, Any, BsT]) -> list[sqla.ColumnElement]:
        """Return the index columns."""
        return [
            *(
                col.label(f"{self.record_type._default_table_name()}.{col_name}")
                for col_name, col in self.root_table.columns.items()
                if self.record_type._cols[col_name] in self.idx
            ),
            *(
                col.label(f"{rel.path_str}.{col_name}")
                for rel in self.merges.rels
                for col_name, col in rel._sql_alias.columns.items()
                if rel.record_type._cols[col_name] in self.idx
            ),
        ]

    @cached_property
    def idx(
        self,
    ) -> Mapping[Col[Any, Any, Any, Any, Any], RecSet[Any, Any, Any, Any, Any, Any]]:
        """Return the index cols."""
        return {c: self for c in self.record_type._primary_keys.values()}

    @cached_property
    def single_key(self) -> Hashable | None:
        """Return the single selected key, if it exists."""
        single_keys = [k for k in self.keys if not isinstance(k, list | slice)]
        if len(single_keys) > 0:
            return single_keys[0]

        return None

    @cached_property
    def root_table(self) -> sqla.FromClause:
        """Get the main table for the current selection."""
        return self._sql_alias

    def _visit_col(
        self,
        element: sqla_visitors.ExternallyTraversible,
        reflist: set[RelSet[Any, Any, Any, Any, Any]] = set(),
        render: bool = False,
        **kw: Any,
    ) -> sqla.ColumnElement | None:
        if isinstance(element, Col):
            if isinstance(element.record_set, RelSet):
                if (
                    not isinstance(self, RelSet)
                    or element.record_set.to_static() != self.to_static()
                ):
                    element._record_set = element.record_set.prefix(self)

                reflist.add(element.record_set)

            return element.record_set._sql_alias.c[element.name] if render else element

        return None

    def _parse_filters(
        self,
        filt: Iterable[sqla.ColumnElement[bool]],
    ) -> tuple[list[sqla.ColumnElement[bool]], RelTree]:
        """Parse filter argument and return SQL expression and join operations."""
        reflist: set[RelSet] = set()
        replace_func = partial(self._visit_col, reflist=reflist, render=False)
        parsed_filt = [
            sqla_visitors.replacement_traverse(f, {}, replace=replace_func)
            for f in filt
        ]
        merge = RelTree(reflist)

        return parsed_filt, merge

    @property
    def _rendered_filters(
        self,
    ) -> list[sqla.ColumnElement[bool]]:
        """Parse filter argument and return SQL expression and join operations."""
        replace_func = partial(self._visit_col, render=True)
        parsed_filt = [
            sqla_visitors.replacement_traverse(f, {}, replace=replace_func)
            for f in self.filters
        ]

        return parsed_filt

    def _append_rel[
        RS: RelSet[Any, Any, Any, Any, Any, Any, Any, Any]
    ](self: RecSet[Any, Any, Any, Any, Any, Any], rel: RS) -> RS:
        rel_filt, rel_filt_merges = self._parse_filters(rel.filters)

        return cast(
            RS,
            copy_and_override(
                rel,
                RelSet[Any, Any, Any, Any, Any, Any, Any, Any],
                _parent_type=self.rec,
                db=self.db,
                keys=[*self.keys, *rel.keys],
                filters=[*self.filters, *rel_filt],
                merges=self.merges * rel.merges.prefix(self) * rel_filt_merges,
            ),
        )

    def suffix(
        self, right: RelSet[RecT2 | None, Any, Any, Any, Any, Any, Any]
    ) -> RelSet[RecT2 | None, Record, Any, RwT, BsT, Record, Record]:
        """Prefix this prop with a relation or record type."""
        if self.res_type is NoneType:
            return right

        rel_path = right._rel_path[1:] if len(right._rel_path) > 1 else (right,)

        prefixed_rel = reduce(
            RecSet._append_rel,
            rel_path,
            self,
        )

        return cast(
            RelSet[RecT2, Record, Any, RwT, BsT, Any, Record],
            prefixed_rel,
        )

    @cached_property
    def rec(self: RecSet[RecT2 | None, Any, Any, Any, Any, Any]) -> type[RecT2]:
        """Reference props of the target record type."""
        return cast(
            type[RecT2],
            type(
                self.record_type.__name__ + "_" + token_hex(5),
                (self.record_type,),
                {"_rel": self},
            ),
        )

    # Overloads: attribute selection:

    # 1. DB-level type selection
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RecSet[Any, Any, Any, Any],
        key: type[RecT2],
    ) -> RecSet[RecT2, IdxT, RwT, BsT, RecT2]: ...

    # 2. Top-level attribute selection, custom index
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RecSet[RecT2, KeyT2 | Filt[KeyT2], Any, Any],
        key: Col[VarT2, Any, RecT2],
    ) -> Col[VarT2, RwT, KeyT, BsT, RecT2]: ...

    # 3. Top-level attribute selection, base index
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2], Any, Any, Any, RecT2],
        key: Col[VarT2, Any, RecT2],
    ) -> Col[VarT2, RwT, KeyT2, BsT, RecT2]: ...

    # 4. Nested attribute selection, custom index
    @overload
    def __getitem__(
        self: RecSet[Any, KeyT2 | Filt[KeyT2], Any, Any],
        key: Col[VarT2, Any, Any],
    ) -> Col[VarT2, RwT, KeyT2 | IdxStart[KeyT2], BsT, Record]: ...

    # 5. Nested attribute selection, base index
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2], Any, Any, Any, Any],
        key: Col[VarT2, Any, Any],
    ) -> Col[VarT2, RwT, KeyT2 | IdxStart[KeyT2], BsT, Record]: ...

    # Overloads: relation selection:

    # 6. Top-level relation selection, singular, base index
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RecSet[Record[KeyT2] | None, BaseIdx | Filt[BaseIdx], Any, Any, RecT2],
        key: RelSet[ResT2, LnT2, SingleIdx | None, Any, LocalStat, RecT2, Any, SelT],
    ) -> RelSet[ResT2, LnT2, KeyT2, RwT, BsT, RecT2, ResT2, SelT]: ...

    # 7. Top-level relation selection, singular, single index
    @overload
    def __getitem__(
        self: RecSet[RecT2 | None, SingleIdx, Any, Any],
        key: RelSet[ResT2, LnT2, SingleIdx, Any, LocalStat, RecT2],
    ) -> RelSet[ResT2, LnT2, SingleIdx, RwT, BsT, RecT2, ResT2, SelT]: ...

    # 8. Top-level relation selection, singular nullable, single index
    @overload
    def __getitem__(
        self: RecSet[RecT2 | None, SingleIdx, Any, Any],
        key: RelSet[ResT2, LnT2, SingleIdx | None, Any, LocalStat, RecT2],
    ) -> RelSet[
        ResT2 | Scalar[None], LnT2, SingleIdx, RwT, BsT, RecT2, ResT2, SelT
    ]: ...

    # 9. Top-level relation selection, singular, custom index
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RecSet[RecT2 | None, KeyT2 | Filt[KeyT2], Any, Any],
        key: RelSet[ResT2, LnT2, SingleIdx | None, Any, LocalStat, RecT2],
    ) -> RelSet[ResT2, LnT2, KeyT2, RwT, BsT, RecT2, ResT2, SelT]: ...

    # 10. Top-level relation selection, base plural, base index
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RecSet[Record[KeyT2] | None, BaseIdx | Filt[BaseIdx], Any, Any, RecT2],
        key: RelSet[
            ResT2,
            LnT2,
            BaseIdx | Filt[BaseIdx],
            Any,
            LocalStat,
            RecT2,
            Record[KeyT3],
        ],
    ) -> RelSet[ResT2, LnT2, tuple[KeyT2, KeyT3], RwT, BsT, RecT2, ResT2, SelT]: ...

    # 11. Top-level relation selection, base plural, single index
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RecSet[RecT2 | None, SingleIdx, Any, Any],
        key: RelSet[
            ResT2,
            LnT2,
            BaseIdx | Filt[BaseIdx],
            Any,
            LocalStat,
            RecT2,
            Record[KeyT3],
        ],
    ) -> RelSet[ResT2, LnT2, KeyT3, RwT, BsT, RecT2, ResT2, SelT]: ...

    # 12. Top-level relation selection, base plural, tuple index
    @overload
    def __getitem__(
        self: RecSet[
            Record[KeyT2] | None, tuple[*IdxTt] | Filt[tuple[*IdxTt]], Any, Any, RecT2
        ],
        key: RelSet[
            ResT2,
            LnT2,
            BaseIdx | Filt[BaseIdx],
            Any,
            LocalStat,
            RecT2,
            Record[KeyT3],
        ],
    ) -> RelSet[
        ResT2,
        LnT2,
        tuple[*IdxTt, KeyT3] | tuple[KeyT2, KeyT3],
        RwT,
        BsT,
        RecT2,
        ResT2,
        SelT,
    ]: ...

    # 13. Top-level relation selection, base plural, custom index
    @overload
    def __getitem__(
        self: RecSet[Any, KeyT2 | Filt[KeyT2], Any, Any, RecT2],
        key: RelSet[
            ResT2,
            LnT2,
            BaseIdx | Filt[BaseIdx],
            Any,
            LocalStat,
            RecT2,
            Record[KeyT3],
        ],
    ) -> RelSet[ResT2, LnT2, tuple[KeyT2, KeyT3], RwT, BsT, RecT2, ResT2, SelT]: ...

    # 14. Top-level relation selection, plural, base index
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2] | None, BaseIdx | Filt[BaseIdx], Any, Any, RecT2],
        key: RelSet[ResT2, LnT2, KeyT3 | Filt[KeyT3], Any, LocalStat, RecT2],
    ) -> RelSet[ResT2, LnT2, tuple[KeyT2, KeyT3], RwT, BsT, RecT2, ResT2, SelT]: ...

    # 15. Top-level relation selection, plural, single index
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RecSet[RecT2 | None, SingleIdx],
        key: RelSet[ResT2, LnT2, KeyT3 | Filt[KeyT3], Any, LocalStat, RecT2],
    ) -> RelSet[ResT2, LnT2, KeyT3, RwT, BsT, RecT2, ResT2, SelT]: ...

    # 16. Top-level relation selection, plural, tuple index
    @overload
    def __getitem__(
        self: RecSet[
            Record[KeyT2] | None, tuple[*IdxTt] | Filt[tuple[*IdxTt]], Any, Any, RecT2
        ],
        key: RelSet[ResT2, LnT2, KeyT3 | Filt[KeyT3], Any, LocalStat, RecT2],
    ) -> RelSet[
        ResT2,
        LnT2,
        tuple[*IdxTt, KeyT3] | tuple[KeyT2, KeyT3],
        RwT,
        BsT,
        RecT2,
        ResT2,
        SelT,
    ]: ...

    # 17. Top-level relation selection, plural, custom index
    @overload
    def __getitem__(
        self: RecSet[RecT2 | None, KeyT2 | Filt[KeyT2], Any, Any],
        key: RelSet[ResT2, LnT2, KeyT3 | Filt[KeyT3], Any, LocalStat, RecT2],
    ) -> RelSet[ResT2, LnT2, tuple[KeyT2, KeyT3], RwT, BsT, RecT2, ResT2, SelT]: ...

    # 18. Nested relation selection, singular, base index
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2] | None, BaseIdx | Filt[BaseIdx], Any, Any, Any],
        key: RelSet[ResT2, LnT2, SingleIdx, Any, LocalStat, ParT2],
    ) -> RelSet[ResT2, LnT2, IdxStart[KeyT2], RwT, BsT, ParT2, ResT2, SelT]: ...

    # 19. Nested relation selection, singular, single index
    @overload
    def __getitem__(
        self: RecSet[Any, SingleIdx, Any, Any, Any],
        key: RelSet[ResT2, LnT2, SingleIdx, Any, LocalStat, ParT2],
    ) -> RelSet[ResT2, LnT2, Hashable | SingleIdx, RwT, BsT, ParT2, ResT2, SelT]: ...

    # 20. Nested relation selection, singular, tuple index
    @overload
    def __getitem__(
        self: RecSet[Any, tuple[*IdxTt] | Filt[tuple[*IdxTt]], Any, Any, Any],
        key: RelSet[ResT2, LnT2, SingleIdx, Any, LocalStat, ParT2],
    ) -> RelSet[ResT2, LnT2, IdxTupStart[*IdxTt], RwT, BsT, ParT2, ResT2, SelT]: ...

    # 21. Nested relation selection, singular, custom index
    @overload
    def __getitem__(
        self: RecSet[Any, KeyT2 | Filt[KeyT2], Any, Any, Any],
        key: RelSet[ResT2, LnT2, SingleIdx, Any, LocalStat, ParT2],
    ) -> RelSet[ResT2, LnT2, IdxStart[KeyT2], RwT, BsT, ParT2, ResT2, SelT]: ...

    # 22. Nested relation selection, base plural, base index
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2] | None, BaseIdx | Filt[BaseIdx], Any, Any, Any],
        key: RelSet[
            ResT2, LnT2, BaseIdx | Filt[BaseIdx], Any, LocalStat, ParT2, Record[KeyT3]
        ],
    ) -> RelSet[
        ResT2, LnT2, IdxStartEnd[KeyT2, KeyT3], RwT, BsT, ParT2, ResT2, SelT
    ]: ...

    # 23. Nested relation selection, base plural, single index
    @overload
    def __getitem__(
        self: RecSet[Any, SingleIdx, Any, Any, Any],
        key: RelSet[
            ResT2, LnT2, BaseIdx | Filt[BaseIdx], Any, LocalStat, ParT2, Record[KeyT3]
        ],
    ) -> RelSet[ResT2, LnT2, IdxEnd[KeyT3], RwT, BsT, ParT2, ResT2, SelT]: ...

    # 24. Nested relation selection, base plural, tuple index
    @overload
    def __getitem__(
        self: RecSet[Any, tuple[*IdxTt] | Filt[tuple[*IdxTt]], Any, Any, Any],
        key: RelSet[
            ResT2, LnT2, BaseIdx | Filt[BaseIdx], Any, LocalStat, ParT2, Record[KeyT3]
        ],
    ) -> RelSet[
        ResT2, LnT2, IdxTupStartEnd[*IdxTt, KeyT3], RwT, BsT, ParT2, ResT2, SelT
    ]: ...

    # 25. Nested relation selection, base plural, custom index
    @overload
    def __getitem__(
        self: RecSet[Any, KeyT2 | Filt[KeyT2], Any, Any, Any],
        key: RelSet[
            ResT2, LnT2, BaseIdx | Filt[BaseIdx], Any, LocalStat, ParT2, Record[KeyT3]
        ],
    ) -> RelSet[
        ResT2, LnT2, IdxStartEnd[KeyT2, KeyT3], RwT, BsT, ParT2, ResT2, SelT
    ]: ...

    # 26. Nested relation selection, plural, base index
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2] | None, BaseIdx | Filt[BaseIdx], Any, Any, Any],
        key: RelSet[ResT2, LnT2, KeyT3 | Filt[KeyT3], Any, LocalStat, ParT2],
    ) -> RelSet[
        ResT2, LnT2, IdxStartEnd[KeyT2, KeyT3], RwT, BsT, ParT2, ResT2, SelT
    ]: ...

    # 27. Nested relation selection, plural, single index
    @overload
    def __getitem__(
        self: RecSet[Any, SingleIdx],
        key: RelSet[ResT2, LnT2, KeyT3 | Filt[KeyT3], Any, LocalStat, ParT2],
    ) -> RelSet[ResT2, LnT2, IdxEnd[KeyT3], RwT, BsT, ParT2, ResT2, SelT]: ...

    # 28. Nested relation selection, plural, tuple index
    @overload
    def __getitem__(
        self: RecSet[Any, tuple[*IdxTt] | Filt[tuple[*IdxTt]]],
        key: RelSet[ResT2, LnT2, KeyT3 | Filt[KeyT3], Any, LocalStat, ParT2],
    ) -> RelSet[
        ResT2, LnT2, IdxTupStartEnd[*IdxTt, KeyT3], RwT, BsT, ParT2, ResT2, SelT
    ]: ...

    # 29. Nested relation selection, plural, custom index
    @overload
    def __getitem__(
        self: RecSet[Any, KeyT2 | Filt[KeyT2]],
        key: RelSet[ResT2, LnT2, KeyT3 | Filt[KeyT3], Any, LocalStat, ParT2],
    ) -> RelSet[
        ResT2, LnT2, IdxStartEnd[KeyT2, KeyT3], RwT, BsT, ParT2, ResT2, SelT
    ]: ...

    # 30. Default relation selection
    @overload
    def __getitem__(
        self: RecSet[Any, Any, Any, Any, Any],
        key: RelSet[ResT2, LnT2, Any, Any, LocalStat, ParT2],
    ) -> RelSet[ResT2, LnT2, Any, RwT, BsT, ParT2, ResT2, SelT]: ...

    # Index filtering and selection:

    # 31. RelSet: Merge selection, single index
    @overload
    def __getitem__(
        self: RelSet[RecT2 | None, LnT2, SingleIdx, Any, Any, ParT2, Any],
        key: RelTree[RecT2, *RecTt],
    ) -> RelSet[
        ResT,
        LnT2,
        BaseIdx,
        RwT,
        BsT,
        ParT2,
        ResTd,
        tuple[ResT, *RecTt],
    ]: ...

    # 32. RelSet: Merge selection, default
    @overload
    def __getitem__(
        self: RelSet[RecT2 | None, LnT2, Any, Any, Any, ParT2, Any],
        key: RelTree[RecT2, *RecTt],
    ) -> RelSet[
        ResT,
        LnT2,
        IdxT,
        RwT,
        BsT,
        ParT2,
        ResTd,
        tuple[ResT, *RecTt],
    ]: ...

    # 33. RelSet: Expression filtering, keep index
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RelSet[Any, LnT2, MdxT2 | Filt[MdxT2], Any, Any, ParT2],
        key: sqla.ColumnElement[bool],
    ) -> RelSet[ResT, LnT2, Filt[MdxT2], RwT, BsT, ParT2, ResTd]: ...

    # 34. RelSet: List selection
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RelSet[Record[KeyT2] | None, LnT2, MdxT2 | Filt[MdxT2], Any, Any, ParT2],
        key: Iterable[KeyT2 | MdxT2],
    ) -> RelSet[ResT, LnT2, Filt[MdxT2], RwT, BsT, ParT2, ResTd]: ...

    # 35. RelSet: Slice selection
    @overload
    def __getitem__(  # type: ignore[reportOverlappingOverload]
        self: RelSet[Any, LnT2, MdxT2 | Filt[MdxT2], Any, Any, ParT2],
        key: slice | tuple[slice, ...],
    ) -> RelSet[ResT, LnT2, Filt[MdxT2], RwT, BsT, ParT2, ResTd]: ...

    # 36. RelSet: Index value selection, record
    @overload
    def __getitem__(
        self: RelSet[
            Record[KeyT2] | None, LnT2, KeyT3 | Index, Any, Any, ParT2, RecT2, Record
        ],
        key: KeyT2 | KeyT3,
    ) -> RecT2: ...

    # 37. RelSet: Index value selection, non-record
    @overload
    def __getitem__(
        self: RelSet[Record[KeyT2] | None, LnT2, KeyT3 | Index, Any, Any, ParT2],
        key: KeyT2 | KeyT3,
    ) -> RelSet[ResT, LnT2, SingleIdx, RwT, BsT, ParT2, ResTd]: ...

    # 38. Merge selection, single index
    @overload
    def __getitem__(
        self: RecSet[Any, SingleIdx, Any, Any],
        key: RelTree[ResT, *RecTt],
    ) -> RecSet[ResT, BaseIdx, RwT, BsT, ResTd, tuple[ResT, *RecTt]]: ...

    # 39. Merge selection, default
    @overload
    def __getitem__(
        self: RecSet[Any, Any, Any, Any],
        key: RelTree[ResT, *RecTt],
    ) -> RecSet[ResT, IdxT, RwT, BsT, ResTd, tuple[ResT, *RecTt]]: ...

    # 40. Expression filtering, keep index
    @overload
    def __getitem__(
        self: RecSet[Any, MdxT2 | Filt[MdxT2], Any, Any], key: sqla.ColumnElement[bool]
    ) -> RecSet[ResT, Filt[MdxT2], RwT, BsT]: ...

    # 41. List selection
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2] | None, MdxT2 | Filt[MdxT2], Any, Any],
        key: Iterable[KeyT2 | MdxT2],
    ) -> RecSet[ResT, Filt[MdxT2], RwT, BsT]: ...

    # 42. Slice selection
    @overload
    def __getitem__(
        self: RecSet[Any, MdxT2 | Filt[MdxT2], Any, Any], key: slice | tuple[slice, ...]
    ) -> RecSet[ResT, Filt[MdxT2], RwT, BsT]: ...

    # 43. Index value selection, record
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2] | None, KeyT3 | Index, Any, Any, RecT2, Record],
        key: KeyT2 | KeyT3,
    ) -> RecT2: ...

    # 44. Index value selection, non-record
    @overload
    def __getitem__(
        self: RecSet[Record[KeyT2] | None, KeyT3 | Index, Any, Any], key: KeyT2 | KeyT3
    ) -> RecSet[ResT, SingleIdx, RwT, BsT]: ...

    # 45. Index value selection, default
    @overload
    def __getitem__(
        self: RecSet[Any, Any, Any, Any], key: Hashable
    ) -> RecSet[ResT, SingleIdx, RwT, BsT]: ...

    # Implementation:

    def __getitem__(  # noqa: D105
        self: RecSet[Record | None, Any, Any, Any, Any, Any],
        key: (
            type[Record]
            | Col
            | RecSet
            | RelTree
            | sqla.ColumnElement[bool]
            | list[Hashable]
            | slice
            | tuple[slice, ...]
            | Hashable
        ),
    ) -> (
        Col[Any, Any, Any, Any, Any]
        | RecSet[Record | None, Any, Any, Any, Any, Any]
        | Record
    ):
        match key:
            case type():
                assert issubclass(
                    key,
                    self.record_type,
                )
                return copy_and_override(self, RecSet[key], res_type=key)
            case Col():
                if isinstance(key.record_set, RelSet) and (
                    not isinstance(self, RelSet)
                    or (key.record_set.to_static() != self.to_static())
                ):
                    return self.suffix(key.record_set)[key]

                return Col(
                    _record_set=self,
                    _type=key._type,
                )
            case RelSet():
                return self.suffix(key)
            case RelTree():
                return copy_and_override(
                    self, type(self), merges=self.merges * key.prefix(self)
                )
            case sqla.ColumnElement():
                filt, filt_merges = self._parse_filters([key])
                return copy_and_override(
                    self,
                    type(self),
                    filters=[*self.filters, filt[0]],
                    merges=self.merges * filt_merges,
                )
            case list() | slice() | tuple() | Hashable():
                if not isinstance(key, list | slice) and not has_type(
                    key, tuple[slice, ...]
                ):
                    assert (
                        self.single_key is None or key == self.single_key
                    ), "Cannot select multiple single record keys"

                key_set = copy_and_override(self, type(self), keys=[*self.keys, key])

                if not is_subtype(self.sel_type, Record):
                    return key_set

                try:
                    return list(iter(key_set))[0]
                except IndexError as e:
                    raise KeyError(key) from e

    @overload
    def get(
        self: RecSet[RecT2 | None, KeyT3 | Index, Any, DynBackendID, Record[KeyT2]],
        key: KeyT2 | KeyT3,
        default: VarT2,
    ) -> RecT2 | VarT2: ...

    @overload
    def get(
        self: RecSet[RecT2 | None, KeyT3 | Index, Any, DynBackendID, Record[KeyT2]],
        key: KeyT2 | KeyT3,
        default: None = ...,
    ) -> RecT2 | None: ...

    def get(
        self: RecSet[Record | None, KeyT3 | Index, Any, DynBackendID, Record],
        key: Hashable | KeyT3,
        default: VarT2 | None = None,
    ) -> Record | VarT2 | None:
        """Get a record by key."""
        try:
            return list(iter(self[key]))[0]
        except KeyError | IndexError:
            return default

    def select(
        self,
        *,
        index_only: bool = False,
    ) -> sqla.Select:
        """Return select statement for this dataset."""
        idx_cols = [rel._sql_alias.c[col.name] for col, rel in self.idx.items()]
        select = sqla.select(
            *idx_cols,
            *(
                col.label(f"{self.record_type._default_table_name()}.{col_name}")
                for col_name, col in self._sql_alias.columns.items()
                if not index_only and col not in idx_cols
            ),
            *(
                col.label(f"{rel.path_str}.{col_name}")
                for rel in self.merges.rels
                for col_name, col in rel._sql_alias.columns.items()
                if not index_only and col not in idx_cols
            ),
        ).select_from(self._sql_alias)

        for join in self._joins():
            select = select.join(*join)

        for filt in self._rendered_filters:
            select = select.where(filt)

        return select

    @overload
    def to_df(
        self: RecSet[Any, Any, Any, Any, Any, tuple],
        kind: type[DfT],
    ) -> tuple[DfT, ...]: ...

    @overload
    def to_df(
        self: RecSet[Any, Any, Any, Any, Any, Record | None], kind: type[DfT]
    ) -> DfT: ...

    @overload
    def to_df(
        self: RecSet[Any, Any, Any, Any, Any, tuple],
        kind: None = ...,
    ) -> tuple[pl.DataFrame, ...]: ...

    @overload
    def to_df(
        self: RecSet[Any, Any, Any, Any, Any, Record | None], kind: None = ...
    ) -> pl.DataFrame: ...

    @overload
    def to_df(
        self: RecSet[Any, Any, Any, Any, Any, Any], kind: None = ...
    ) -> pl.DataFrame | tuple[pl.DataFrame, ...]: ...

    def to_df(
        self: RecSet[Record | None, Any, Any, Any, Any, Any],
        kind: type[DfT] | None = None,
    ) -> DfT | tuple[DfT, ...]:
        """Download selection."""
        select = self.select()

        idx_cols = [col.name for col in self.idx]

        main_prefix = self.record_type._default_table_name() + "."
        main_cols = {
            col: col[len(main_prefix) :]
            for col in select.columns.keys()
            if col.startswith(main_prefix)
        }

        extra_cols = {
            rel: {
                col: col[len(rel.path_str) + 1 :]
                for col in select.columns.keys()
                if col.startswith(rel.path_str)
            }
            for rel in self.merges.rels
        }

        merged_df = None
        if kind is pd.DataFrame:
            with self.db.engine.connect() as con:
                merged_df = pd.read_sql(select, con)
                merged_df = merged_df.set_index(idx_cols)
        else:
            merged_df = pl.read_database(
                str(select.compile(self.db.engine)), self.db.engine
            )

        main_df, *extra_dfs = cast(
            tuple[DfT, ...],
            (
                merged_df[list(main_cols.keys())].rename(main_cols),
                *(
                    merged_df[list(cols.keys())].rename(cols)
                    for cols in extra_cols.values()
                ),
            ),
        )

        return main_df, *extra_dfs

    @overload
    def __iter__(
        self: RecSet[RecT2 | None, Any, Any, Any, Any, Record | None]
    ) -> Iterator[RecT2]: ...

    @overload
    def __iter__(
        self: RecSet[RecT2 | None, Any, Any, Any, Any, tuple[*RecTt]]
    ) -> Iterator[tuple[RecT2, *tuple[Record, ...]]]: ...

    def __iter__(  # noqa: D105
        self: RecSet[RecT2 | None, Any, Any, Any, Any, Any],
    ) -> Iterator[RecT2 | tuple[RecT2, *tuple[Record, ...]]]:
        dfs = self.to_df()
        if isinstance(dfs, pl.DataFrame):
            dfs = (dfs,)

        rec_types: tuple[type[Record], ...] = (self.record_type, *self.merges.types)
        idx_cols = [
            f"{rel.path_str}.{pk}"
            for rel in self.merges.rels
            for pk in rel.record_type._primary_keys
        ]

        for rt in rec_types:
            if rt not in self.db._instance_map:
                self.db._instance_map[rt] = {}

        def iterator() -> Generator[RecT2 | tuple[RecT2, *tuple[Record, ...]]]:
            for rows in zip(*(df.iter_rows(named=True) for df in dfs)):
                main_row = rows[0]
                idx = tuple(main_row[i] for i in idx_cols)
                idx = idx[0] if len(idx) == 1 else idx

                rec_list = []
                for rec_type, row in zip(rec_types[1:], rows[1:]):
                    new_rec = self.record_type._from_partial_dict(row)
                    rec = self.db._instance_map[rec_type].get(new_rec._index) or new_rec
                    rec._db = self.db

                    rec_list.append(rec)

                yield tuple(rec_list) if len(rec_list) > 1 else rec_list[0]

        return iterator()

    def __imatmul__(
        self: RecSet[RecT2 | None, MdxT, RW, Any, ResT2],
        other: RecSet[RecT2 | None, MdxT, Any, Any] | RecInput[RecT2, MdxT],
    ) -> RecSet[ResT2, MdxT, RW, Any, ResT2]:
        """Aligned assignment."""
        self._mutate(other, mode="update")
        return cast(RecSet[ResT2, MdxT, RW, Any, ResT2], self)

    def __iand__(
        self: RecSet[RecT2 | None, MdxT, RW, Any, RecT2],
        other: RecSet[RecT2 | None, MdxT, Any, Any] | RecInput[RecT2, MdxT],
    ) -> RecSet[ResT2, MdxT, RW, Any, ResT2]:
        """Replacing assignment."""
        self._mutate(other, mode="replace")
        return cast(RecSet[ResT2, MdxT, RW, Any, ResT2], self)

    def __ior__(
        self: RecSet[RecT2 | None, MdxT, RW, Any, ResT2],
        other: RecSet[RecT2 | None, MdxT, Any, Any] | RecInput[RecT2, MdxT],
    ) -> RecSet[ResT2, MdxT, RW, Any, ResT2]:
        """Upserting assignment."""
        self._mutate(other, mode="upsert")
        return cast(RecSet[ResT2, MdxT, RW, Any, ResT2], self)

    def __iadd__(
        self: RecSet[RecT2 | None, MdxT, RW, Any, ResT2],
        other: RecSet[RecT2 | None, MdxT, Any, Any] | RecInput[RecT2, MdxT],
    ) -> RecSet[ResT2, MdxT, RW, Any, ResT2]:
        """Inserting assignment."""
        self._mutate(other, mode="insert")
        return cast(RecSet[ResT2, MdxT, RW, Any, ResT2], self)

    def __isub__(
        self: RecSet[RecT2 | None, MdxT, RW, Any, ResT2],
        other: RecSet[RecT2 | None, MdxT, Any, Any] | Iterable[MdxT] | MdxT,
    ) -> RecSet[ResT2, MdxT, RW, Any, ResT2]:
        """Deletion."""
        raise NotImplementedError("Delete not supported yet.")

    @overload
    def __lshift__(
        self: RecSet[Record[KeyT2] | None, BaseIdx, Any, Any, RecT2 | None],
        other: RecSet[RecT2 | None, Any, Any, Any] | RecInput[RecT2, KeyT2],
    ) -> list[KeyT2]: ...

    @overload
    def __lshift__(
        self: RecSet[Any, Any, Any, Any, RecT2 | None],
        other: RecSet[RecT2 | None, KeyT, Any, Any] | RecInput[RecT2, KeyT],
    ) -> list[KeyT]: ...

    def __lshift__(
        self: RecSet[Any, Any, Any, Any, RecT2 | None],
        other: RecSet[RecT2 | None, Any, Any, Any] | RecInput[RecT2, Any],
    ) -> list:
        """Injection."""
        raise NotImplementedError("Inject not supported yet.")

    @overload
    def __rshift__(
        self: RecSet[Any, KeyT2, RW, Any, RecT2 | None], other: KeyT2 | Iterable[KeyT2]
    ) -> dict[KeyT, RecT2]: ...

    @overload
    def __rshift__(
        self: RecSet[Record[KeyT2], BaseIdx, RW, Any, RecT2 | None],
        other: KeyT2 | Iterable[KeyT2],
    ) -> dict[KeyT2, RecT2]: ...

    def __rshift__(
        self: RecSet[Any, Any, RW, Any, RecT2 | None],
        other: Hashable | Iterable[Hashable],
    ) -> dict[Any, RecT2]:
        """Extraction."""
        raise NotImplementedError("Extract not supported yet.")

    # 1. Type deletion
    @overload
    def __delitem__(
        self: RecSet[RecT2 | None, Any, RW, Any, Any], key: type[RecT2]
    ) -> None: ...

    # 2. List deletion
    @overload
    def __delitem__(
        self: RecSet[
            Record[KeyT2] | None,
            BaseIdx | Filt[BaseIdx] | KeyT3 | Filt[KeyT3],
            RW,
            Any,
            RecT2 | None,
        ],
        key: Iterable[KeyT2 | KeyT3],
    ) -> None: ...

    # 3. Index value deletion
    @overload
    def __delitem__(
        self: RecSet[
            Record[KeyT2] | None,
            BaseIdx | Filt[BaseIdx] | KeyT3 | Filt[KeyT3],
            RW,
            Any,
            RecT2 | None,
        ],
        key: KeyT2 | KeyT3,
    ) -> None: ...

    # 4. Slice deletion
    @overload
    def __delitem__(
        self: RecSet[Record | None, Any, RW, Any], key: slice | tuple[slice, ...]
    ) -> None: ...

    # 5. Expression filter deletion
    @overload
    def __delitem__(
        self: RecSet[Record | None, Any, RW, Any], key: sqla.ColumnElement[bool]
    ) -> None: ...

    # Implementation:

    def __delitem__(  # noqa: D105
        self: RecSet[Record | None, Any, RW, LocalDyn, Any, Any],
        key: (
            type[Record]
            | list[Hashable]
            | Hashable
            | slice
            | tuple[slice, ...]
            | sqla.ColumnElement[bool]
        ),
    ) -> None:
        if not isinstance(key, slice) or key != slice(None):
            del self[key][:]

        select = self.select()

        tables = {
            self.db._get_table(rec, writable=True)
            for rec in self.record_type._record_bases
        }

        statements = []

        for table in tables:
            # Prepare delete statement.
            if self.db.engine.dialect.name in (
                "postgres",
                "postgresql",
                "duckdb",
                "mysql",
                "mariadb",
            ):
                # Delete-from.
                statements.append(
                    table.delete().where(
                        reduce(
                            sqla.and_,
                            (
                                table.c[col.name] == select.c[col.name]
                                for col in table.primary_key.columns
                            ),
                        )
                    )
                )
            else:
                # Correlated update.
                raise NotImplementedError("Correlated update not supported yet.")

        # Execute delete statements.
        with self.db.engine.begin() as con:
            for statement in statements:
                con.execute(statement)

    @overload
    def extract(
        *,
        self: RecSet[Record | None, Any, Any, Any, Any],
        aggs: Mapping[RelSet, Agg] | None = ...,
        to_backend: None = ...,
        overlay_with_schema: bool = ...,
    ) -> DB[RW, BsT]: ...

    @overload
    def extract(
        *,
        self: RecSet[Record | None, Any, Any, Any, Any],
        aggs: Mapping[RelSet, Agg] | None = ...,
        to_backend: Backend[BtT],
        overlay_with_schema: bool = ...,
    ) -> DB[RW, BtT]: ...

    def extract(
        self: RecSet[
            Record | None,
            Any,
            Any,
            Any,
            Any,
        ],
        aggs: Mapping[RelSet, Agg] | None = None,
        to_backend: Backend[BtT] | None = None,
        overlay_with_schema: bool = False,
    ) -> DB[RW, BsT | BtT]:
        """Extract a new database instance from the current selection."""
        # Get all rec types in the schema.
        rec_types = {self.record_type, *self.record_type._rel_types}

        # Get the entire subdag from this selection.
        all_paths_rels = {
            r
            for rel in self.record_type._rels.values()
            for r in rel._get_subdag(rec_types)
        }

        # Extract rel paths, which contain an aggregated rel.
        aggs_per_type: dict[type[Record], list[tuple[RelSet, Agg]]] = {}
        if aggs is not None:
            for rel, agg in aggs.items():
                for path_rel in all_paths_rels:
                    if rel in path_rel._rel_path:
                        aggs_per_type[rel.parent_type] = [
                            *aggs_per_type.get(rel.parent_type, []),
                            (rel, agg),
                        ]
                        all_paths_rels.remove(path_rel)

        replacements: dict[type[Record], sqla.Select] = {}
        for rec in rec_types:
            # For each table, create a union of all results from the direct routes.
            selects = [
                self[rel].select()
                for rel in all_paths_rels
                if issubclass(rec, rel.record_type)
            ]
            replacements[rec] = sqla.union(*selects).select()

        aggregations: dict[type[Record], sqla.Select] = {}
        for rec, rec_aggs in aggs_per_type.items():
            selects = []
            for rel, agg in rec_aggs:
                src_select = self[rel].select()
                selects.append(
                    sqla.select(
                        *[
                            (
                                src_select.c[sa.name]
                                if isinstance(sa, Col)
                                else sqla_visitors.replacement_traverse(
                                    sa,
                                    {},
                                    replace=lambda element, **kw: (
                                        src_select.c[element.name]
                                        if isinstance(element, Col)
                                        else None
                                    ),
                                )
                            ).label(ta.name)
                            for ta, sa in agg.map.items()
                        ]
                    )
                )

            aggregations[rec] = sqla.union(*selects).select()

        # Create a new database overlay for the results.
        overlay_db = copy_and_override(
            self.db,
            DB,
            write_to_overlay=f"temp_{token_hex(10)}",
            overlay_with_schema=overlay_with_schema,
            schema=None,
            types=rec_types,
            _def_types={},
            _metadata=sqla.MetaData(),
            _instance_map={},
        )

        # Overlay the new tables onto the new database.
        for rec in replacements:
            overlay_db[rec] &= replacements[rec]

        for rec, agg_select in aggregations.items():
            overlay_db[rec] &= agg_select

        # Transfer table to the new database.
        if to_backend is not None:
            other_db = copy_and_override(
                overlay_db,
                DB,
                b_id=to_backend.b_id,
                url=to_backend.url,
                _def_types={},
                _metadata=sqla.MetaData(),
                _instance_map={},
            )

            for rec in set(replacements) | set(aggregations):
                other_db[rec] &= overlay_db[rec].to_df()

            overlay_db = other_db

        return overlay_db

    def __len__(self) -> int:
        """Return the number of records in the dataset."""
        with self.db.engine.connect() as conn:
            res = conn.execute(
                sqla.select(sqla.func.count()).select_from(self.select().subquery())
            ).scalar()
            assert isinstance(res, int)
            return res

    def __contains__(self: RecSet[Any, Any, Any, DynBackendID], key: Hashable) -> bool:
        """Check if a record is in the dataset."""
        return len(self[key]) > 0

    def __clause_element__(self) -> sqla.Subquery:
        """Return subquery for the current selection to be used inside SQL clauses."""
        return self.select().subquery()

    def _joins(self, _subtree: DataDict | None = None) -> list[Join]:
        """Extract join operations from the relation tree."""
        joins = []
        _subtree = _subtree or self.merges.dict

        for rel, next_subtree in _subtree.items():
            parent = (
                rel._parent_rel._sql_alias
                if isinstance(rel._parent_rel, RelSet)
                else self.root_table
            )

            temp_alias_map = {rec: rec._db._sql_alias for rec in rel.inter_joins.keys()}

            joins.extend(
                (
                    temp_alias_map[rec],
                    reduce(
                        sqla.or_,
                        (
                            reduce(
                                sqla.and_,
                                (
                                    parent.c[lk.name] == temp_alias_map[rec].c[rk.name]
                                    for lk, rk in join_on.items()
                                ),
                            )
                            for join_on in joins
                        ),
                    ),
                )
                for rec, joins in rel.inter_joins.items()
            )

            target_table = rel._sql_alias

            joins.append(
                (
                    target_table,
                    reduce(
                        sqla.or_,
                        (
                            reduce(
                                sqla.and_,
                                (
                                    temp_alias_map[lk.parent_type].c[lk.name]
                                    == target_table.c[rk.name]
                                    for lk, rk in join_on.items()
                                ),
                            )
                            for join_on in rel.target_joins
                        ),
                    ),
                )
            )

            joins.extend(self._joins(next_subtree))

        return joins

    def _get_subdag(
        self,
        backlink_records: set[type[Record]] | None = None,
        _traversed: set[RelSet[Record | None, Any, Any, Any, LocalStat]] | None = None,
    ) -> set[RelSet[Record | None, Any, Any, Any, LocalStat]]:
        """Find all paths to the target record type."""
        backlink_records = backlink_records or set()
        _traversed = _traversed or set()

        # Get relations of the target type as next relations
        next_rels = set(self.record_type._rels.values())

        for backlink_record in backlink_records:
            next_rels |= backlink_record._backrels_to_rels(self.record_type)

        # Filter out already traversed relations
        next_rels = {rel for rel in next_rels if rel not in _traversed}

        # Add next relations to traversed set
        _traversed |= next_rels

        next_rels = {rel.prefix(self) for rel in next_rels}

        # Return next relations + recurse
        return next_rels | {
            rel
            for next_rel in next_rels
            for rel in next_rel._get_subdag(backlink_records, _traversed)
        }

    def _gen_idx_match_expr(
        self,
        values: Sequence[slice | list[Hashable] | Hashable | sqla.ColumnElement],
    ) -> sqla.ColumnElement[bool] | None:
        """Generate SQL expression for matching index values."""
        if values == slice(None):
            return None

        exprs = [
            (
                idx.label(None).in_(val)
                if isinstance(val, list)
                else (
                    idx.label(None).between(val.start, val.stop)
                    if isinstance(val, slice)
                    else idx.label(None) == val
                )
            )
            for idx, val in zip(self.idx, values)
        ]

        if len(exprs) == 0:
            return None

        return reduce(sqla.and_, exprs)

    @cached_property
    def _has_list_index(self) -> bool:
        return (
            isinstance(self, RelSet)
            and self.path_idx is not None
            and len(self.path_idx) == 1
            and is_subtype(list(self.path_idx)[0].value_type, int)
        )

    def _df_to_table(
        self,
        df: pd.DataFrame | pl.DataFrame,
    ) -> sqla.TableClause:
        if isinstance(df, pd.DataFrame) and any(
            name is None for name in df.index.names
        ):
            idx_names = [vs.name for vs in self.idx]
            df.index.set_names(idx_names, inplace=True)

        value_table = sqla.table(
            f"{self.record_type._default_table_name()}_{token_hex(5)}"
        )

        if isinstance(df, pd.DataFrame):
            df.reset_index().to_sql(
                value_table.name,
                self.db.engine,
                if_exists="replace",
                index=False,
            )
        else:
            df.write_database(str(value_table), self.db.engine)

        return value_table

    @staticmethod
    def _idx_map(idx_names: list[str], idx: Any) -> dict[str, Hashable]:
        if len(idx_names) == 1:
            return {idx_names[0]: idx}

        assert isinstance(idx, tuple) and len(idx) == len(idx_names)
        return {idx_name: idx_val for idx_name, idx_val in zip(idx_names, idx)}

    def _records_to_df(self, records: dict[Any, Record]) -> pl.DataFrame:
        idx_names = [vs.name for vs in self.idx]

        col_data = [
            {
                **self._idx_map(idx_names, idx),
                **{p.name: getattr(rec, p.name) for p in type(rec)._cols.values()},
            }
            for idx, rec in records.items()
        ]

        # Transform attribute data into DataFrame.
        return pl.DataFrame(col_data)

    def _mutate(  # noqa: C901
        self: RecSet[RecT2 | None, Any, RW, DynBackendID, Any, Any],
        value: RecSet[RecT2 | None, Any, Any, Any, Any] | RecInput[RecT2, Any],
        mode: Literal["update", "upsert", "replace", "insert"] = "update",
    ) -> None:
        mutations: list[
            tuple[
                type[Record] | RecSet[RecT2 | None, Any, RW, DynBackendID, Any, Any],
                sqla.FromClause,
                Literal["update", "upsert", "replace", "insert"],
            ]
        ] = []

        records: dict[Any, Record] | None = None
        self_sel = self if isinstance(self, RelSet) else self.record_type

        match value:
            case sqla.Select():
                mutations.append((self_sel, value.subquery(), mode))
            case RecSet():
                if value.db != self.db:
                    remote_db = value if isinstance(value, DB) else value.extract()
                    for s in remote_db._def_types:
                        if remote_db.b_id == self.db.b_id:
                            mutations.append((s, remote_db._get_table(s), "upsert"))
                        else:
                            mutations.append(
                                (s, self._df_to_table(remote_db[s].to_df()), "upsert")
                            )

                mutations.append((self_sel, value.select().subquery(), mode))
            case pd.DataFrame() | pl.DataFrame():
                mutations.append((self_sel, self._df_to_table(value), mode))
            case Record():
                records = {value._index: value}
            case Mapping() if has_type(value, Mapping[Any, Record]):
                records = dict(value)
            case Iterable():
                records = {idx: rec for idx, rec in enumerate(value)}

        if records is not None:
            db_grouped = {
                db: dict(recs)
                for db, recs in groupby(
                    sorted(records.items(), key=lambda x: x[1]._db.b_id),
                    lambda x: None if not x[1]._connected else x[1]._db,
                )
            }

            local_records = {**db_grouped.get(self.db, {}), **db_grouped.get(None, {})}

            remote_records = {
                db: recs
                for db, recs in db_grouped.items()
                if db is not None and db != self.db
            }

            if local_records:
                # Get the column data.
                df_data = self._records_to_df(local_records)
                mutations.append((self_sel, self._df_to_table(df_data), mode))

            for db, recs in remote_records.items():
                rec_ids = [rec._index for rec in recs.values()]
                remote_set = db[self.record_type][rec_ids]

                remote_db = (
                    db
                    if all(rec._root for rec in recs.values())
                    else remote_set.extract()
                )
                for s in remote_db._def_types:
                    if remote_db.b_id == self.db.b_id:
                        mutations.append((s, remote_db._get_table(s), "upsert"))
                    else:
                        mutations.append(
                            (s, self._df_to_table(remote_db[s].to_df()), "upsert")
                        )

                mutations.append((self_sel, remote_set.select().subquery(), mode))

        for rec, value_table, sub_mode in mutations:
            # Get the statements to perform the mutation.
            self.db[rec]._mutate_table(value_table, sub_mode)

            # Drop the temporary table, if any.
            if isinstance(value_table, sqla.TableClause):
                with self.db.engine.begin() as con:
                    con.execute(sqla.drop(value_table))

        return

    def _mutate_table(  # noqa: C901
        self: RecSet[RecT2 | None, Any, RW, BsT, Any, Any],
        value_table: sqla.FromClause,
        mode: Literal["update", "upsert", "replace", "insert"] = "update",
    ) -> None:
        cols_by_table = {
            self.db._get_table(rec, writable=True): {
                a for a in self.record_type._cols.values() if a.parent_type is rec
            }
            for rec in self.record_type._record_bases
        }

        statements: list[sqla.Executable] = []

        if mode == "replace":
            # Delete all records in the current selection.
            select = self.select()

            for table in cols_by_table:
                # Prepare delete statement.
                if self.db.engine.dialect.name in (
                    "postgres",
                    "postgresql",
                    "duckdb",
                    "mysql",
                    "mariadb",
                ):
                    # Delete-from.
                    statements.append(
                        table.delete().where(
                            reduce(
                                sqla.and_,
                                (
                                    table.c[col.name] == select.c[col.name]
                                    for col in table.primary_key.columns
                                ),
                            )
                        )
                    )
                else:
                    # Correlated update.
                    raise NotImplementedError("Correlated update not supported yet.")

        if mode in ("replace", "upsert", "insert"):
            # Construct the insert statements.

            assert len(self.filters) == 0, "Can only upsert into unfiltered datasets."

            for table, cols in cols_by_table.items():
                # Do an insert-from-select operation, which updates on conflict:
                statement = table.insert().from_select(
                    [a.name for a in cols],
                    value_table,
                )

                if mode in ("replace", "upsert"):
                    if isinstance(statement, postgresql.Insert):
                        # For Postgres / DuckDB, use: https://docs.sqlalchemy.org/en/20/dialects/postgresql.html#updating-using-the-excluded-insert-values
                        statement = statement.on_conflict_do_update(
                            index_elements=[
                                col.name for col in table.primary_key.columns
                            ],
                            set_=dict(statement.excluded),
                        )
                    elif isinstance(statement, mysql.Insert):
                        # For MySQL / MariaDB, use: https://docs.sqlalchemy.org/en/20/dialects/mysql.html#insert-on-duplicate-key-update-upsert
                        statement = statement.prefix_with("INSERT INTO")
                        statement = statement.on_duplicate_key_update(
                            **statement.inserted
                        )
                    else:
                        # For others, use CTE: https://docs.sqlalchemy.org/en/20/core/selectable.html#sqlalchemy.sql.expression.Select.cte
                        raise NotImplementedError(
                            "Upsert not supported for this database dialect."
                        )

                statements.append(statement)
        else:
            # Construct the update statements.

            assert isinstance(self.root_table, sqla.Table), "Base must be a table."

            # Derive current select statement and join with value table, if exists.
            value_join_on = (
                sqla.text("TRUE")
                if self.single_key is not None
                else reduce(
                    sqla.and_,
                    (
                        self.root_table.c[idx_col.name] == idx_col
                        for idx_col in value_table.primary_key
                    ),
                )
            )
            select = self.select().join(
                value_table,
                value_join_on,
            )

            for table, cols in cols_by_table.items():
                col_names = {a.name for a in cols}
                values = {
                    c_name: c
                    for c_name, c in value_table.columns.items()
                    if c_name in col_names
                }

                # Prepare update statement.
                if self.db.engine.dialect.name in (
                    "postgres",
                    "postgresql",
                    "duckdb",
                    "mysql",
                    "mariadb",
                ):
                    # Update-from.
                    statements.append(
                        table.update()
                        .values(values)
                        .where(
                            reduce(
                                sqla.and_,
                                (
                                    table.c[col.name] == select.c[col.name]
                                    for col in table.primary_key.columns
                                ),
                            )
                        )
                    )
                else:
                    # Correlated update.
                    raise NotImplementedError("Correlated update not supported yet.")

        # Execute delete / insert / update statements.
        with self.db.engine.begin() as con:
            for statement in statements:
                con.execute(statement)

        if mode in ("replace", "upsert") and isinstance(self, RelSet):
            # Update incoming relations from parent records.
            if self.direct_rel is not True:
                if issubclass(self.fk_record_type, self.parent_type):
                    # Case: parent links directly to child (n -> 1)
                    for fk, pk in self.direct_rel.fk_map.items():
                        self.parent[fk] @= value_table.select().with_only_columns(
                            value_table.c[pk.name]
                        )
                else:
                    # Case: parent and child are linked via assoc table (n <--> m)
                    # Update link table with new child indexes.
                    assert issubclass(self.link_type, Record)
                    ls = self.links
                    ls &= value_table.select()

            # Note that the (1 <- n) case is already covered by updating
            # the child record directly, which includes all its foreign keys.

        return

    def __setitem__(
        self,
        key: Any,
        other: RecSet[Any, Any, Any, Any, Any, Any] | Col[Any, Any, Any, Any, Any],
    ) -> None:
        """Catchall setitem."""
        return


@dataclass(kw_only=True, eq=False)
class Col(  # type: ignore[reportIncompatibleVariableOverride]
    Attr[VarTi, RwT, ParT],
    sqla.ColumnClause[VarTi],
    Generic[VarTi, RwT, CdxT, BsT, ParT],
):
    """Reference an attribute of a record."""

    def __post_init__(self) -> None:  # noqa: D105
        # Initialize fields required by SQLAlchemy superclass.
        self.table = None
        self.is_literal = False

    _record_set: RecSet[ParT | None, Any, RwT, BsT] | None = None

    index: bool = False
    primary_key: bool = False
    sql_getter: Callable[[sqla.FromClause], sqla.ColumnElement] | None = None

    @property
    def record_set(self) -> RecSet[ParT | None, Any, RwT, BsT]:
        """Return the record set of the column."""
        return cast(
            RecSet[ParT | None, Any, RwT, BsT],
            self._record_set if self._record_set is not None else self.parent_type._db,
        )

    @cached_property
    def sql_type(self) -> sqla_types.TypeEngine:
        """Column key."""
        return sqla_types.to_instance(self.value_type)  # type: ignore[reportCallIssue]

    def all(self) -> sqla.CollectionAggregate[bool]:
        """Return a SQL ALL expression for this attribute."""
        return sqla.all_(self)

    def any(self) -> sqla.CollectionAggregate[bool]:
        """Return a SQL ANY expression for this attribute."""
        return sqla.any_(self)

    def __hash__(self) -> int:
        """Hash the Col."""
        return gen_int_hash((Prop.__hash__(self), self.record_set))

    @overload
    def __get__(
        self, instance: None, owner: type[RecT2]
    ) -> Col[VarTi, RwT, CdxT, LocalStat, RecT2]: ...

    @overload
    def __get__(
        self: Prop[Any, Any, Any], instance: RecT2, owner: type[RecT2]
    ) -> VarTi: ...

    @overload
    def __get__(self, instance: object | None, owner: type | None) -> Self: ...

    def __get__(  # noqa: D105
        self, instance: object | None, owner: type | type[RecT2] | None
    ) -> Col[Any, Any, Any, Any, Any] | VarTi | Self:
        return super().__get__(instance, owner)

    # Plural selection
    @overload
    def __getitem__(
        self: Col[Any, Any, KeyT],
        key: Iterable[KeyT] | slice | tuple[slice, ...],
    ) -> Col[VarTi, RwT, KeyT, BsT, ParT]: ...

    # Single value selection
    @overload
    def __getitem__(
        self: Col[Any, Any, KeyT], key: KeyT
    ) -> Col[VarTi, RwT, SingleIdx, BsT, ParT]: ...

    # Implementation:

    def __getitem__(  # noqa: D105
        self: Col[Any, Any, KeyT],
        key: list[Hashable] | slice | tuple[slice, ...] | Hashable,
    ) -> Col[Any, Any, Any, Any, Any]:
        return copy_and_override(
            self, Col, _record_set=self.record_set[cast(slice, key)]
        )

    def __setitem__(
        self,
        key: Any,
        other: Col[Any, Any, Any, Any, Any],
    ) -> None:
        """Catchall setitem."""
        return

    def select(self) -> sqla.Select:
        """Return select statement for this column."""
        selection_table = self.record_set.select().subquery()

        return sqla.select(
            *(selection_table.c[a.name] for a in self.record_set.idx),
            selection_table.c[self.name],
        ).select_from(selection_table)

    @overload
    def to_series(self: Col, kind: type[pd.Series]) -> pd.Series: ...

    @overload
    def to_series(self: Col, kind: type[pl.Series] = ...) -> pl.Series: ...

    def to_series(
        self,
        kind: type[pd.Series | pl.Series] = pl.Series,
    ) -> pd.Series | pl.Series:
        """Download selection."""
        select = self.select()
        engine = self.record_set.db.engine

        if kind is pd.Series:
            with engine.connect() as con:
                return pd.read_sql(select, con).set_index(
                    [c.key for c in self._idx_cols]
                )[self.name]

        return pl.read_database(str(select.compile(engine)), engine)[self.name]

    def __imatmul__(
        self: Col[Any, RW, Any, Any, Any],
        value: ValInput | Col[VarT, Any, KeyT, BsT],
    ) -> Col[VarT, RW, KeyT, BsT, ParT]:
        """Do item-wise, broadcasting assignment on this value set."""
        match value:
            case pd.Series() | pl.Series():
                series = value
            case Mapping():
                series = pd.Series(dict(value))
            case Iterable():
                series = pd.Series(dict(enumerate(value)))
            case _:
                series = pd.Series({self.single_key or 0: value})

        df = series.rename(self.name).to_frame()

        self.record_set._mutate(df)
        return self


@dataclass(kw_only=True, eq=False)
class RelSet(
    RecSet[ResT, IdxT, RwT, BsT, ResTd, SelT],
    Prop[ResT | Iterable[ResT], RwT, ParT],
    Generic[ResT, LnT, IdxT, RwT, BsT, ParT, ResTd, SelT],
):
    """Relation set class."""

    index: bool = False
    primary_key: bool = False

    on: DirectLink[Record] | BackLink[Record] | BiLink[Record, Any] | None = None
    order_by: Mapping[Col[Any, Any, Any, LocalStat], int] | None = None
    map_by: Col[Any, Any, Any, LocalStat] | None = None

    def __hash__(self) -> int:
        """Hash the RelSet."""
        return gen_int_hash((RecSet.__hash__(self), Prop.__hash__(self)))

    @property
    def record_type(  # type: ignore[reportIncompatibleMethodOverride]
        self: RelSet[RecT2 | None, Any, Any, Any, Any, Any, Any, Any]
    ) -> type[RecT2]:
        """Record type of the set."""
        return cast(type[RecT2], self._type.record_type())

    @overload
    def __get__(
        self: RelSet[Any, Any, Any, Any, LocalStat, Any, Any],
        instance: None,
        owner: type[ParT2],
    ) -> RelSet[ResT, LnT, IdxT, RwT, LocalStat, ParT2, ResT, SelT]: ...

    @overload
    def __get__(
        self: RelSet[RecT2 | None, Any, SingleIdx, Any, LocalStat, Any, Any],
        instance: ParT2,
        owner: type[ParT2],
    ) -> RecT2 | None: ...

    @overload
    def __get__(
        self: RelSet[RecT2, Any, SingleIdx, Any, LocalStat, Any, Any],
        instance: ParT2,
        owner: type[ParT2],
    ) -> RecT2: ...

    @overload
    def __get__(
        self: RelSet[Any, Any, SingleIdx, Any, LocalStat, ParT, Any],
        instance: ParT2,
        owner: type[ParT2],
    ) -> RelSet[ResT, LnT, IdxT, RwT, BsT, ParT, ResT, Record]: ...

    @overload
    def __get__(
        self: RelSet[Any, Any, Any, Any, LocalStat, Any, Any],
        instance: object | None,
        owner: type,
    ) -> RelSet[ResT, LnT, IdxT, RwT, LocalStat, ParT, ResT, SelT]: ...

    def __get__(  # noqa: D105
        self: RelSet[Any, Any, Any, Any, LocalStat, Any, Any],
        instance: object | None,
        owner: type | type[RecT2],
    ) -> RelSet[Any, Any, Any, Any, Any, Any, Any, Any] | Record | None:
        # TODO: Implement support for default values of relations.
        # TODO: Implement getters + setters for relations.

        if has_type(instance, Record[Hashable]):
            if self.direct_rel is True:
                single_self = cast(
                    RelSet[Record, Any, SingleIdx, Any, LocalStat, Any, Any], self
                )
                return instance._db[type(instance)][single_self].get(instance._index)

            return copy_and_override(
                instance._db[type(instance)][instance._index][self],
                RelSet[ResT, LnT, IdxT, RwT, BsT, RecT2, ResTd, Record],
                sel_type=Record,
            )
        elif issubclass(owner, Record):
            return copy_and_override(
                self,
                RelSet[ResT, LnT, IdxT, RwT, BsT, RecT2, ResTd, Record],
                _parent_type=cast(type[RecT2], owner),
            )

        return self

    def __set__(  # noqa: D105
        self: RelSet[Record[KeyT2], Any, Any, RW, LocalStat, ParT2, RecT2],
        instance: ParT2,
        value: (
            RelSet[RecT2, LnT, IdxT, Any, Any, ParT2, RecT2] | RecInput[RecT2, KeyT2]
        ),
    ) -> None:
        rec = cast(Record[Hashable], instance)
        instance._db[type(rec)][rec._index][self]._mutate(value)

    @staticmethod
    def _get_tag(rec_type: type[Record]) -> RelSet | None:
        """Retrieve relation-tag of a record type, if any."""
        try:
            rel = getattr(rec_type, "_rel")
            return rel if isinstance(rel, RelSet) else None
        except AttributeError:
            return None

    @cached_property
    def _parent_rel(self) -> RelSet[ParT, Any, Any, RwT, BsT, Any, ParT] | None:
        """Parent relation of this Rel."""
        return cast(
            RelSet[ParT, Any, Any, RwT, BsT, Any, ParT],
            (
                self._get_tag(self.parent_type)
                if issubclass(self.parent_type, Record)
                else None
            ),
        )

    @cached_property
    def _rel_path(
        self,
    ) -> tuple[RecSet[Record | None, BaseIdx, RwT, BsT, Record], *tuple[RelSet, ...]]:
        """Path from base record type to this Rel."""
        if self._parent_rel is None:
            return (self.parent,)

        return cast(
            tuple[
                RecSet[Record | None, BaseIdx, RwT, BsT, Record], *tuple[RelSet, ...]
            ],
            (
                *self._parent_rel._rel_path,
                self,
            ),
        )

    @cached_property
    def root_set(self) -> RecSet[Record | None, BaseIdx, RwT, BsT, Record]:
        """Root record type of the set."""
        return self._rel_path[0]

    @cached_property
    def idx(
        self,
    ) -> Mapping[Col[Any, Any, Any, Any, Any], RecSet[Any, Any, Any, Any, Any, Any]]:
        """Return the index cols."""
        return self.path_idx if self.path_idx is not None else super().idx

    @cached_property
    def root_table(self) -> sqla.FromClause:
        """Get the main table for the current selection."""
        return self.root_set._sql_alias

    def to_static(
        self,
    ) -> RelSet[ResT, LnT, IdxT, RwT, LocalStat, ParT, ResTd]:
        """Return backend-less version of this RelSet."""
        tmpl = cast(
            RelSet[ResT, LnT, IdxT, RwT, LocalStat, ParT, ResTd],
            self,
        )
        return copy_and_override(
            tmpl,
            type(tmpl),
            db=DB(b_id=LocalBackend.static, types={self.record_type}),
            merges=RelTree(),
        )

    def prefix(
        self,
        left: type[ParT] | RecSet[ParT | None, Any, Any, Any, Any, Any],
    ) -> RelSet[ResT, LnT, IdxT, RwT, BsT, ParT, ResTd, SelT]:
        """Prefix this prop with a relation or record type."""
        current_root = self.root_set.record_type
        new_root = left if isinstance(left, RecSet) else left._rel(current_root)

        if new_root.res_type is NoneType:
            return self

        rel_path = self._rel_path[1:] if len(self._rel_path) > 1 else (self,)

        prefixed_rel = reduce(
            RelSet._append_rel,
            rel_path,
            new_root,
        )

        return cast(
            RelSet[ResT, LnT, IdxT, RwT, BsT, ParT, ResTd, SelT],
            prefixed_rel,
        )

    @cached_property
    def path_idx(
        self,
    ) -> Mapping[Col[Any, Any, Any, Any], RecSet[Any, Any, Any, Any]] | None:
        """Get the path index of the relation."""
        p = {
            col: rel
            for rel in self._rel_path[1:]
            for col in (
                [rel.map_by]
                if rel.map_by is not None
                else rel.order_by.keys() if rel.order_by is not None else []
            )
        }

        if len(p) == 0:
            return None

        return {
            **{
                c: self.root_set
                for c in self.root_set.record_type._primary_keys.values()
            },
            **p,
        }

    @cached_property
    def path_str(self) -> str:
        """String representation of the relation path."""
        prefix = (
            self.parent_type.__name__
            if len(self._rel_path) == 0
            else self._rel_path[-1].path_str
        )
        return f"{prefix}.{self.name}"

    @cached_property
    def parent(self) -> RecSet[ParT, Any, RwT, BsT, ParT]:
        """Parent set of this Rel."""
        if self._parent_rel is not None:
            tmpl = cast(RelSet[ParT, Any, Any, RwT, BsT, Any, ParT], self)
            return copy_and_override(
                tmpl,
                type(tmpl),
                _parent_type=self._parent_rel.parent_type,
                on=self._parent_rel.on,
                map_by=self._parent_rel.map_by,
                order_by=self._parent_rel.order_by,
            )

        tmpl = cast(RecSet[ParT, Any, RwT, BsT, ParT], self)
        return copy_and_override(tmpl, type(tmpl))

    @property
    def link_type(self) -> type[LnT]:
        """Get the link record type."""
        return cast(
            type[LnT],
            self._type.link_type(),
        )

    @property
    def links(
        self: RelSet[Any, RecT2, Any, Any, Any, Record, Any],
    ) -> RecSet[RecT2, KeyT, RwT, BsT, RecT2]:
        """Get the link set."""
        r = self.parent_type._rel(self.link_type)
        return cast(RecSet[RecT2, KeyT, RwT, BsT, RecT2], self.parent[r])

    @cached_property
    def ln(self: RelSet[Any, RecT2, Any, Any, Any, Any]) -> type[RecT2]:
        """Reference props of the link record type."""
        return self.links.rec if self.links is not None else cast(type[RecT2], Record)

    @cached_property
    def fk_record_type(self) -> type[Record]:
        """Record type of the foreign key."""
        match self.on:
            case type():
                return self.on
            case RelSet():
                return self.on.parent_type
            case tuple():
                link = self.on[0]
                assert isinstance(link, RecSet)
                return link.parent_type
            case None if issubclass(self.link_type, Record):
                return self.link_type
            case dict() | Col() | Iterable() | None:
                return self.parent_type

    @cached_property
    def direct_rel(
        self,
    ) -> RelSet[ParT, LnT, SingleIdx, RwT, BsT, Record] | Literal[True]:
        """Direct rel."""
        on = self.on
        if on is None and issubclass(self.link_type, Record):
            on = self.link_type

        match on:
            case type():
                rels = [
                    r
                    for r in on._rels.values()
                    if issubclass(self.parent_type, r.record_type)
                    and r.direct_rel is True
                ]
                assert len(rels) == 1, "Direct relation must be unique."
                return cast(
                    RelSet[ParT, LnT, SingleIdx, RwT, BsT, Record],
                    rels[0],
                )
            case RelSet():
                return cast(
                    RelSet[ParT, LnT, SingleIdx, RwT, BsT, Record],
                    on,
                )
            case tuple():
                link = on[0]
                assert isinstance(link, RelSet)
                return cast(RelSet[ParT, LnT, SingleIdx, RwT, BsT, Record], link)
            case dict() | Col() | Iterable() | None:
                return True

    @cached_property
    def counter_rel(
        self: RelSet[RecT2 | None, LnT, Any, RwT, BsT, ParT]
    ) -> RelSet[ParT, LnT, Any, RwT, BsT, RecT2]:
        """Counter rel."""
        if self.direct_rel is not True and issubclass(
            self.direct_rel.parent_type, self.record_type
        ):
            return cast(RelSet[ParT, LnT, Any, RwT, BsT, RecT2], self.direct_rel)

        return cast(
            RelSet[ParT, LnT, Any, RwT, BsT, RecT2],
            self.record_type._rel(self.parent_type),
        )

    @cached_property
    def fk_map(
        self,
    ) -> bidict[Col[Hashable], Col[Hashable]]:
        """Map source foreign keys to target cols."""
        target = self.record_type
        on = self.on
        if on is None and issubclass(self.link_type, Record):
            on = self.link_type

        match on:
            case type() | RelSet() | tuple():
                return bidict()
            case dict():
                return bidict(
                    {
                        Col[Hashable](
                            _name=fk.name,
                            _type=PropType(Col[cast(type[Hashable], fk.value_type)]),
                            _parent_type=self.record_type,
                        ): cast(Col[Hashable], pk)
                        for fk, pk in on.items()
                    }
                )
            case Col() | Iterable():
                cols = on if isinstance(on, Iterable) else [on]
                source_cols = [
                    Col[Hashable](
                        _name=col.name,
                        _type=PropType(Col[cast(type[Hashable], col.value_type)]),
                        _parent_type=self.record_type,
                    )
                    for col in cols
                ]
                target_cols = target._primary_keys.values()
                fk_map = dict(zip(source_cols, target_cols))

                assert all(
                    is_subtype(
                        self.parent_type._defined_cols[fk_col.name].value_type,
                        pk_col.value_type,
                    )
                    for fk_col, pk_col in fk_map.items()
                    if pk_col.typedef is not None
                ), "Foreign key value types must match primary key value types."

                return bidict(fk_map)
            case None:
                return bidict(
                    {
                        Col[Hashable](
                            _name=f"{self.name}_{target_col.name}",
                            _type=PropType(Col[target_col.value_type, Any, Record]),
                            _parent_type=self.record_type,
                        ): cast(Col[Hashable], target_col)
                        for target_col in target._primary_keys.values()
                    }
                )
            case _:
                return bidict()

    @cached_property
    def inter_joins(
        self,
    ) -> dict[
        type[Record],
        list[Mapping[Col[Hashable], Col[Hashable]]],
    ]:
        """Intermediate joins required by this rel."""
        on = self.on
        if on is None and issubclass(self.link_type, Record):
            on = self.link_type

        match on:
            case RelSet():
                # Relation is defined via other relation
                other_rel = on
                assert isinstance(
                    other_rel, RecSet
                ), "Back-reference must be an explicit relation"

                if issubclass(other_rel.record_type, self.parent_type):
                    # Supplied record type object is a backlinking relation
                    return {}
                else:
                    # Supplied record type object is a forward relation
                    # on a relation table
                    back_rels = [
                        rel
                        for rel in other_rel.parent_type._rels.values()
                        if issubclass(rel.record_type, self.parent_type)
                        and len(rel.fk_map) > 0
                    ]

                    return {
                        other_rel.parent_type: [
                            back_rel.fk_map.inverse for back_rel in back_rels
                        ]
                    }
            case type():
                if issubclass(on, self.record_type):
                    # Relation is defined via all direct backlinks of given record type.
                    return {}

                # Relation is defined via relation table
                back_rels = [
                    rel
                    for rel in on._rels.values()
                    if issubclass(rel.record_type, self.parent_type)
                    and len(rel.fk_map) > 0
                ]

                return {on: [back_rel.fk_map.inverse for back_rel in back_rels]}
            case tuple() if has_type(on, tuple[RelSet, RelSet]):
                # Relation is defined via back-rel + forward-rel
                # on a relation table.
                back, _ = on
                assert len(back.fk_map) > 0, "Back relation must be direct."
                assert issubclass(back.parent_type, Record)
                return {back.parent_type: [back.fk_map.inverse]}
            case _:
                # Relation is defined via foreign key attributes
                return {}

    @cached_property
    def target_joins(
        self,
    ) -> list[Mapping[Col[Hashable], Col[Hashable]]]:
        """Mappings of column keys to join the target on."""
        on = self.on
        if on is None and issubclass(self.link_type, Record):
            on = self.link_type

        match on:
            case RelSet():
                # Relation is defined via other relation or relation table
                other_rel = on
                assert (
                    len(other_rel.fk_map) > 0
                ), "Backref or forward-ref on relation table must be a direct relation"
                return [
                    (
                        other_rel.fk_map.inverse
                        if issubclass(other_rel.record_type, self.parent_type)
                        else other_rel.fk_map
                    )
                ]

            case type():
                if issubclass(on, self.record_type):
                    # Relation is defined via all direct backlinks of given record type.
                    back_rels = [
                        rel
                        for rel in on._rels.values()
                        if issubclass(rel.record_type, self.parent_type)
                        and len(rel.fk_map) > 0
                    ]
                    assert len(back_rels) > 0, "No direct backlinks found."
                    return [back_rel.fk_map.inverse for back_rel in back_rels]

                # Relation is defined via relation table
                fwd_rels = [
                    rel
                    for rel in on._rels.values()
                    if issubclass(rel.record_type, self.parent_type)
                    and len(rel.fk_map) > 0
                ]
                assert (
                    len(fwd_rels) > 0
                ), "No direct forward rels found on relation table."
                return [fwd_rel.fk_map for fwd_rel in fwd_rels]

            case tuple():
                assert has_type(on, tuple[RelSet, RelSet])
                # Relation is defined via back-rel + forward-rel
                # on a relation table.
                _, fwd = on
                assert len(fwd.fk_map) > 0, "Forward relation must be direct."
                return [fwd.fk_map]

            case _:
                # Relation is defined via foreign key attributes
                return [self.fk_map]


class LocalBackend(StrEnum):
    """Local backend."""

    static = "local-stat"
    dynamic = "local-dyn"


@dataclass(kw_only=True, eq=False)
class Backend(Generic[BkT]):
    """Data backend."""

    b_id: BkT = LocalBackend.dynamic
    """Unique name to identify this database's backend by."""

    url: sqla.URL | CloudPath | HttpFile | Path | None = None
    """Connection URL or path."""


@dataclass(kw_only=True, eq=False)
class DB(RecSet[None, BaseIdx, RwT, BkT, Record, None], Backend[BkT]):
    """Database."""

    db: DB[RwT, BkT] = field(init=False, repr=False)

    types: (
        Mapping[
            type[Record],
            Literal[True] | Require | str | sqla.TableClause,
        ]
        | set[type[Record]]
        | None
    ) = None
    schema: (
        type[Schema]
        | Mapping[
            type[Schema],
            Literal[True] | Require | str,
        ]
        | None
    ) = None

    write_to_overlay: str | None = None
    overlay_with_schema: bool = True

    validate_on_init: bool = False
    create_cross_fk: bool = False

    _def_types: Mapping[
        type[Record], Literal[True] | Require | str | sqla.TableClause
    ] = field(default_factory=dict)
    _subs: dict[type[Record], sqla.TableClause] = field(default_factory=dict)

    _metadata: sqla.MetaData = field(default_factory=sqla.MetaData)
    _instance_map: dict[type[Record], dict[Hashable, Record]] = field(
        default_factory=dict
    )

    def __post_init__(self):  # noqa: D105
        self.db = self

        if self.types is not None:
            types = self.types
            if isinstance(types, set):
                types = cast(
                    dict[type[Record], Literal[True]], {rec: True for rec in self.types}
                )

            self._subs = {
                **self._subs,
                **{
                    rec: (sub if isinstance(sub, sqla.TableClause) else sqla.table(sub))
                    for rec, sub in types.items()
                    if not isinstance(sub, Require) and sub is not True
                },
            }
            self._def_types = {**self._def_types, **types}

        if self.schema is not None:
            if isinstance(self.schema, Mapping):
                self._subs = {
                    **self._subs,
                    **{
                        rec: sqla.table(rec._default_table_name(), schema=schema_name)
                        for schema, schema_name in self.schema.items()
                        for rec in schema._record_types
                    },
                }
                schemas = self.schema
            else:
                schemas = {self.schema: True}

            self._def_types = cast(
                dict,
                {
                    **self._def_types,
                    **{
                        rec: (
                            req
                            if not isinstance(req, str)
                            else sqla.table(rec._default_table_name(), schema=req)
                        )
                        for schema, req in schemas.items()
                        for rec in schema._record_types
                    },
                },
            )

        self.res_type = get_lowest_common_base(self._def_types.keys())

        if self.validate_on_init:
            self.validate()

        if self.write_to_overlay is not None and self.overlay_with_schema:
            self._ensure_schema_exists(self.write_to_overlay)

    def __hash__(self) -> int:
        """Hash the DB."""
        return gen_int_hash(
            (
                self.b_id,
                self.url,
                self._subs,
            )
        )

    @cached_property
    def _schema_types(
        self,
    ) -> Mapping[type[Record], Literal[True] | Require | str | sqla.TableClause]:
        """Set of all schema types in this DB."""
        types = dict(self._def_types).copy()

        for rec in types:
            types = {
                **{r: True for r in rec._rel_types},
                **types,
            }

        return types

    @property
    def record_type(self: RecSet[Any, Any, Any, Any, Any, Any]) -> type[Record]:
        """Set of all record types in this DB."""
        return Record

    @cached_property
    def backend_type(
        self,
    ) -> Literal["sql-connection", "sqlite-file", "excel-file", "in-memory"]:
        """Type of the backend."""
        match self.url:
            case Path() | CloudPath():
                return "excel-file" if "xls" in self.url.suffix else "sqlite-file"
            case sqla.URL():
                return (
                    "sqlite-file"
                    if self.url.drivername == "sqlite"
                    else "sql-connection"
                )
            case HttpFile():
                url = yarl.URL(self.url.url)
                typ = (
                    ("excel-file" if "xls" in Path(url.path).suffix else "sqlite-file")
                    if url.scheme in ("http", "https")
                    else None
                )
                if typ is None:
                    raise ValueError(f"Unsupported URL scheme: {url.scheme}")
                return typ
            case None:
                return "in-memory"

    @cached_property
    def engine(self) -> sqla.engine.Engine:
        """SQLA Engine for this DB."""
        # Create engine based on backend type
        # For Excel-backends, use duckdb in-memory engine
        return (
            sqla.create_engine(
                self.url if isinstance(self.url, sqla.URL) else str(self.url)
            )
            if (
                self.backend_type == "sql-connection"
                or self.backend_type == "sqlite-file"
            )
            else (sqla.create_engine(f"duckdb:///:memory:{self.b_id}"))
        )

    @cached_property
    def assoc_types(self) -> set[type[Record]]:
        """Set of all association tables in this DB."""
        assoc_types = set()
        for rec in self._def_types:
            pks = set([col.name for col in rec._primary_keys.values()])
            fks = set(
                [col.name for rel in rec._rels.values() for col in rel.fk_map.keys()]
            )
            if pks == fks:
                assoc_types.add(rec)

        return assoc_types

    @cached_property
    def relation_map(self) -> dict[type[Record], set[RelSet]]:
        """Maps all tables in this DB to their outgoing or incoming relations."""
        rels: dict[type[Record], set[RelSet]] = {
            table: set() for table in self._def_types
        }

        for rec in self._def_types:
            for rel in rec._rels.values():
                rels[rec].add(rel)
                rels[rel.record_type].add(rel)

        return rels

    def describe(self) -> dict[str, str | dict[str, str] | None]:
        """Return a description of this database.

        Returns:
            Mapping of table names to table descriptions.
        """
        schema_desc = {}
        if isinstance(self.schema, type):
            schema_ref = PyObjectRef.reference(self.schema)

            schema_desc = {
                "schema": {
                    "repo": schema_ref.repo,
                    "package": schema_ref.package,
                    "class": f"{schema_ref.module}.{schema_ref.object}",
                }
            }

            if schema_ref.object_version is not None:
                schema_desc["schema"]["version"] = schema_ref.object_version
            elif schema_ref.package_version is not None:
                schema_desc["schema"]["version"] = schema_ref.package_version

            if schema_ref.repo_revision is not None:
                schema_desc["schema"]["revision"] = schema_ref.repo_revision

            if schema_ref.docs_url is not None:
                schema_desc["schema"]["docs"] = schema_ref.docs_url

        return {
            **schema_desc,
            "backend": (
                str(self.url)
                if self.url is not None and self.backend_type != "in-memory"
                else None
            ),
        }

    def to_graph(
        self, nodes: Sequence[type[Record]]
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """Export links between select database objects in a graph format.

        E.g. for usage with `Gephi`_

        .. _Gephi: https://gephi.org/
        """
        node_tables = [self[n] for n in nodes]

        # Concat all node tables into one.
        node_dfs = [
            n.to_df(kind=pd.DataFrame)
            .reset_index()
            .assign(table=n.record_type._default_table_name())
            for n in node_tables
        ]
        node_df = (
            pd.concat(node_dfs, ignore_index=True)
            .reset_index()
            .rename(columns={"index": "node_id"})
        )

        directed_edges = reduce(set.union, (self.relation_map[n] for n in nodes))

        undirected_edges: dict[type[Record], set[tuple[RelSet, ...]]] = {
            t: set() for t in nodes
        }
        for n in nodes:
            for at in self.assoc_types:
                if len(at._rels) == 2:
                    left, right = (r for r in at._rels.values())
                    assert left is not None and right is not None
                    if left.record_type == n:
                        undirected_edges[n].add((left, right))
                    elif right.record_type == n:
                        undirected_edges[n].add((right, left))

        # Concat all edges into one table.
        edge_df = pd.concat(
            [
                *[
                    node_df.loc[
                        node_df["table"]
                        == str((rel.parent_type or Record)._default_table_name())
                    ]
                    .rename(columns={"node_id": "source"})
                    .merge(
                        node_df.loc[
                            node_df["table"]
                            == str(rel.record_type._default_table_name())
                        ],
                        left_on=[c.name for c in rel.fk_map.keys()],
                        right_on=[c.name for c in rel.fk_map.values()],
                    )
                    .rename(columns={"node_id": "target"})[["source", "target"]]
                    .assign(ltr=",".join(c.name for c in rel.fk_map.keys()), rtl=None)
                    for rel in directed_edges
                ],
                *[
                    self[assoc_table]
                    .to_df(kind=pd.DataFrame)
                    .merge(
                        node_df.loc[
                            node_df["table"]
                            == str(left_rel.record_type._default_table_name())
                        ].dropna(axis="columns", how="all"),
                        left_on=[c.name for c in left_rel.fk_map.keys()],
                        right_on=[c.name for c in left_rel.fk_map.values()],
                        how="inner",
                    )
                    .rename(columns={"node_id": "source"})
                    .merge(
                        node_df.loc[
                            node_df["table"]
                            == str(left_rel.record_type._default_table_name())
                        ].dropna(axis="columns", how="all"),
                        left_on=[c.name for c in right_rel.fk_map.keys()],
                        right_on=[c.name for c in right_rel.fk_map.values()],
                        how="inner",
                    )
                    .rename(columns={"node_id": "target"})[
                        list(
                            {
                                "source",
                                "target",
                                *(a for a in (left_rel.parent_type or Record)._cols),
                            }
                        )
                    ]
                    .assign(
                        ltr=",".join(c.name for c in right_rel.fk_map.keys()),
                        rtl=",".join(c.name for c in left_rel.fk_map.keys()),
                    )
                    for assoc_table, rels in undirected_edges.items()
                    for left_rel, right_rel in rels
                ],
            ],
            ignore_index=True,
        )

        return node_df, edge_df

    def validate(self) -> None:
        """Perform pre-defined schema validations."""
        types = {}

        if self.types is not None:
            types |= {
                rec: (isinstance(req, Require) and req.present) or req is True
                for rec, req in self._def_types.items()
            }

        if isinstance(self.schema, Mapping):
            types |= {
                rec: isinstance(req, Require) and req.present
                for schema, req in self.schema.items()
                for rec in schema._record_types
            }

        tables = {self._get_table(rec): required for rec, required in types.items()}

        inspector = sqla.inspect(self.engine)

        # Iterate over all tables and perform validations for each
        for table, required in tables.items():
            has_table = inspector.has_table(table.name, table.schema)

            if not has_table and not required:
                continue

            # Check if table exists
            assert has_table

            db_columns = {
                c["name"]: c for c in inspector.get_columns(table.name, table.schema)
            }
            for column in table.columns:
                # Check if column exists
                assert column.name in db_columns

                db_col = db_columns[column.name]

                # Check if column type and nullability match
                assert isinstance(db_col["type"], type(column.type))
                assert db_col["nullable"] == column.nullable or column.nullable is None

            # Check if primary key is compatible
            db_pk = inspector.get_pk_constraint(table.name, table.schema)
            if len(db_pk["constrained_columns"]) > 0:  # Allow source tbales without pk
                assert set(db_pk["constrained_columns"]) == set(
                    table.primary_key.columns.keys()
                )

            # Check if foreign keys are compatible
            db_fks = inspector.get_foreign_keys(table.name, table.schema)
            for fk in table.foreign_key_constraints:
                matches = [
                    (
                        set(db_fk["constrained_columns"]) == set(fk.column_keys),
                        (
                            db_fk["referred_table"].lower()
                            == fk.referred_table.name.lower()
                        ),
                        set(db_fk["referred_columns"])
                        == set(f.column.name for f in fk.elements),
                    )
                    for db_fk in db_fks
                ]

                assert any(all(m) for m in matches)

    def execute[
        *T
    ](
        self,
        stmt: sqla.Select[tuple[*T]] | sqla.Insert | sqla.Update | sqla.Delete,
    ) -> sqla.Result[tuple[*T]]:
        """Execute a SQL statement in this database's context."""
        stmt = self._parse_expr(stmt)
        with self.engine.begin() as conn:
            return conn.execute(self._parse_expr(stmt))

    def to_set(
        self: DB[RW, Any],
        data: pd.DataFrame | pl.DataFrame | sqla.Select,
        fks: Mapping[str, Col] | None = None,
    ) -> RecSet[DynRecord, Any, RO, BsT]:
        """Create a temporary dataset instance from a DataFrame or SQL query."""
        name = (
            f"temp_df_{gen_str_hash(data, 10)}"
            if isinstance(data, pd.DataFrame | pl.DataFrame)
            else f"temp_{token_hex(5)}"
        )

        rec = dynamic_record_type(name, props=props_from_data(data, fks))
        ds = RecSet[DynRecord, BaseIdx, RW, BsT, DynRecord](res_type=rec, db=self.db)

        ds &= data

        return ds

    def _get_table(
        self, rec: type[Record] | None = None, writable: bool = False
    ) -> sqla.Table:
        rec_type = rec or self.record_type

        if writable and self.write_to_overlay is not None and rec not in self._subs:
            # Create an empty overlay table for the record type
            self._subs[rec_type] = sqla.table(
                (
                    (self.write_to_overlay + "_" + rec_type._default_table_name())
                    if not self.overlay_with_schema
                    else rec_type._default_table_name()
                ),
                schema=self.write_to_overlay if self.overlay_with_schema else None,
            )

        table_name = rec_type._sql_table_name(self._subs)

        if table_name in self._metadata.tables:
            return self._metadata.tables[table_name]

        new_metadata = sqla.MetaData()
        for t in self._metadata.tables.values():
            t.to_metadata(new_metadata)

        table = rec_type._table(new_metadata, self._subs)

        # Create any missing tables in the database.
        if self.b_id is not None:
            new_tables = set(new_metadata.tables) - set(self._metadata.tables)
            if len(new_tables) > 0:
                new_metadata.create_all(self.engine, checkfirst=True)

        return table

    def _get_joined_table(
        self, rec: type[Record] | None = None
    ) -> sqla.Table | sqla.Join:
        rec_type = rec or self.record_type

        new_metadata = sqla.MetaData()
        for t in self._metadata.tables.values():
            t.to_metadata(new_metadata)

        table = rec_type._joined_table(new_metadata, self._subs)

        # Create any missing tables in the database.
        if self.b_id is not None:
            new_tables = set(new_metadata.tables) - set(self._metadata.tables)
            if len(new_tables) > 0:
                new_metadata.create_all(self.engine, checkfirst=True)

        return table

    def _parse_schema_items(
        self,
        element: sqla_visitors.ExternallyTraversible,
        **kw: Any,
    ) -> sqla.ColumnElement | sqla.FromClause | None:
        if isinstance(element, RelSet):
            return element._sql_alias
        elif isinstance(element, Col):
            return element.record_set._sql_alias.c[element.name]
        elif has_type(element, type[Record]):
            return self._get_table(element)

        return None

    def _parse_expr[CE: sqla.ClauseElement](self, expr: CE) -> CE:
        """Parse an expression in this database's context."""
        return cast(
            CE,
            sqla_visitors.replacement_traverse(
                expr, {}, replace=self._parse_schema_items
            ),
        )

    def _ensure_schema_exists(self, schema_name: str) -> str:
        """Ensure that the table exists in the database, then return it."""
        if not sqla.inspect(self.engine).has_schema(schema_name):
            with self.engine.begin() as conn:
                conn.execute(sqla.schema.CreateSchema(schema_name))

        return schema_name

    def _table_exists(self, sqla_table: sqla.Table) -> bool:
        """Check if a table exists in the database."""
        return sqla.inspect(self.engine).has_table(
            sqla_table.name, schema=sqla_table.schema
        )

    def _create_sqla_table(self, sqla_table: sqla.Table) -> None:
        """Create SQL-side table from Table class."""
        if not self.create_cross_fk:
            # Create a temporary copy of the table object and remove external FKs.
            # That way, local metadata will retain info on the FKs
            # (for automatic joins) but the FKs won't be created in the DB.
            sqla_table = sqla_table.to_metadata(sqla.MetaData())  # temporary metadata
            _remove_external_fk(sqla_table)

        sqla_table.create(self.engine)

    def _load_from_excel(self, record_types: list[type[Record]] | None = None) -> None:
        """Load all tables from Excel."""
        assert self.b_id is not None
        assert self.backend_type == "excel-file", "Backend must be an Excel file."
        assert isinstance(self.url, Path | CloudPath | HttpFile)

        path = self.url.get() if isinstance(self.url, HttpFile) else self.url

        with open(path, "rb") as file:
            for rec in record_types or self._def_types:
                pl.read_excel(
                    file, sheet_name=rec._default_table_name()
                ).write_database(
                    str(self._get_table(rec, writable=True)), str(self.engine.url)
                )

    def _save_to_excel(
        self, record_types: Iterable[type[Record]] | None = None
    ) -> None:
        """Save all (or selected) tables to Excel."""
        assert self.b_id is not None
        assert self.backend_type == "excel-file", "Backend must be an Excel file."
        assert isinstance(self.url, Path | CloudPath | HttpFile)

        file = BytesIO() if isinstance(self.url, HttpFile) else self.url.open("wb")

        with ExcelWorkbook(file) as wb:
            for rec in record_types or self._def_types:
                pl.read_database(
                    f"SELECT * FROM {self._get_table(rec)}",
                    self.engine,
                ).write_excel(wb, worksheet=rec._default_table_name())

        if isinstance(self.url, HttpFile):
            assert isinstance(file, BytesIO)
            self.url.set(file)

    def _delete_from_excel(self, record_types: Iterable[type[Record]]) -> None:
        """Delete selected table from Excel."""
        assert self.b_id is not None
        assert self.backend_type == "excel-file", "Backend must be an Excel file."
        assert isinstance(self.url, Path | CloudPath | HttpFile)

        file = BytesIO() if isinstance(self.url, HttpFile) else self.url.open("wb")

        wb = openpyxl.load_workbook(file)
        for rec in record_types or self._def_types:
            del wb[rec._default_table_name()]

        if isinstance(self.url, HttpFile):
            assert isinstance(file, BytesIO)
            self.url.set(file)


class Rel(RelSet[ResT, None, SingleIdx, RwT, BsT, ParT, ResTd, SelT]):
    """Singular relation."""


class RecUUID(Record[UUID]):
    """Record type with a default UUID primary key."""

    _template = True
    _id: Col[UUID] = prop(primary_key=True, default_factory=uuid4)


class RecHashed(Record[int]):
    """Record type with a default hashed primary key."""

    _template = True

    _id: Col[int] = prop(primary_key=True, init=False)

    def __post_init__(self) -> None:  # noqa: D105
        self._id = gen_int_hash(
            {a.name: getattr(self, a.name) for a in self._defined_cols.values()}
        )


class Scalar(Record[KeyT], Generic[VarT2, KeyT]):
    """Dynamically defined record type."""

    _template = True

    _id: Col[KeyT] = prop(primary_key=True, default_factory=uuid4)
    _value: Col[VarT2]


class DynRecordMeta(RecordMeta):
    """Metaclass for dynamically defined record types."""

    def __getitem__(cls: type[Record], name: str) -> Col[Any, Any, Record]:
        """Get dynamic attribute by dynamic name."""
        return Col(_name=name, _type=PropType(Col[cls]), _parent_type=cls)

    def __getattr__(cls: type[Record], name: str) -> Col[Any, Any, Record]:
        """Get dynamic attribute by name."""
        if not TYPE_CHECKING and name.startswith("__"):
            return super().__getattribute__(name)

        return Col(_name=name, _type=PropType(Col[cls]), _parent_type=cls)


class DynRecord(Record, metaclass=DynRecordMeta):
    """Dynamically defined record type."""

    _template = True


a = DynRecord


def dynamic_record_type(
    name: str, props: Iterable[Prop[Any, Any]] = []
) -> type[DynRecord]:
    """Create a dynamically defined record type."""
    return type(name, (DynRecord,), {p.name: p for p in props})


class Link(RecHashed, Generic[RecT2, RecT3]):
    """Automatically defined relation record type."""

    _template = True

    _from: RelSet[RecT2]
    _to: RelSet[RecT3]


class Schema:
    """Group multiple record types into a schema."""

    _record_types: set[type[Record]]
    _rel_record_types: set[type[Record]]

    def __init_subclass__(cls) -> None:  # noqa: D105
        subclasses = get_subclasses(cls, max_level=1)
        cls._record_types = {s for s in subclasses if isinstance(s, Record)}
        cls._rel_record_types = {rr for r in cls._record_types for rr in r._rel_types}
        super().__init_subclass__()


@dataclass
class Require:
    """Mark schema or record type as required."""

    present: bool = True


type DataDict = dict[RelSet, DataDict]
